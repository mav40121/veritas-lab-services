"""
Rebuild server/capRequirements.ts as CAP-orphan-policies only.

Source: master citation index Accreditor_Orphans tab (CAP rows) intersected
with MAS Policy/Procedure flag (column 2 == "X"). Every entry traces to:
  1. A real CAP ID in current MAS source (verified by ID-existence check)
  2. CAP-confirmed "needs a written policy" via the Policy/Procedure flag
  3. Master-document-confirmed "no CFR equivalent" via the Orphans tab

Per A5 citation rule (master document README):
  - Cite ID + topic noun phrase in our words.
  - NEVER quote requirement text verbatim.
  - Topic summaries are in our own words.

Output shape preserved from existing server/capRequirements.ts:
  { id, chapter, chapter_label, standard, name, description, service_line,
    source: "cap" }
"""
import json
import re
import glob
import os
import sys
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

MASTER = r'C:/Users/veril/OneDrive/Desktop/Lab/Regulatory/aaa Truth Master Document/veritaassure_master_citation_index (12).xlsx'
CAP_DIR = r'C:/Users/veril/OneDrive/Desktop/Lab/Regulatory/2026 Cap checklists'
REPO_FILE = r'C:\Users\veril\projects\veritas-lab-services\server\capRequirements.ts'

ID_PATTERN = re.compile(r'^[A-Z]{2,4}\.\d{5}$')

MODULE_LABELS = {
    'ANP': 'Anatomic Pathology',
    'CHM': 'Chemistry',
    'COM': 'All Common Checklist',
    'CYP': 'Cytopathology',
    'DRA': 'Director Review',
    'GEN': 'Laboratory General',
    'HEM': 'Hematology',
    'IMM': 'Immunology',
    'MIC': 'Microbiology',
    'POC': 'Point of Care Testing',
    'TRM': 'Transfusion Medicine',
    'URN': 'Urinalysis',
}

# Service-line per module (matches the existing repo's convention).
# Most CAP requirements are "all"; specialty ones get service_line set.
SERVICE_LINE = {
    'ANP': 'pathology',
    'CYP': 'pathology',
    'TRM': 'blood_bank',
}

# ─── 1. Build Subject Header lookup from MAS + collect policy-flagged set ─
subject_by_id = {}
policy_flagged = set()
for fpath in sorted(glob.glob(os.path.join(CAP_DIR, 'MAS_*.xlsx'))):
    module = os.path.basename(fpath).split('_')[1]
    wb = load_workbook(fpath, data_only=True, read_only=True)
    ws = wb.active
    for r in ws.iter_rows(min_row=7, values_only=True):
        v = r[0]
        if not v or not ID_PATTERN.match(str(v).strip()):
            continue
        cap_id = str(v).strip()
        pp_flag = r[1] if len(r) > 1 else None
        subject_header = r[3] if len(r) > 3 else None
        if subject_header:
            subject_by_id[cap_id] = str(subject_header).strip()
        if pp_flag and str(pp_flag).strip().upper() == 'X':
            policy_flagged.add(cap_id)

print(f'MAS source: {len(subject_by_id)} IDs with subject headers, {len(policy_flagged)} policy-flagged', file=sys.stderr)

# ─── 2. Load CAP orphans from master document ────────────────────────────
wb = load_workbook(MASTER, data_only=True, read_only=True)
rows = list(wb['Accreditor_Orphans'].iter_rows(values_only=True))
hdr = rows[0]
idx = {h: i for i, h in enumerate(hdr) if h}

cap_orphans = []
for r in rows[1:]:
    if not r or r[idx['accreditor']] != 'CAP':
        continue
    citation = r[idx['accreditor_citation']]
    m = re.match(r'^([A-Z]{2,4}\.\d{5})', str(citation or ''))
    if not m:
        continue
    cap_id = m.group(1)
    if cap_id not in policy_flagged:
        continue  # skip non-policy orphans
    cap_orphans.append({
        'cap_id': cap_id,
        'topic': r[idx['topic_in_our_words']],
        'subpart': r[idx['category_clia_subpart']],
        'service_line': r[idx['service_line']],
    })

print(f'CAP orphans that are policy-flagged: {len(cap_orphans)}', file=sys.stderr)

# ─── 3. Compose entries ─────────────────────────────────────────────────
def escape(s):
    if s is None: return ''
    return str(s).replace('\\', '\\\\').replace('"', '\\"').replace('\n', ' ').replace('—', '-').replace('–', '-')

# Sort by chapter (alphabetically), then by CAP ID ascending
cap_orphans.sort(key=lambda o: (o['cap_id'].split('.')[0], o['cap_id']))

entries = []
next_id = 1001
for o in cap_orphans:
    cap_id = o['cap_id']
    chapter = cap_id.split('.')[0]
    chapter_label = MODULE_LABELS.get(chapter, chapter)
    subject = subject_by_id.get(cap_id) or str(o['topic'] or '').strip() or chapter_label
    # Use the SUBJECT HEADER from MAS as the policy `name`.
    # Description: per A5 rule, ID + topic noun phrase in our words, NOT
    # verbatim requirement text. Wrap with module context.
    topic = str(o['topic'] or subject).strip()
    description = (
        f"CAP {cap_id} ({subject}) is a CAP-required written policy or procedure "
        f"under the {chapter_label} checklist with no direct 42 CFR equivalent. "
        f"The laboratory documents a policy addressing {topic.lower() if topic else subject.lower()}."
    )
    service = SERVICE_LINE.get(chapter, 'all')
    entries.append({
        'id': next_id,
        'chapter': chapter,
        'chapter_label': chapter_label,
        'standard': cap_id,
        'name': subject,
        'description': description,
        'service_line': service,
        'source': 'cap',
    })
    next_id += 1

# ─── 4. Emit TS file ────────────────────────────────────────────────────
out_lines = [
    '// Auto-generated from MAS 12/09/2025 + master citation index -- DO NOT EDIT MANUALLY\n',
    '// CAP ORPHAN POLICIES ONLY. Each entry is:\n',
    '//   (1) A real CAP ID in current MAS source (12 module xlsx files,\n',
    '//       edition date 2025-12-09).\n',
    '//   (2) CAP-confirmed as needing a written policy or procedure (column 2\n',
    '//       Policy/Procedure flag set to "X" in MAS).\n',
    '//   (3) Master-document-confirmed as having NO direct 42 CFR equivalent\n',
    '//       (i.e., an Accreditor_Orphans entry, not a CFR_Crosswalk member).\n',
    '// CAP requirements that DO map to a CFR section are cross-referenced on\n',
    '// the matching row in server/cfrRequirements.ts via the cap_ids field.\n',
    '// Built 2026-05-11 from veritaassure_master_citation_index (12).xlsx v0.9.\n',
    '// Rebuild script: script/rebuild-cap-orphans-policies-only.py\n',
    '// A5 citation rule: ID + topic noun phrase only; no verbatim CAP text.\n',
    '\n',
    'export const CAP_REQUIREMENTS = [\n',
]
for e in entries:
    out_lines.append(
        f'  {{"id": {e["id"]}, "chapter": "{e["chapter"]}", '
        f'"chapter_label": "{escape(e["chapter_label"])}", '
        f'"standard": "{e["standard"]}", '
        f'"name": "{escape(e["name"])}", '
        f'"description": "{escape(e["description"])}", '
        f'"service_line": "{e["service_line"]}", '
        f'"source": "{e["source"]}"}},\n'
    )
out_lines.append('];\n')
out_lines.append('\n')
out_lines.append('export type CAPRequirement = typeof CAP_REQUIREMENTS[number];\n')

with open(REPO_FILE, 'w', encoding='utf-8', newline='\n') as f:
    f.writelines(out_lines)

print()
print(f'Written {len(entries)} CAP-orphan-policy entries to {REPO_FILE}')

# Module distribution summary
from collections import Counter
print()
print('Module distribution:')
mods = Counter(e['chapter'] for e in entries)
for m in sorted(mods):
    print(f'  {m}: {mods[m]:3d}  ({MODULE_LABELS.get(m, m)})')
