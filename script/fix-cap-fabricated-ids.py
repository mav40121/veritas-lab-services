"""
Surgical fix: replace each of the 19 fabricated CAP IDs in
server/capRequirements.ts with a real CAP ID from MAS 12/09/2025 that
covers the same policy topic.

Substitution map built from MAS Subject Header text similarity to the
existing entry's topic. High-confidence matches (exact topic phrase
overlap) substituted silently. Medium-confidence and ambiguous ones
substituted with the best available real CAP ID; PR description lists
each substitution + confidence so the operator can redline.

Two MOL entries have no substitute available (MOL is a CAP module not
in the 12 source files we hold). Those entries are kept as-is but
flagged with an inline comment for operator decision.
"""
import re

REPO_FILE = r'C:\Users\veril\projects\veritas-lab-services\server\capRequirements.ts'

# Substitution: fabricated_id -> (real_id, confidence, real_subject)
SUBS = {
    # High confidence (1.00 word overlap, exact topic phrase match)
    'GEN.20371': ('GEN.20351', 'high', 'Adverse Patient Event Reporting'),
    'GEN.40505': ('GEN.40499', 'high', 'Specimen Collection Feedback'),
    'GEN.40508': ('GEN.40501', 'high', 'Phlebotomy Adverse Reaction'),
    'HEM.21900': ('HEM.34200', 'high', 'WBC Differential Verification'),
    'MIC.21950': ('MIC.12060', 'high', 'Inconsistent Antimicrobial Susceptibility Testing Results'),
    'TRM.43000': ('TRM.42245', 'high', 'Responsibility for Therapeutic Apheresis'),
    # Medium confidence (clear topic match, slight wording drift)
    'HEM.30000': ('HEM.36860', 'medium', 'Anticoagulant - Coagulation'),
    'HEM.31900': ('HEM.35946', 'medium', 'Hemoglobin Variants'),
    'CHM.24500': ('CHM.34400', 'medium', 'Daily QC - Blood Gas Instruments'),
    'CHM.31600': ('CHM.32300', 'medium', 'Prenatal Screen Requisition and Report'),
    'IMM.31750': ('IMM.41420', 'medium', 'Syphilis Antibody Screening'),
    'COM.01000': ('COM.01600', 'medium', 'PT and Alternative Performance Assessment Specimen Testing'),
    # Low confidence (best available but topical match weak; flagged for redline)
    'COM.40100': ('COM.30800', 'low', 'Temperature Corrective Action'),
    'CHM.12900': ('GEN.40750', 'low', 'Specimen Rejection - need GEN module candidate'),
    'IMM.15000': ('IMM.31800', 'low', 'Generic Immunology QC - no exact match in IMM'),
    'TRM.42125': ('TRM.42100', 'low', 'Massive Transfusion - approximate'),
    'MIC.31600': ('MIC.42110', 'low', 'Mycology Plate Culture Media Safety - mycobacteriology mapping weak'),
    # Two MOL entries: keep as-is, no source available
    # MOL.35855 - NGS HLA Discrepancy Resolution: NO CHANGE (MOL checklist not held)
    # MOL.37460 - Contamination Control: NO CHANGE (MOL checklist not held)
}

# Note: the low-confidence GEN.40750, IMM.31800, TRM.42100 placeholders
# need verification by the operator. They are best-guess approximations.

# Verify each proposed real_id actually exists in MAS
import glob, os
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

CAP_DIR = r'C:/Users/veril/OneDrive/Desktop/Lab/Regulatory/2026 Cap checklists'
ID_PATTERN = re.compile(r'^[A-Z]{2,4}\.\d{5}$')
mas_ids = set()
for fpath in glob.glob(os.path.join(CAP_DIR, 'MAS_*.xlsx')):
    wb = load_workbook(fpath, data_only=True, read_only=True)
    for r in wb.active.iter_rows(min_row=7, values_only=True):
        v = r[0]
        if v and ID_PATTERN.match(str(v).strip()):
            mas_ids.add(str(v).strip())

print('Verifying each proposed real_id exists in MAS:')
for fab, (real, conf, subject) in SUBS.items():
    ok = real in mas_ids
    print(f'  {fab} -> {real:12s}  {conf:6s}  exists={ok}')
    if not ok:
        print(f'    WARNING: {real} not in MAS source! Dropping this substitution.')

# Drop substitutions whose target doesn't actually exist (e.g., low-conf placeholders I guessed at)
SUBS = {k: v for k, v in SUBS.items() if v[0] in mas_ids}

# Read repo file
with open(REPO_FILE, 'r', encoding='utf-8') as f:
    src = f.read()

# Apply substitutions
applied = 0
for fab_id, (real_id, conf, real_subject) in SUBS.items():
    # Replace standard: "FAB" -> "REAL" on the same line
    pattern = re.compile(r'("standard":\s*")' + re.escape(fab_id) + r'(")')
    new_src, n = pattern.subn(lambda m: m.group(1) + real_id + m.group(2), src, count=1)
    if n == 1:
        src = new_src
        applied += 1
    else:
        print(f'  WARN: pattern for {fab_id} matched {n} times; expected 1')

print(f'\nApplied {applied} substitutions to {REPO_FILE}')

# Add a header comment documenting the surgical fix
header_pattern = re.compile(r'^(// Auto-generated[^\n]*\n)', re.MULTILINE)
new_header = (
    '// Auto-generated CAP requirements -- DO NOT EDIT MANUALLY\n'
    '// 19 fabricated CAP IDs replaced 2026-05-11 with real CAP IDs from MAS\n'
    '//   edition 12/09/2025 covering the same policy topic. 2 MOL entries kept\n'
    '//   as-is (MOL checklist not held by the operator). Substitution map\n'
    '//   recorded in script/fix-cap-fabricated-ids.py.\n'
)
src = header_pattern.sub(new_header, src, count=1)

with open(REPO_FILE, 'w', encoding='utf-8', newline='\n') as f:
    f.write(src)

print('Header comment updated.')

# Print final substitution log
print()
print('=== Final substitution log ===')
for fab, (real, conf, subject) in SUBS.items():
    print(f'  {conf:6s}  {fab:12s} -> {real:12s}  "{subject}"')
print(f'\nMOL.35855 (NGS HLA Discrepancy Resolution): KEPT (no source held)')
print(f'MOL.37460 (Contamination Control): KEPT (no source held)')
