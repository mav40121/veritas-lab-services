"""
Augment server/cfrRequirements.ts entries with accreditor cross-reference
arrays (cap_ids, tjc_ids, cola_ids, aabb_ids) drawn from the master citation
index's CFR_Crosswalk + CAP_Uncompressed tabs.

Architecture: CFR is the spine. Each CFR row carries the matching CAP/TJC/
COLA/AABB IDs as cross-references (not as separate rows). Accreditor files
(capRequirements.ts etc.) will then be reducible to orphans-only without
losing surveyor coverage.
"""
import re
import sys
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

MASTER = r'C:/Users/veril/OneDrive/Desktop/Lab/Regulatory/aaa Truth Master Document/veritaassure_master_citation_index (12).xlsx'
REPO_FILE = r'C:\Users\veril\projects\veritas-lab-services\server\cfrRequirements.ts'

# ─── 1. Load master document ────────────────────────────────────────────────
wb = load_workbook(MASTER, data_only=True, read_only=True)

# CFR_Crosswalk shape (from prior inspection):
#   col 1 = CFR Citation, col 2 = Section Title, col 3 = Topic, col 4 = Subpart,
#   col 5 = Service Line, col 6 = Federal Linkage,
#   col 7 = CAP IDs, col 8 = TJC IDs, col 9 = COLA IDs, col 10 = AABB IDs
# Note: openpyxl gives None for the first column ("row_id" doesn't exist in
# CFR_Crosswalk; the first cell is empty). Header row indicates real columns
# start at index 1.

cross_rows = list(wb['CFR_Crosswalk'].iter_rows(values_only=True))
header = cross_rows[0]
print(f'CFR_Crosswalk header: {header}', file=sys.stderr)

# Map header name -> column index
hdr_idx = {h: i for i, h in enumerate(header) if h}

# CAP_Uncompressed: maps a CFR row_id to its uncompressed CAP ID list.
# Used to expand range notation like "CHM.13000-CHM.13125 (3 IDs)".
uncompressed_rows = list(wb['CAP_Uncompressed'].iter_rows(values_only=True))
print(f'CAP_Uncompressed header: {uncompressed_rows[0]}', file=sys.stderr)
# Header: row_id | CFR Row ID | CFR Citation | Section Title | Compressed CAP IDs | Uncompressed CAP IDs | Runs Compressed JSON
uncompressed_by_citation = {}
for r in uncompressed_rows[1:]:
    if not r or not r[2]:
        continue
    citation = str(r[2]).strip()
    uncompressed_cell = r[5] if len(r) > 5 else None
    if uncompressed_cell:
        ids = [x.strip() for x in str(uncompressed_cell).split('\n') if x.strip()]
        # Strip "(topic)" suffix from each ID
        clean_ids = []
        for idstr in ids:
            m = re.match(r'^([A-Z]{2,4}\.\d{4,5})', idstr)
            if m:
                clean_ids.append(m.group(1))
        uncompressed_by_citation[citation] = clean_ids

print(f'CAP_Uncompressed rows: {len(uncompressed_by_citation)}', file=sys.stderr)

# ─── 2. Parse each crosswalk row's accreditor cells ─────────────────────────
def parse_cap_cell(cell, citation):
    """CAP cells may contain range notation. Expand via CAP_Uncompressed
    when this citation has an entry there; otherwise extract IDs literally."""
    if not cell:
        return []
    # Prefer uncompressed list when available
    if citation in uncompressed_by_citation:
        return uncompressed_by_citation[citation]
    out = []
    for raw in str(cell).split('\n'):
        raw = raw.strip()
        if not raw:
            continue
        # Match "GEN.13750" or "GEN.13750 (topic)" or "GEN.13750-GEN.13900 (3 IDs)"
        # For range notation without uncompressed counterpart, just emit endpoints
        m_range = re.match(r'^([A-Z]{2,4}\.\d{4,5})-([A-Z]{2,4}\.\d{4,5})', raw)
        if m_range:
            out.append(m_range.group(1))
            out.append(m_range.group(2))
            continue
        m_single = re.match(r'^([A-Z]{2,4}\.\d{4,5})', raw)
        if m_single:
            out.append(m_single.group(1))
    return out

def parse_tjc_cell(cell):
    """TJC cells: 'QSA.04.07.01 (topic)' format."""
    if not cell:
        return []
    out = []
    for raw in str(cell).split('\n'):
        raw = raw.strip()
        if not raw:
            continue
        m = re.match(r'^([A-Z]{2,5}\.\d{2}\.\d{2}\.\d{2})', raw)
        if m:
            out.append(m.group(1))
    return out

def parse_cola_cell(cell):
    """COLA cells: 'PER 3 (topic)' format. Strip the trailing dot+E/R if present."""
    if not cell:
        return []
    out = []
    for raw in str(cell).split('\n'):
        raw = raw.strip()
        if not raw:
            continue
        # Match "PER 3" or "PER 3*E" or "COLA APM 13" etc.
        m = re.match(r'^(?:COLA\s+)?([A-Z]{1,5})\s+(\d+(?:\.\d+)?)', raw)
        if m:
            out.append(f'{m.group(1)} {m.group(2)}')
    return out

def parse_aabb_cell(cell):
    """AABB cells: '5.14.1 (topic)' format."""
    if not cell:
        return []
    out = []
    for raw in str(cell).split('\n'):
        raw = raw.strip()
        if not raw:
            continue
        # Comma-separated within a line is possible: "PER 3 (...), PER 4 (...)"
        for part in re.split(r',\s*(?=[\dA-Z])', raw):
            m = re.match(r'^(\d+(?:\.\d+){0,3})', part.strip())
            if m:
                out.append(m.group(1))
    return out

# Build citation -> xrefs map
xref_by_citation = {}
for r in cross_rows[1:]:
    if not r or not r[hdr_idx['CFR Citation']]:
        continue
    citation = str(r[hdr_idx['CFR Citation']]).strip()
    xref_by_citation[citation] = {
        'cap_ids':  parse_cap_cell(r[hdr_idx['CAP IDs']],  citation),
        'tjc_ids':  parse_tjc_cell(r[hdr_idx['TJC IDs']]),
        'cola_ids': parse_cola_cell(r[hdr_idx['COLA IDs']]),
        'aabb_ids': parse_aabb_cell(r[hdr_idx['AABB IDs']]),
    }

print(f'CFR_Crosswalk rows parsed: {len(xref_by_citation)}', file=sys.stderr)
# Sample
sample = list(xref_by_citation.items())[5]
print(f'  Sample: {sample[0]!r}', file=sys.stderr)
print(f'    cap_ids: {sample[1]["cap_ids"][:5]}{"..." if len(sample[1]["cap_ids"])>5 else ""} ({len(sample[1]["cap_ids"])} total)', file=sys.stderr)
print(f'    tjc_ids: {sample[1]["tjc_ids"]}', file=sys.stderr)
print(f'    cola_ids: {sample[1]["cola_ids"]}', file=sys.stderr)
print(f'    aabb_ids: {sample[1]["aabb_ids"]}', file=sys.stderr)

# ─── 3. Read repo file ──────────────────────────────────────────────────────
with open(REPO_FILE, 'r', encoding='utf-8') as f:
    src = f.read()

# Normalize citation for matching: strip §, normalize whitespace
def normalize_citation(s):
    s = s.replace('§', '').replace('§', '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# Build normalized lookup
xref_normalized = {normalize_citation(k): v for k, v in xref_by_citation.items()}

# Match each entry line. Same regex as the rebuild script.
entry_pat = re.compile(
    r'(\{"id":\s*\d+,\s*"chapter":\s*"[^"]*",\s*"chapter_label":\s*"[^"]*",\s*'
    r'"standard":\s*")([^"]+)(",\s*"name":\s*"[^"]*",\s*"description":\s*"(?:[^"\\]|\\.)*",\s*'
    r'"service_line":\s*"[^"]*",\s*"source":\s*"[^"]*")(\})'
)

def emit_arr(ids):
    if not ids:
        return '[]'
    return '[' + ', '.join(f'"{i}"' for i in ids) + ']'

matched = 0
unmatched = []
unmatched_citations = set()

def augment(m):
    global matched
    head, std, mid, tail = m.group(1), m.group(2), m.group(3), m.group(4)
    norm = normalize_citation(std)
    xref = xref_normalized.get(norm)
    if xref is None:
        # Try with paragraph-qualifier stripped
        norm_base = re.sub(r'\([a-z\d]+\).*$', '', norm).strip()
        xref = xref_normalized.get(norm_base)
    if xref is None:
        unmatched_citations.add(std)
        cap_ids = tjc_ids = cola_ids = aabb_ids = []
    else:
        matched += 1
        cap_ids = xref['cap_ids']
        tjc_ids = xref['tjc_ids']
        cola_ids = xref['cola_ids']
        aabb_ids = xref['aabb_ids']
    addition = (
        f', "cap_ids": {emit_arr(cap_ids)}'
        f', "tjc_ids": {emit_arr(tjc_ids)}'
        f', "cola_ids": {emit_arr(cola_ids)}'
        f', "aabb_ids": {emit_arr(aabb_ids)}'
    )
    return head + std + mid + addition + tail

new_src, n_subs = entry_pat.subn(augment, src)

# Update the file header comments to reflect the augmentation
header_comment = (
    '// Auto-generated from eCFR + master citation index -- DO NOT EDIT MANUALLY\n'
    '// 42 CFR Part 493 portion rebuilt 2026-05-11 against official eCFR XML\n'
    '//   https://www.ecfr.gov/api/versioner/v1/full/2026-05-07/title-42.xml?part=493\n'
    '// Accreditor cross-references (cap_ids, tjc_ids, cola_ids, aabb_ids) added\n'
    '// 2026-05-11 from the master citation index CFR_Crosswalk + CAP_Uncompressed\n'
    '// tabs (veritaassure_master_citation_index (12).xlsx). CFR is the spine;\n'
    '// the accreditor arrays are cross-references on each CFR row, not separate\n'
    '// rows. Empty arrays mean the master document has no accreditor mapping for\n'
    '// that CFR citation (CFR sections that are administrative-only or\n'
    '// out-of-scope for the supported accreditors).\n'
    '// 21 CFR / 29 CFR / 45 CFR / 42 CFR 482-485 entries preserved from prior\n'
    '// build pending similar source-grounded rebuild against their respective\n'
    '// eCFR endpoints. Source: 42 CFR Part 493 (public domain U.S. government work),\n'
    '// 21 CFR Parts 606/610/640, 29 CFR 1910, 45 CFR 164.\n'
)
new_src = re.sub(
    r'^// Auto-generated[\s\S]*?(?=export const CFR_REQUIREMENTS)',
    header_comment + '\n',
    new_src,
    count=1,
)

with open(REPO_FILE, 'w', encoding='utf-8', newline='\n') as f:
    f.write(new_src)

print()
print(f'Substitutions made: {n_subs}')
print(f'Matched to master document: {matched}')
print(f'Unmatched (out-of-scope CFR titles or no mapping): {n_subs - matched}')
print(f'Distinct unmatched citations (top 10):')
for c in sorted(unmatched_citations)[:10]:
    print(f'  {c!r}')

# Quick spot-check: a §493.1256 row should have many cross-refs
print()
print('Spot-check: §493.1256 cross-refs in master document:')
for k, v in xref_by_citation.items():
    if '493.1256' in k:
        print(f'  {k!r}')
        print(f'    cap_ids: {v["cap_ids"][:8]}{"..." if len(v["cap_ids"])>8 else ""}')
        print(f'    tjc_ids: {v["tjc_ids"]}')
        print(f'    cola_ids: {v["cola_ids"]}')
        print(f'    aabb_ids: {v["aabb_ids"]}')
        break
