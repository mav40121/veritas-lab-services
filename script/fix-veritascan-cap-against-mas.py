"""
Manual CAP mapping for veritaScanData.ts — same discipline as the TJC
rebuild in PR #104. Each scan item gets a hand-picked CAP ID from MAS
12/09/2025 based on reading the question and choosing the best-matching
Subject Header in the relevant CAP module.

Format: scan_item_id -> "CAP.ID  # Subject Header from MAS (rationale)"
"""
import re, json, sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO = r'C:\Users\veril\projects\veritas-lab-services'
SCAN_FILE = os.path.join(REPO, 'client/src/lib/veritaScanData.ts')

# id -> (cap_id, subject_header_for_log)
# Picks reflect the scan item's intent matched against MAS Subject Headers.
# Comments encode the choice rationale.

MAPPING = {
    # ─── Quality Systems & QC ───────────────────────────────────────────────
    2:   ('GEN.30000', 'Monitoring Analytic Performance'),       # QC at required frequency
    3:   ('GEN.30000', 'Monitoring Analytic Performance'),       # L-J charts maintained/reviewed (cross-cutting QC monitoring)
    4:   ('GEN.20318', 'Corrective and Preventive Action'),      # Corrective action when QC out
    5:   ('COM.50500', 'Quality Control Plan Elements'),         # QC acceptability ranges established
    6:   ('DRA.10460', 'Director Responsibility - PT/QC'),       # Director reviews QC records
    7:   ('COM.30450', 'New Reagent Lot and Shipment Confirmation of Acceptability'),  # New QC lot validated
    8:   ('GEN.30000', 'Monitoring Analytic Performance'),       # Westgard multi-rule (lives under analytic performance monitoring)
    9:   ('GEN.20318', 'Corrective and Preventive Action'),      # QC failure -> patient result review
    10:  ('DRA.10460', 'Director Responsibility - PT/QC'),       # End-of-month QC review by supervisor/director
    11:  ('GEN.20316', 'QMS Indicators of Quality'),             # PI program with objectives + outcomes
    12:  ('COM.30000', 'Critical Result Notification'),          # Critical values policy
    13:  ('COM.30100', 'Critical Result Read-Back'),             # Critical value read-back
    16:  ('COM.06300', 'Specimen Rejection Criteria'),           # Specimen rejection criteria
    18:  ('GEN.20326', 'Assessment of the QMS Implementation'),  # QM plan reviewed/updated annually
    20:  ('GEN.20318', 'Corrective and Preventive Action'),      # PT failures -> RCA

    # ─── Calibration & Verification ─────────────────────────────────────────
    21:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),  # may not exist; verify below
    22:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),
    23:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),
    24:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),
    26:  ('COM.04250', 'Comparability of Instruments and Methods - Nonwaived Testing'),  # Method comparison
    27:  ('COM.04250', 'Comparability of Instruments and Methods - Nonwaived Testing'),
    28:  ('COM.04300', 'Comparability Criteria - Nonwaived Testing'),
    30:  ('COM.40300', 'Verification of Test Performance Specifications - FDA-cleared/approved Tests'),  # EP15 precision
    31:  ('COM.40500', 'Analytical Interferences'),              # was: Reportable range - need a better fit
    32:  ('COM.40300', 'Verification of Test Performance Specifications - FDA-cleared/approved Tests'),  # Accuracy verification
    33:  ('COM.30980', 'Waived Test Implementation and Approval'),  # Waived excluded from cal ver
    34:  ('GEN.20377', 'Record and Material Retention - General Laboratory'),  # 2y/10y retention
    35:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),  # Sign-off date used
    36:  ('COM.40300', 'Verification of Test Performance Specifications - FDA-cleared/approved Tests'),  # Factory-locked exempt
    37:  ('COM.40300', 'Verification of Test Performance Specifications - FDA-cleared/approved Tests'),  # Calver materials true values
    38:  ('COM.40000', 'Calibration Verification - Nonwaived Tests'),  # Cal ver schedule

    # ─── Proficiency Testing ────────────────────────────────────────────────
    39:  ('COM.01300', 'PT Participation'),                       # Enrolled in PT
    40:  ('COM.01600', 'PT and Alternative Performance Assessment Specimen Testing'),  # PT samples handled same
    41:  ('COM.01300', 'PT Participation'),                       # PT submitted on time
    42:  ('COM.01400', 'PT Attestation Statement'),               # PT results reviewed by director
    43:  ('COM.01700', 'PT and Alternative Performance Assessment Result Evaluation'),  # Unsuccessful PT -> RCA
    44:  ('GEN.20377', 'Record and Material Retention - General Laboratory'),  # PT records 2y retention
    45:  ('COM.01800', 'PT Interlaboratory Communication'),        # No communication about PT samples
    46:  ('COM.01500', 'Alternative Performance Assessment'),      # APA for unregulated analytes
    47:  ('COM.01300', 'PT Participation'),                        # PT enrollment certs on file
    48:  ('COM.01700', 'PT and Alternative Performance Assessment Result Evaluation'),  # Corrective action per PT result
    49:  ('COM.04250', 'Comparability of Instruments and Methods - Nonwaived Testing'),  # PT comparison between instruments
    50:  ('COM.01600', 'PT and Alternative Performance Assessment Specimen Testing'),    # PT personnel = patient testing
    51:  ('COM.01700', 'PT and Alternative Performance Assessment Result Evaluation'),   # PT trend analysis
    52:  ('COM.01600', 'PT and Alternative Performance Assessment Specimen Testing'),    # PT same manner/time as patient
    53:  ('COM.01600', 'PT and Alternative Performance Assessment Specimen Testing'),    # PT material handling

    # ─── Personnel & Competency ─────────────────────────────────────────────
    # Note: most personnel topics live in GEN.55000-56000 range
    55:  ('DRA.10100', 'Laboratory Director Qualifications'),
    56:  ('DRA.10200', 'Section Director/Technical Supervisor Qualifications'),
    57:  ('GEN.55450', 'Personnel Training'),       # Testing personnel qualifications
    59:  ('GEN.55600', 'Competency Assessment Frequency - Nonwaived Testing'),  # Annual competency
    60:  ('GEN.55500', 'Competency Assessment Elements - Nonwaived Testing'),   # 6 CLIA methods
    61:  ('DRA.10460', 'Director Responsibility - PT/QC'),         # Director signs competency
    62:  ('GEN.55700', 'Competency Corrective Action'),            # Remedial training
    63:  ('COM.04100', 'Supervisory Review for High Complexity Testing'),  # Infrequent test concurrent QC
    64:  ('DRA.10460', 'Director Responsibility - PT/QC'),         # Director attestation on staff quals
    66:  ('GEN.55450', 'Personnel Training'),                       # Safety/regulatory training records
    67:  ('DRA.11425', 'Director Responsibility - Documenting Delegation of Specific Functions'),  # Delegation
    68:  ('GEN.55500', 'Competency Assessment Elements - Nonwaived Testing'),  # Knowledge of QC rules
    69:  ('GEN.55450', 'Personnel Training'),                       # Training for new tests
    70:  ('GEN.54400', 'Personnel Records'),                        # Staff roster + testing privileges
    71:  ('GEN.54400', 'Personnel Records'),                        # State licensure
    72:  ('GEN.55750', 'Continuing Education'),                     # CE records (may not exist; verify)
    73:  ('DRA.10432', 'Director On-Site Visits - Laboratories Subject to US Regulations'),  # Director physical presence

    # ─── Test Management & Procedures ───────────────────────────────────────
    74:  ('COM.10000', 'Policy and Procedure Manual'),             # Written SOPs every test
    75:  ('COM.10100', 'Policy and Procedure Review'),             # SOP review every 2y
    76:  ('COM.10300', 'Knowledge of Policies and Procedures'),    # SOPs accessible at point of use
    77:  ('COM.10500', 'Discontinued Policies and Procedures'),    # Discontinued SOPs retained 2y
    78:  ('GEN.20361', 'CLIA Certificate Type'),                   # CLIA complexity matches
    81:  ('GEN.20377', 'Record and Material Retention - General Laboratory'),  # Patient results retention
    82:  ('GEN.40491', 'Primary Specimen Container Labeling'),     # Specimen labeling minimum IDs
    83:  ('COM.06000', 'Specimen Collection Manual'),              # Specimen collection procedures
    84:  ('COM.30000', 'Critical Result Notification'),            # Panic value notification
    85:  ('COM.40640', 'Clinical Performance Characteristics Validation - Laboratory-developed Tests'),  # LDT validated
    86:  ('COM.40250', 'Manufacturer\'s Instructions'),            # Package inserts current
    87:  ('COM.10050', 'Procedure Manual Elements'),               # Reflex testing criteria
    90:  ('COM.40500', 'Analytical Interferences'),                # Biotin interference
    91:  ('GEN.43022', 'LIS Testing'),                              # LIS interface validation (matches by LIS Testing concept)

    # ─── Equipment & Maintenance ────────────────────────────────────────────
    93:  ('COM.30675', 'Instrument and Equipment Records'),        # Maintenance records
    94:  ('COM.30750', 'Temperature Checks'),                       # Temp monitoring logs
    95:  ('COM.30800', 'Temperature Corrective Action'),            # Temp excursion corrective
    96:  ('COM.30525', 'Maintenance and Function Checks - Waived Tests'),  # function checks (waived equivalent for ranges)
    97:  ('COM.30650', 'Instrument Troubleshooting'),               # Repair records
    98:  ('GEN.41500', 'Defined Water Types'),                      # Water quality
    99:  ('COM.30350', 'Reagent Storage and Handling - Nonwaived Tests'),  # Reagent lot acceptance
    100: ('COM.30350', 'Reagent Storage and Handling - Nonwaived Tests'),  # Reagent storage conditions
    101: ('COM.30650', 'Instrument Troubleshooting'),               # Out-of-service instruments
    102: ('COM.30820', 'Quantitative Pipette Accuracy and Reproducibility'),  # Pipette calibration
    103: ('GEN.41017', 'Centrifuge Operating Speeds'),              # Centrifuge speed/timer
    104: ('COM.30400', 'Reagent Expiration Date - Nonwaived Tests'),  # Reagent expiration
    105: ('COM.30450', 'New Reagent Lot and Shipment Confirmation of Acceptability - Nonwaived Tests'),  # New lot acceptance
    106: ('COM.30650', 'Instrument Troubleshooting'),               # Instrument downtime log

    # ─── Safety & Environment ───────────────────────────────────────────────
    109: ('GEN.76000', 'Chemical Hygiene Plan'),                    # Chemical hygiene plan
    110: ('GEN.74100', 'Safety Data Sheets'),                       # SDS accessible (verify GEN.74100 exists)
    111: ('GEN.73900', 'Personal Protective Equipment'),            # PPE (verify exists)
    112: ('GEN.74800', 'Viral Exposure'),                           # Hep B vaccination
    113: ('GEN.74200', 'Fire Safety'),                              # Fire safety equipment (verify)
    114: ('GEN.74400', 'Biohazardous Waste Disposal'),              # Biological waste (verify)
    115: ('GEN.74600', 'Emergency Procedures'),                     # Emergency procedures (verify)
    116: ('GEN.74800', 'Viral Exposure'),                           # Exposure incident reports
    117: ('GEN.74300', 'Laboratory Ventilation'),                   # Lab ventilation
    118: ('GEN.73950', 'Eyewash Stations'),                         # Eyewash stations
    119: ('GEN.74050', 'Chemical Storage'),                         # Chemical storage segregated
    121: ('GEN.20326', 'Assessment of the QMS Implementation'),     # Annual safety inspection

    # ─── Blood Bank & Transfusion (TRM module) ──────────────────────────────
    # Items 122-141: items 131 was IM.02.01.01 EP4 in TJC; CAP cited TRM.60000.
    # Need to map each to real TRM IDs.
    122: ('TRM.40550', 'Forward/Reverse Typing'),                   # ABO/Rh typing
    123: ('TRM.40650', 'Serologic Crossmatch'),                     # Crossmatch
    124: ('TRM.40600', 'Unexpected Antibody Screen'),               # Antibody screen
    125: ('TRM.40950', 'Clerical Identification and Transfusion Records Final Check'),  # 2-person verify
    126: ('TRM.41650', 'Transfusion Reaction Recognition'),          # Transfusion reaction workup
    127: ('TRM.30970', 'Donor and Transfusion-Related Fatality Notifications'),  # FDA fatality reporting
    128: ('TRM.32208', 'Collection/Processing Equipment'),           # Blood storage temp (verify - may need TRM.41500 area)
    129: ('TRM.31375', 'Inventory Control'),                         # Inventory management + traceability
    130: ('TRM.41000', 'Blood Administration Procedure'),            # Massive Transfusion Protocol (closest in TRM)
    131: ('TRM.32250', 'Record Retention - Transfusion Medicine'),   # BB records retention
    132: ('TRM.40700', 'Selection of Blood Components'),             # Emergency release uncrossmatched
    133: ('TRM.40925', 'Blood/Component Compatibility Label or Tag'),  # Visual inspection before issue
    134: ('TRM.40720', 'Provisions for Special Components'),         # Irradiation/leukoreduction
    135: ('TRM.40875', 'Transfusion Service Medical Director Responsibility'),  # Utilization review
    136: ('TRM.41000', 'Blood Administration Procedure'),            # Consent for transfusion (procedure)
    137: ('TRM.40230', 'Specimen Labeling for Pretransfusion Testing'),  # ISBT 128 labeling
    138: ('TRM.41025', 'Transfusionist Training'),                   # Blood admin training nursing
    139: ('TRM.40720', 'Provisions for Special Components'),         # Special transfusion req (CMV neg, etc)
    140: ('TRM.32300', 'Receipt of Blood'),                          # Donor records/deferrals
    141: ('TRM.31241', 'Reagent QC'),                                # BB QC reagent daily

    # ─── Point of Care Testing (POC module) ─────────────────────────────────
    142: ('POC.06800', 'Authorized POCT Personnel'),                # POCT coordinator
    143: ('POC.06875', 'Competency Assessment - Waived Testing'),   # POCT operators trained
    144: ('POC.06915', 'Competency Assessment Frequency - Nonwaived Testing'),  # Annual POCT competency
    145: ('POC.07300', 'Daily QC - Nonwaived Tests'),               # POCT QC frequency
    146: ('POC.07484', 'QC Corrective Action'),                     # POCT QC corrective
    147: ('POC.03810', 'Manufacturer\'s Instructions'),             # POCT devices maintained
    148: ('POC.07550', 'Monthly QC Review'),                        # POCT lot/exp recorded
    149: ('POC.04400', 'Results in Medical Record'),                # POCT results in EMR
    150: ('POC.07037', 'QC - Waived Tests'),                        # POCT PT/APA
    151: ('POC.09300', 'Procedure Manual'),                         # POCT P&P accessible
    152: ('POC.08300', 'Recalibration/Calibration Verification Criteria'),  # POCT glucose correlation
    153: ('COM.30000', 'Critical Result Notification'),             # POCT critical values
    154: ('GEN.20361', 'CLIA Certificate Type'),                    # POCT inventory + CLIA coverage
    155: ('POC.03700', 'Error Detection and Correction'),           # POCT oversight committee
    156: ('POC.06925', 'Competency Corrective Action'),             # POCT operator lockout
    157: ('POC.08300', 'Recalibration/Calibration Verification Criteria'),  # POCT glucose meter FDA 2016
    158: ('GEN.43022', 'LIS Testing'),                              # POCT connectivity

    # ─── Leadership & Governance ────────────────────────────────────────────
    159: ('DRA.10460', 'Director Responsibility - PT/QC'),          # Director reviews/signs policies
    160: ('DRA.10500', 'Director Responsibility - Communication'),  # Lab reports to C-suite
    161: ('GEN.26791', 'Terms of Accreditation'),                   # External agency recommendations acted on
    162: ('GEN.20361', 'CLIA Certificate Type'),                    # CLIA cert posted
    163: ('GEN.26791', 'Terms of Accreditation'),                   # Notify accreditor of changes
    166: ('DRA.10500', 'Director Responsibility - Communication'),  # Hospital patient safety participation
    168: ('GEN.23584', 'Interim Self-Inspection'),                  # Annual self-assessment / mock inspection
}

# Verify each proposed CAP ID actually exists in MAS
import glob, warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

CAP_DIR = r'C:/Users/veril/OneDrive/Desktop/Lab/Regulatory/2026 Cap checklists'
ID_PATTERN = re.compile(r'^[A-Z]{2,4}\.\d{5}$')
mas_ids = set()
mas_subjects = {}
for fp in glob.glob(os.path.join(CAP_DIR, 'MAS_*.xlsx')):
    wb = load_workbook(fp, data_only=True, read_only=True)
    for r in wb.active.iter_rows(min_row=7, values_only=True):
        if r[0] and ID_PATTERN.match(str(r[0]).strip()):
            cap_id = str(r[0]).strip()
            mas_ids.add(cap_id)
            mas_subjects[cap_id] = (str(r[3]).strip() if len(r) > 3 and r[3] else '')

# Verify mapping
print('=== Verifying each proposed CAP ID exists in MAS ===')
bad = []
for item_id, (cap_id, label) in MAPPING.items():
    if cap_id in mas_ids:
        actual = mas_subjects[cap_id]
        match = '✓' if actual.lower() == label.lower() else '~'
        print(f'  {match} item {item_id:3d}: {cap_id:12s} "{actual[:65]}"  (proposed: "{label[:50]}")')
    else:
        bad.append((item_id, cap_id, label))
        print(f'  ✗ item {item_id:3d}: {cap_id:12s} NOT IN MAS  (proposed: "{label}")')

print(f'\nTotal verified: {len(MAPPING) - len(bad)} / {len(MAPPING)}')
if bad:
    print(f'Bad picks (need to fix): {len(bad)}')
    for item_id, cap_id, label in bad:
        print(f'  item {item_id}: {cap_id} - needs alternate')
