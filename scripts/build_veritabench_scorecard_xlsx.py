#!/usr/bin/env python3
"""Build the VeritaBench module audit scorecard (internal deliverable for Michael).

4-lens read-only audit (reliability/multi-lab, export/PDF vs Sec5/6, math/compliance-
copy/data-truth, UX), each finding independently verified by the main agent against
current code. VeritaBench = the Operations umbrella: Productivity Tracker + Staffing
Analyzer + PI (performance indicators). The /api/inventory/* routes that also live in
server/veritabench.ts belong to VeritaStock and are OUT OF SCOPE (separate audit).

Main-agent verifications / recalibrations vs the sub-agents:
  - CONFIRMED HIGH: PI-entries cross-account IDOR (unvalidated metric_id + global
    UNIQUE(metric_id,year,month) + account-less read-back), the fabricated peer/
    industry-benchmark capability (hardcoded BENCHMARKS, zero peer aggregation exists),
    the error-as-empty across all 3 ops dashboards (PI worst), and the silent staffing
    hourly-grid autosave.
  - The paid-vs-productive forecast basis (B7), the CFR subsection (B6), the PI
    weighted-vs-unweighted average (B8), and the VeritaPace/VeritaBench branding (B22)
    are methodology/regulatory/brand judgment calls -> flagged for Michael's ruling,
    not fixed unilaterally.
  - CLEAN confirmed: the Leverage Report PDF has NEITHER VeritaResponse bug (identity
    from the labs table, not getUserById; real Page-X-of-Y footer; Perplexity author);
    forecast math guards divide-by-zero; delete actions use confirm dialogs.

Usage:  python scripts/build_veritabench_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaBench_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaBench_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 20)]

ROWS = [
    (1, "SECURITY / cross-account IDOR: POST /api/pi/entries lets any suite-plan user overwrite AND read another ACCOUNT's PI quality-indicator entries by passing a foreign metric_id. The upsert conflict target and the read-back both omit account_id.",
        "Security/Multi-lab", "server/veritabench.ts:2370-2388 ; schema db.ts:4687-4699 (UNIQUE(metric_id,year,month))",
        "HIGH",
        "The handler takes metric_id straight from the body with NO ownership SELECT (every sibling child-write validates the parent by `id AND account_id`; this one does not). The INSERT ... ON CONFLICT(metric_id, year, month) DO UPDATE SET value/volume/notes does NOT include account_id in the conflict target or the SET, and pi_entries has a GLOBAL UNIQUE(metric_id, year, month). Exploit: attacker (account B) POSTs {metric_id: <A's metric id, a small int>, year, month, value:0}; the conflict fires on A's row and overwrites A's value/volume/notes (account_id stays A). The read-back `SELECT * FROM pi_entries WHERE metric_id=? AND year=? AND month=?` (no account filter) then returns A's row, so B also READS A's data. B can also pre-empt an unused (metric,year,month) so A can never file it. Cross-tenant write + read + denial.",
        "Validate `SELECT 1 FROM pi_metrics WHERE id=? AND account_id=?` (404 if not owned) before the upsert; filter the read-back by account_id; add account_id to the uniqueness. Fix first.",
        "OPEN - HIGH (security)"),
    (2, "Fabricated capability: the product advertises 'peer comparison / how you compare to peer labs / industry benchmarks / Peer Benchmarking' and a public dollar 'labor savings vs benchmark' figure, but it computes NO peer aggregation. It compares one lab's ratio to three hardcoded constant bands with no source.",
        "Data-truth/Copy", "client/src/pages/VeritaBenchPage.tsx:63-67 (BENCHMARKS) + 331,353,359,370,416 ; DemoPage.tsx:356,1161",
        "HIGH",
        "BENCHMARKS = { community:{.15,.22}, trauma:{.09,.13}, reference:{.06,.09} } are hardcoded constants; a grep for any aggregation over other labs' productivity returns NOTHING. Yet the copy says 'how you compare to peer labs' (353), 'compares to industry benchmarks' (359), SEO 'peer comparisons' (331), 'Peer Benchmarking' (416), and the public no-login DemoPage computes `${annualSavings}/yr in labor savings vs. benchmark` (356) + 'instantly benchmark against industry standards' (1161). The module's own Excel disclaimer contradicts this ('not benchmarks against an external standard', veritabench.ts:437). This is the fabricated-capability + unsourced-statistic class.",
        "Either source the bands to a citable authority (e.g. CAP LMIP) and label them a static reference range, or drop 'peer/industry benchmark' language for 'target band you configure'; remove the public $-savings-vs-benchmark claim.",
        "OPEN - HIGH (data-truth)"),
    (3, "Error-as-empty across ALL THREE ops dashboards: every GET swallows failures (empty catch, no !res.ok) and falls through to an empty/zero state. Worst: a failed PI-departments load drops an EXISTING lab into the 'Set Up Your PI Program' first-time screen, inviting duplicate departments/metrics.",
        "Reliability/UX", "VeritaBenchPage.tsx:203-208 ; VeritaBenchStaffingPage.tsx:670-675 ; VeritaBenchPIPage.tsx:397-408 (isFirstTime = !loading && departments.length===0)",
        "HIGH",
        "loadData/loadStudies/loadDepartments all do `if (res.ok) setState(...)` with no else + `catch {}`. On a 500/403 the productivity Dashboard renders Current/YTD/YoY/OT as '-' (identical to a new account), the staffing list shows 'No staffing studies yet', and PI computes isFirstTime=true -> the full-page library picker, actively instructing an existing lab to rebuild its PI program (duplicate creation). A director reads zeros/first-time as operational truth on the surface that drives staffing + CFO decisions. Same class the sweep fixed in VeritaLab/QC/Response; VeritaBench never got the error card + Retry.",
        "Add an error state to every GET; render a distinct 'Couldn't load, Retry' card (copy VeritaQCAppPage:518-520); only show first-time/empty after a confirmed 2xx-empty response.",
        "OPEN - HIGH, clean fix ready"),
    (4, "The staffing hourly-data grid autosaves SILENTLY: a debounced POST with no !res.ok check, an empty catch, and no 'Saving/Saved' indicator. A failed save of the 6-week x 7-day x 24-hour grid (the basis of the adequacy determination) is invisible; the tech assumes it saved.",
        "Reliability/UX", "client/src/pages/VeritaBenchStaffingPage.tsx:190-208 (debounce) + 718-729 (handleSaveData)",
        "HIGH",
        "handleSaveData: `try { await fetch(.../data, POST) } catch {}` then loadStudyData() -- no res.ok, empty catch, no toast either way, no save-state UI. Hand-entering the hourly grid is the most tedious input in the module; a 500/token-expiry loses it with zero signal, and that grid feeds the workload analysis + the 42 CFR 493.1445 adequacy attestation. Real data-loss on a compliance input.",
        "Surface save state (pending -> Saved / destructive toast on !res.ok and in catch); reconcile the reloaded grid so a failed save is visible.",
        "OPEN - HIGH, clean fix ready"),
    (5, "Multi-lab: the ops routes trust a CLIENT-SUPPLIED ?labId with no membership validation (resolveOpsLabId), so a seat user of a multi-lab owner can read/write a SIBLING lab's productivity, forecast, staffing, and PI by swapping ?labId. Cross-account is blocked by account_id; the residue is cross-lab within one owner account.",
        "Security/Multi-lab", "server/veritabench.ts:162-167 (resolveOpsLabId) consumed at 173,183,227,244,258,264,288,307,331,369,540,555",
        "MEDIUM",
        "resolveOpsLabId returns req.scope?.labId (undefined here -- no labScopeMiddleware) else the raw `?labId` query param (Number-coerced, but never membership-validated). resolveLegacyLabId (which DOES validate the X-Active-Lab-Id header) is imported but used only on the out-of-scope inventory routes. Every ops query ANDs account_id = req.ownerUserId ?? req.userId, so a foreign owner's data is blocked; but distinct customer/demo labs hosted under one shared owner account (e.g. comped secondary labs under account 17) have no server-side per-lab boundary -- a per-lab-restricted seat user can pass any sibling ?labId.",
        "Route ?labId through resolveLegacyLabId, or verify the labId belongs to account_id before trusting it.",
        "OPEN - MED (security)"),
    (6, "Wrong CFR subsection in the user-facing staffing-adequacy help: '42 CFR 493.1445(e)(5)'. (e)(5) is the 'ensure QC and QA programs' clause; the 'employ a sufficient number of laboratory personnel' requirement is 493.1445(e)(11).",
        "Data-truth", "client/src/pages/VeritaBenchStaffingPage.tsx:561 (user-facing) + 527 (comment) ; server/veritabench.ts:580 (comment)",
        "MEDIUM",
        "The adequacy panel help reads '...determination that staffing is adequate for the volume and complexity of testing, per 42 CFR 493.1445(e)(5). Record it here.' Verified against Cornell LII: 493.1445 (director responsibilities) is correct, but (e)(5) = 'ensure the QC and QA programs are established and maintained'; the personnel-sufficiency clause is (e)(11). Wrong subsection in customer-facing compliance help. Regulatory citation = your ruling; confirm (e)(11) on eCFR before shipping.",
        "Change the citation to 42 CFR 493.1445(e)(11) in all three spots after you confirm the subsection.",
        "OPEN - your regulatory ruling"),
    (7, "Forecast->FTE basis mismatch: the FTE budget divides a PRODUCTIVE-hours allowance by 2080 PAID hours/FTE, understating the budget by the PTO factor (~10-15%) and therefore OVERSTATING the gap and the capital ask in the CFO leverage report.",
        "Math/Data-truth", "shared/operationsForecast.ts:14-52 (DEFAULT_HOURS_PER_FTE_YEAR=2080 'Paid hours'; annualHourAllowance = goalRatio(productive) x volume, / 2080)",
        "MEDIUM",
        "goalRatio is documented + entered as productive (worked) hours per test ('Monthly Productive Hours' UI, productive_hours column, comment 'productive hours / billable tests'). fteBudget = allowance / DEFAULT_HOURS_PER_FTE_YEAR (2080, labeled 'Paid hours per FTE'). A 1.0 FTE works ~1768-1872 PRODUCTIVE hours/yr (2080 paid minus PTO). Dividing a productive allowance by 2080 paid understates FTE, so chainGap.fteGap = staffingModelFte - fteBudget is inflated. NOTE: the module deliberately reproduces the 'Fantasy Medical Center' paper (0.12 x 450000 / 2080 = 25.96 FTE, 'paper: 25.9'), so the paper itself uses 2080 -- this is only a defect if the ratio is truly productive. Confirm against the source methodology.",
        "Confirm paid vs productive against the source paper. If productive, divide by productive-hours-per-FTE (make the ~0.85 factor explicit); else relabel the ratio + field + UI as paid/worked hours.",
        "OPEN - your methodology ruling"),
    (8, "PI YTD / quarterly averages are UNWEIGHTED means of monthly rates and ignore the collected volume denominator, so for a rate indicator (contamination/1000, TAT%, rejection%) the aggregate is not the pooled rate; it can flip the surveyor-facing green/yellow/red status.",
        "Math/Data-truth", "server/veritabench.ts:2489-2512 (ytdAvg = allVals.reduce/len; volume stored but unused)",
        "MEDIUM",
        "quarters/ytdAvg/pyAvg are simple arithmetic means of monthly `value`. Each entry also stores `volume` (the denominator) but it is never used. For variable monthly volumes the pooled rate (sum value*volume / sum volume) differs materially from the mean of rates (e.g. 5% on 1000 + 20% on 100 -> pooled 6.4% vs unweighted 12.5%). The unweighted average feeds ytdStatus (2512), so it can change the green/red benchmark color a surveyor sees.",
        "If these are rate indicators, compute volume-weighted YTD/quarterly aggregates when volume is present; else label 'average of monthly values' so it isn't mistaken for a pooled rate. Your ruling on intent.",
        "OPEN - your ruling"),
    (9, "The productivity + staffing EXCEL exports resolve lab identity from the account-owner's users row only (never the labs table / selected labId), so a multi-lab export prints the OWNER's lab name/CLIA on a workbook whose data is correctly lab-scoped, and falls back to the PERSON's name when clia_lab_name is null.",
        "Export/Multi-lab", "server/veritabench.ts:381-385 (productivity) + 2057-2061 (staffing)",
        "MEDIUM",
        "`const ownerRow = SELECT clia_lab_name, clia_number, name FROM users WHERE id = accountId; labName = ownerRow.clia_lab_name || ownerRow.name || 'Laboratory'`. The Leverage PDF does this right (reads the labs row by labId), and the productivity route even HAS labId in scope (used to filter data rows) but not for identity. Conditional version of the VeritaResponse identity bug: right data, wrong lab header for a secondary lab; person-name leak when clia_lab_name is null.",
        "Mirror the Leverage PDF: resolve the labs row by resolveOpsLabId(req); fall back through clia_lab_name, never to users.name.",
        "OPEN"),
    (10, "The shared license stamp (applyLicenseToExcelJS) OVERWRITES the route-set identity header/footer on every sheet, dropping the CLIA number from the printed oddHeader/oddFooter identity layers. Sec 6.3 requires the CLIA on every sheet. Repo-wide (affects every Excel export).",
        "Export/Sec 6", "shared/licenseExceljs.ts:86-97 (called after veritabench.ts:507-508/2185-2186 set identity)",
        "MEDIUM",
        "The routes set ws.headerFooter.oddHeader/oddFooter carrying '<lab>  CLIA: <clia>'. Then setHeaderFooter replaces both with `&RLicensed: <clia_lab_name>` / a copyright+license line -- the license context (bencheLicenseCtx) carries NO CLIA. So the lab NAME survives but the CLIA is gone from identity layers (b) and (c); only the About-sheet row (layer a) keeps it. The route's identity header/footer lines are dead code. Shared helper -> fix-the-class.",
        "Have setHeaderFooter APPEND the license line to the existing header/footer (or carry the CLIA in the license context) so the Sec 6.3 identity survives.",
        "OPEN - fix-the-class (all Excel exports)"),
    (11, "Em-dashes in the customer-facing Excel About-sheet cells + titles (8 shipping instances across the productivity + staffing exports). Sec 3 / Sec 6.6 ban em-dashes in every cell/header/About paragraph.",
        "Export/Copy", "server/veritabench.ts:431,434,437,443 (productivity) + 2074,2107,2113,2119 (staffing)",
        "MEDIUM",
        "e.g. 2074 `VeritaBench Staffing Analyzer \\u2014 ${study.name}` and 437 '...applied to the entered values \\u2014 they are not benchmarks...'. These are live cell values (not overwritten). The Leverage PDF is CLEAN (zero em-dashes). Same class as the VeritaTrack/VeritaLab Excel About em-dash fixes.",
        "Replace each \\u2014 with a comma, colon, or period (e.g. title -> 'VeritaBench Staffing Analyzer: ...').",
        "OPEN - clean fix ready"),
    (12, "The core metric is INVERTED in marketing vs the app: the app computes + shows 'productive hours per billable test' (~0.06-0.22, lower-is-better), but the copy calls it 'billable tests per paid hour' (the reciprocal, higher-is-better) and says 'paid' where the app uses 'productive'.",
        "Data-truth/Copy", "client/src/pages/VeritaBenchPage.tsx:413,331,370,374,359,362 vs the computed ratio (72-80)",
        "MEDIUM",
        "Product card 'Tests per Paid Hour' (413), SEO 'Billable tests per paid hour benchmarking' (331), body 'billable tests per paid hour' (370). The app displays productive-hours-per-test and even labels the YoY card 'Improving (lower is better)'. 'Tests per paid hour' is the inverse throughput metric; 'paid' conflicts with the 'productive' hours actually used. Both directions can't be right.",
        "Standardize marketing on 'productive hours per billable test (lower is better)' to match the app; fix 'paid' -> 'productive'.",
        "OPEN"),
    (13, "Dated accreditor-manual reference surfaced to logged-in customers: the PI starter library cites 'AABB Standards 34th ed.', shown in the metric picker ('Source: {metric.source}'). Sec 3 bans naming an accreditor manual by edition.",
        "Copy/Data-truth", "client/src/lib/piStarterLibrary.ts:47,56 ; surfaced VeritaBenchPIPage.tsx:299-301",
        "MEDIUM",
        "source: 'CAP Q-Probes; AABB Standards 34th ed.' and 'AABB Standards 34th ed.; ISO 15189'. 'AABB Standards 34th ed.' names an accreditor manual by edition. (The many 'CAP Q-Probes <year>' entries cite published Q-Probes STUDIES, a gray area the ban may not target -- confirm whether audit.py scans this file.)",
        "Change 'AABB Standards 34th ed.' to 'AABB Standards' (undated); sweep the file for other edition/year manual refs.",
        "OPEN"),
    (14, "The 'CFO report' PDF is built from the SAVED forecast row, not the on-screen inputs, so a director who edits the goal without clicking 'Save goal' hands the CFO a PDF whose FTE/gap/capital numbers contradict the screen.",
        "Reliability/UX", "VeritaBenchPage.tsx:183-199 (generateReport posts empty body) vs live useMemo fcResult 127-136 ; server buildForecastResponse reads productivity_forecasts",
        "MEDIUM",
        "The panel is a live useMemo; generateReport() POSTs an empty body and the server recomputes from the persisted productivity_forecasts row. The button is enabled whenever fcResult exists, with no 'unsaved changes' indicator. Edit goal 0.12->0.10 (panel updates), click CFO report without Save goal -> the PDF shows the old numbers. A leverage/decision artifact that can contradict the screen.",
        "Generate from current inputs, or disable 'CFO report' until saved and show a 'Save goal to refresh report' hint when the form is dirty.",
        "OPEN"),
    (15, "The staffing-adequacy attestation (a 42 CFR 493.1445 compliance record) files on one click with NO confirm dialog, and its submit has a silent network-failure path (try/finally, no catch), unlike the compliance sweep's sign-off pattern.",
        "Data quality/UX", "client/src/pages/VeritaBenchStaffingPage.tsx:540-552 (AdequacyPanel.submit)",
        "MEDIUM",
        "'Record determination' fires the POST immediately -- no ConfirmDialog, no explicit affirmation. submit() has try/finally with no catch, so a transport failure only flips busy off with no toast (server 4xx/5xx ARE toasted at 550). A misclick files the determination with no second look; a network drop files nothing with no warning. 'Clear' (598) also has no confirm.",
        "Add a ConfirmDialog summarizing the determination/signer/title before filing; add a catch with a destructive toast.",
        "OPEN"),
    (16, "PI 'Save All' reports success even when the server rejected the entries: the loop never checks res.ok and toasts 'Saved' unconditionally; the reload then blanks the typed values, contradicting the toast.",
        "Reliability/UX", "client/src/pages/VeritaBenchPIPage.tsx:465-491 (handleSaveAll) ; same in handleAddFromLibrary 570-635",
        "MEDIUM",
        "`for (const m of metrics) { await fetch(/api/pi/entries, POST) }` with no res.ok check; `toast({title:'Saved'})` fires regardless; a non-2xx does not throw. loadEntries() then repaints the pre-save state. False-success on a quality-metric record.",
        "Collect per-request res.ok; toast an accurate result ('Saved 4 of 6; 2 failed'); same in handleAddFromLibrary.",
        "OPEN"),
    (17, "The staffing study-detail load fails silently, so the Analysis tab + the adequacy determination are computed on all-zero data (blank heatmaps, 0 weekly volume) that looks like a low-volume lab rather than a fetch failure.",
        "Reliability/UX", "client/src/pages/VeritaBenchStaffingPage.tsx:677-685 (loadStudyData) consumed by AnalysisView 324-522",
        "MEDIUM",
        "loadStudyData: `catch {}`, no !res.ok. On failure studyData stays [] and AnalysisView renders Weekly Volume 0, blank heatmaps, staffing-vs-demand '-'. A director could read the Analysis tab and file an adequacy determination against what is actually a load error.",
        "Add an error state on the study-detail load with Retry; block the Analysis/adequacy view when the load errored.",
        "OPEN"),
    (18, "Class: most mutations across all three pages are silent on a server (!res.ok) failure -- only a network throw toasts. Deletes, exports, and creates do `if (res.ok) { toast; reload }` with no else, so a 404/500 does nothing and the user assumes success.",
        "Reliability/UX", "VeritaBenchPage.tsx:273-290 ; VeritaBenchStaffingPage.tsx:696-716,731-740 ; VeritaBenchPIPage.tsx:554-566",
        "MEDIUM",
        "handleDelete/handleExport (productivity), handleCreate/handleDelete/handleExport (staffing), handleDeleteMetric (PI) all guard `if (res.ok)` with no else and only a network catch. StudyListView also closes the create dialog optimistically (145) before the result is known, so a silent create failure leaves no dialog, no study, no error. The good pattern (handleSave/saveForecast/saveGrid) already exists in the same files.",
        "Add an else toast on every !res.ok (read err.error); don't close the create dialog until success.",
        "OPEN"),
    (19, "PI Add/Edit Metric has no pending state, so a double-click POSTs twice and creates duplicate metrics (no client-side unique guard).",
        "UX/Data quality", "client/src/pages/VeritaBenchPIPage.tsx:519-552 (handleSaveMetric) + button 1159",
        "MEDIUM",
        "handleSaveMetric has no `saving` flag; the button is disabled only on empty name. A double-click issues two POSTs -> duplicate metrics. The productivity/forecast/grid saves already have a busy flag.",
        "Add a busy flag; disable + show 'Saving...' during the request.",
        "OPEN"),
    (20, "Forecast + staffing-grid loads swallow errors and leave inputs blank; because 'Save goal'/'Save grid' then overwrite the persisted row, a failed load followed by a save destroys the prior saved model.",
        "Reliability", "VeritaBenchPage.tsx:138-162 (loadForecast) ; VeritaBenchStaffingPage.tsx:638-651 (loadGrid)",
        "MEDIUM",
        "Both return/catch {} silently on failure, leaving the form empty. The user, seeing a blank forecast, re-enters + saves, clobbering the saved productivity_forecasts / staffing_grid row that failed to load.",
        "Toast on load failure and/or block save until a successful load, so a transient GET error can't clobber saved data.",
        "OPEN"),
    (21, "The staffing Averages sheet lacks a real auto-filter (Sec 6): the route sets autoFilter only as a protection PERMISSION and never assigns wsAvg.autoFilter = {...} (the productivity sheet does).",
        "Export/Sec 6", "server/veritabench.ts:2193 (permission only) vs 493 (productivity sets a real autoFilter)",
        "LOW-MED",
        "Sec 6 requires auto-filter on all columns of a customer-facing workbook. The productivity export sets ws.autoFilter at 493; the staffing Averages sheet only enables the autoFilter protection permission (2193) and never assigns the range, so the shipped sheet has no filter.",
        "Assign wsAvg.autoFilter = { from/to } like the productivity sheet.",
        "OPEN"),
    (22, "Brand inconsistency: this module renders as 'VeritaPace(TM)' in the SEO/hero/leverage-PDF and 'VeritaBench(TM)' in the route gates/About titles/one filename, with the productivity export named VeritaBench server-side but VeritaPace client-side. VeritaBench and VeritaPace are DISTINCT modules in the suite.",
        "Copy/Brand", "leverageReport.ts:186,163 ; veritabench.ts:356,524 vs VeritaBenchPage.tsx:287,331,370,416 (route gates say VeritaBench(TM))",
        "MEDIUM",
        "The Leverage PDF footer + the productivity SEO/hero render 'VeritaPace(TM)' (consistent with the leverage-chain design + the file comment 'VeritaPace Operations Leverage Report'), while the routes (/veritabench/*), plan gates ('VeritaBench(TM) requires a suite subscription'), and About-sheet titles say 'VeritaBench'. The productivity download is named 'VeritaBench-Productivity' server-side (veritabench.ts:524) but 'VeritaPace-Productivity' client-side (VeritaBenchPage.tsx:287). TM (not R) is satisfied everywhere; the issue is WHICH brand. Your call on the canonical name for this surface.",
        "Decide VeritaBench vs VeritaPace for this module; align the About titles, both export filenames, the SEO/hero, and the PDF footer.",
        "OPEN - your brand ruling"),
    (23, "The Leverage PDF stamps another lab's name + CLIA when a foreign ?labId is passed (identity-disclosure only; the numbers stay account-scoped and empty).",
        "Security/Export", "server/veritabench.ts:338-341 (SELECT lab_name, clia_number FROM labs WHERE id = <raw ?labId>)",
        "LOW",
        "labRow = labs WHERE id = ?labId with no ownership check; the header prints that lab's name/CLIA. All forecast numbers come from account-scoped buildForecastResponse (empty for a foreign lab), so only the identity header leaks. Same root as B5.",
        "Validate labId against the account before the identity lookup (fixes with B5).",
        "OPEN"),
    (24, "Excel author metadata ships as 'Michael Veri / Veritas Lab Services, LLC', not 'Perplexity Computer': the route sets wb.creator = 'Perplexity Computer' but applyLicenseToExcelJS overwrites it with AUTHOR_META.",
        "Export/Metadata", "server/veritabench.ts:377,2053 (dead) vs shared/licenseExceljs.ts:163-164",
        "LOW",
        "wb.creator = 'Perplexity Computer' is overwritten repo-wide by the license stamp. The Leverage PDF author is correct (Perplexity Computer via stampPdfAuthor). Resolve by aligning the standard or the override.",
        "Either update the Excel standard, or have the license stamp keep 'Perplexity Computer' as creator.",
        "OPEN"),
    (25, "The Leverage Report PDF ships without the per-recipient license band / copyright + license-terms appendix that every other generated PDF/XLSX carries (it calls only stampPdfAuthor).",
        "Export", "server/leverageReport.ts:192-206 (no injectLicenseHtml / licenseAugmentedFooterTemplate)",
        "LOW",
        "licenseStamp.ts states the stamp is 'applied to every PDF/XLSX the website generates'; the Leverage PDF is an exception -- no license band, no appendix, no per-recipient watermark on a customer-facing PDF.",
        "Apply the license stamp to the Leverage PDF for parity.",
        "OPEN"),
    (26, "LOW batch (copy + a11y polish): (a) DemoPage labels a high ratio 'Overstaffed' + shows a $-savings figure, contradicting the product's own 'depends on test mix' disclaimer; (b) marketing peer-group names 'Clinic/Community/Hospital' don't match the coded tiers community/trauma/reference; (c) 'Productive %' denominator mislabeled 'total worked hours' (really total PAID); (d) Excel About section order deviates from Sec 6.2 + no labeled provenance block; (e) bare 'laboratory director' (no 'or designee') in the Leverage PDF header line.",
        "Copy/Sec 6", "DemoPage.tsx:322,352-357 ; VeritaBenchPage.tsx:370 vs 63-67 ; veritabench.ts:434 ; veritabench.ts:430-443 ; leverageReport.ts:163",
        "LOW",
        "(a) 'Overstaffed' + savings claim vs the disclaimer that a good/bad ratio depends on mix/automation. (b) copy tiers != coded tiers. (c) productive/(productive+non_productive) is a share of PAID hours, not 'worked'. (d) About order is title->identity->about->how-to->disclaimer vs Sec 6.2 title->identity->disclaimer->how-to->provenance. (e) header says 'For laboratory director to CFO review'; footer correctly uses 'or designee'.",
        "Soften 'Overstaffed' -> 'Above target range'; align peer-group names to the 3 tiers; relabel the % denominator; reorder the About sheet; add 'or designee' to the header.",
        "OPEN"),
    (27, "LOW batch (UX/a11y): icon-only Edit/Delete/back/remove-row buttons across all 3 pages have no aria-label; no explicit 'Add PI department' control (only implicit via the library); a red dashboard metric is a dead-end (no link to document RCA / no overdue indicator); the productivity Dashboard flashes '-'/'No data' during load (no loading gate on the default tab).",
        "UX/a11y", "VeritaBenchPage.tsx:688-689,487-533 ; VeritaBenchStaffingPage.tsx:775,114,869 ; VeritaBenchPIPage.tsx:1026,1029,792-800,570-635",
        "LOW",
        "Screen-reader users hear unlabeled buttons; a lab wanting a department outside the 11-item library has no direct create; a yellow/red PI count on the dashboard doesn't link to the RCA workflow; the default productivity tab has no loading gate so it flashes empty until data arrives.",
        "Add aria-labels; add an explicit Add-Department control; link dashboard red metrics to the RCA flow; gate the Dashboard tab on loading.",
        "OPEN"),
]

PASS_NOTES = [
    "The Leverage Report PDF is CLEAN of both VeritaResponse bug classes: lab identity is resolved from the labs table by labId (leverageReport reads labRow.lab_name/clia_number, NOT the typed getUserById), so it is multi-lab-safe and prints the real lab name + CLIA with a proper 'Not on file' fallback; the footer is a real template carrying Page X of Y (not the empty-footer bug); author metadata is 'Perplexity Computer' via stampPdfAuthor; TM (not R) throughout; ZERO em-dashes in leverageReport.ts; no forbidden URLs (only the info@ email). Its gaps are cosmetic (VeritaPace/VeritaBench naming, one bare 'laboratory director' header phrase, and no license appendix).",
    "The forecast + staffing math (shared/operationsForecast.ts) is robust against garbage input: fteBudget guards hoursPerFteYear > 0, projectedProductivity guards annualVolume > 0, staffingGridFte guards weeklyHoursPerFte > 0 and coerces every input with Number(x) || 0, so empty/NaN inputs yield 0 rather than NaN/Infinity, and it reproduces the FMC worked example. The client productivity ratio()/otPct() also guard divide-by-zero (return null, showing 'Enter a goal...' not NaN). chainGap keeps the gap in FTE units and never multiplies by a cost-per-FTE, avoiding a whole class of capital-math errors. The ARITHMETIC is sound; the open question is the paid-vs-productive BASIS (#7), not a divide-by-zero or sign error.",
    "Cross-ACCOUNT scoping is correct on productivity / forecast / staffing-grid / staffing-studies: every read and every :id mutation ANDs account_id, and the staffing :id/data, attest-adequacy, and PI :id/rca child-writes validate the parent by `id AND account_id`. The account-less IDOR (#1) is isolated to POST /api/pi/entries; the rest of the module does NOT repeat it. The ?labId gap (#5) is cross-lab-within-owner, not cross-account.",
    "UX fundamentals that ARE handled well: all three pages use AlertDialog confirms on deletes; several mutations (month save, saveForecast, generateReport, saveGrid, the adequacy server-error path) have proper pending + success + destructive-error toasts; the forecast shows 'Enter a productivity goal and volume' instead of NaN; wide tables are overflow-x-auto (mobile); the plan gate uses the required allowlist (SUITE_PLANS); and the adequacy attestation captures + displays signer/title/date. The systemic gaps are READ-path error states (#3) and a cluster of WRITE-path silent failures (#4/#16/#17/#18/#20), not the happy path.",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaBench Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaBench - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-12)   [inventory = VeritaStock, out of scope]"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 92
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
for sev in ("HIGH", "MEDIUM", "LOW-MED", "LOW"):
    print(f"  {sev}: {sum(1 for x in ROWS if x[4]==sev)}")
