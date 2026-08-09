#!/usr/bin/env python3
"""Build the VeritaMap module audit scorecard (internal deliverable for Michael).

Every row was verified against current code (post PR #979 Mayo removal and PR #980
silent-save fix). Severity is calibrated, not inflated; findings that did not
reproduce are recorded as dropped so the document is honest about what was checked.

Usage:
  python scripts/build_veritamap_scorecard_xlsx.py
Outputs:
  C:/Users/veril/Downloads/VeritaMap_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaMap_Scorecard.xlsx"

TEAL = "01696F"
WHITE = "FFFFFF"
ALT = "EBF3F8"
TEXT = "28251D"
GREEN = "437A22"   # fixed / done
RED = "A12C7B"     # high
AMBER = "964219"   # medium
GRAY = "7A7974"    # low / note

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# columns: header, width
COLS = [
    ("#", 5),
    ("Finding", 40),
    ("Area", 16),
    ("File : Line", 34),
    ("Severity", 12),
    ("Evidence (verified against code)", 52),
    ("Recommendation", 46),
    ("Status", 20),
]

# Each row verified this session. sev drives the Severity cell color.
ROWS = [
    (1, "Silent-save on lab critical values + AMR: bare fetch, no res.ok check, "
        "local state updated even on 401/500/validation reject, so the UI showed "
        "'Saved' when the server rejected it.",
        "Data integrity", "VeritaMapMapPage.tsx:1920-1937", "HIGH",
        "Both save handlers did await fetch(...) then setState unconditionally; every "
        "other fetch in the file checks res.ok. On the exact critical-value entry path.",
        "Handlers now check res.ok, toast the server reason, throw on failure, and update "
        "state only on success; manual Save button surfaces the error state.",
        "FIXED - PR #980 / 7ad9050"),
    (2, "Correlation CFR citation is inconsistent across the module (same requirement, "
        "two different sections).",
        "Regulatory cite", "VeritaMapPage.tsx:17; VeritaMapMapPage.tsx:1264,2597; veritamapData.ts:146",
        "MEDIUM",
        "Cited as 42 CFR 493.1255(b)(3) in one place and 493.1213 in three. CLIA's "
        "comparison-of-results requirement is 493.1281; 493.1255 is calibration verification.",
        "Standardize to one section. Confirm the correct citation against the Master "
        "Citation Index before changing (Michael's regulatory ruling).",
        "OPEN - needs your citation ruling"),
    (3, "Export About-sheet column guide is stale / off-by-one (a Serial Number column "
        "shifted every letter).",
        "Export doc", "server/veritamapData.ts:144-150", "LOW-MED",
        "Guide says 'Column G: CFR Section' but G is 'Number of Instruments'; 'Column H: "
        "Correlation' but H is CFR Section. Predates and is unrelated to the Mayo removal.",
        "Rewrite the column guide against the real header array (routes.ts:13616-13628). "
        "Deterministic; deferred out of the Mayo PR to keep that scope clean.",
        "OPEN - doc-accuracy fix"),
    (4, "Competency-methods overstatement: 'All 6 competency assessment methods must be "
        "documented for every testing personnel member per test.'",
        "Regulatory copy", "VeritaMapPage.tsx:15", "LOW-MED",
        "Not all 6 elements apply to every test (e.g., PT-sample element 5, maintenance "
        "element 4 may be N/A). 'per test' overstates.",
        "Reword to 'all applicable competency methods' or similar. Your regulatory call on "
        "the exact framing.",
        "OPEN - your call"),
    (5, "cal-ver cadence phrasing: 'Non-waived tests require cal ver every 6 months.'",
        "Regulatory copy", "VeritaMapMapPage.tsx:457", "LOW",
        "6 months is the CLIA floor (493.1255(b)(2)); phrasing omits the lot-change / "
        "maintenance / QC-drift triggers and the full-recalibration alternative.",
        "Optional precision tweak. Your call whether the simplification is acceptable "
        "for the UI hint.",
        "OPEN - your call"),
    (6, "'validate' used for lab actions (labs verify, manufacturers validate).",
        "Copy rule", "VeritaMapResourcesPage.tsx:48,150,151", "LOW",
        "'each lab must establish and validate its own critical value thresholds'; "
        "'must be validated before clinical use'; 'laboratory-validated values'. In the "
        "kept Mayo-link block and the Resources disclaimer.",
        "Consider 'establish and verify' / 'verified'. Domain-loose usage; low priority.",
        "OPEN - your call"),
    (7, "Two-store model (instrument_tests source of truth vs veritamap_tests denormalized "
        "copy) can drift, incl. analyte-name casing.",
        "Architecture", "server/db.ts:4624; rebuildMapTests / resync-complexity", "NOTE",
        "Boot reconciliation at db.ts:4624 is a bounded one-shot orphan DELETE (convergent, "
        "not the cascading UPDATE/INSERT boot anti-pattern). Nightly consistency guard + "
        "resync-complexity admin endpoint already monitor drift.",
        "No action required now. A deeper casing-match audit is optional if drift recurs.",
        "NOTE - monitored"),
    (8, "TJC 'QSA.04.05.01' hardcode (flagged by an audit agent).",
        "Copy", "n/a", "DROPPED",
        "grep of VeritaMap client pages + veritamapData.ts is clean; the claim did not "
        "reproduce in current code.",
        "None. Recorded as dropped so the audit is honest about what was checked.",
        "DROPPED - not reproduced"),
]

SEV_COLOR = {"HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY, "NOTE": GRAY, "DROPPED": GRAY}

wb = Workbook()
ws = wb.active
ws.title = "VeritaMap Scorecard"

# Title bar
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "VeritaMap - Module Audit Scorecard   (verified against code 2026-07-10; post PR #979 + #980)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

# Header row
hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

# Data rows
r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = border
    # Severity color
    sev = row[4]
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(sev, TEXT))
    # Status color: green when it starts with FIXED
    stc = ws.cell(row=r, column=8)
    if str(row[7]).startswith("FIXED"):
        stc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{r-1}"
ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} rows")
