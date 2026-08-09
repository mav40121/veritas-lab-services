#!/usr/bin/env python3
"""Build the VeritaScan module audit scorecard (internal deliverable for Michael).

Findings came from a 4-lens read-only audit (reliability/multi-lab, export/PDF,
compliance/no-PHI, UX) and were then each verified by the main agent against
current code. Severity is calibrated by the main agent, not the sub-agents:
the multi-lab item was downgraded HIGH->MEDIUM (within-account, not cross-tenant),
and one agent's severities were normalized. The no-PHI / URL-pointers-only
invariant was verified INTACT and is recorded as a pass.

Usage:  python scripts/build_veritascan_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaScan_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaScan_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 16), ("File : Line", 34),
        ("Severity", 11), ("Evidence (verified against code)", 50),
        ("Recommendation", 44), ("Status", 22)]

ROWS = [
    (1, "Item-count drift: the checklist actually has 173 items but every label says '168', and the public page lists a DIFFERENT 10-domain taxonomy than the app.",
        "Public accuracy", "veritaScanData.ts:1; VeritaScanPage.tsx:30,45,68,102,170,233,237; VeritaScanAppPage.tsx:226; routes.ts:15064",
        "MED-HIGH",
        "grep -c of item entries = 173; the app renders SCAN_ITEMS.length (=173) dynamically while all marketing/SEO/PDF/Excel copy hardcodes 168. Public DOMAINS (Lab Administration, Facility & Safety, ...) do not match data-file DOMAINS (Quality Systems & QC, Calibration & Verification, ...).",
        "Decide whether 173 is intended (5 items were added without updating copy). If yes, drive count + domain list from veritaScanData.ts everywhere. Domain taxonomy reconciliation is your marketing call.",
        "OPEN - needs your ruling (is 173 intended?)"),
    (2, "Legacy scan-by-id / items routes lack the lab_id equality guard their export twins have; a multi-lab owner on an unprefixed URL can read/write the wrong one of THEIR OWN labs.",
        "Multi-lab", "server/routes.ts:14238-14313; useActiveLabId.ts:41-45; VeritaScanScanPage.tsx:609-614",
        "MEDIUM",
        "userCanAccessScan (11324) authorizes by ownership OR membership in the scan's own lab, so this is WITHIN-account, not cross-tenant. useActiveLabId derives labId only from the URL prefix; a bare /veritascan-app/:id yields null -> legacy route with no scan.lab_id===activeLab guard (export handlers at 14948/15242 do have it).",
        "Add the scan.lab_id === req.scope.labId 404 guard to the 4 legacy routes (mirror the export handlers), or reject multi-lab on legacy like the competency guard (#978).",
        "OPEN"),
    (3, "Auto-save failure is invisible: onError resets to idle with no toast, so a rejected save silently loses the user's edits.",
        "Reliability", "VeritaScanScanPage.tsx:688",
        "MEDIUM",
        "The mutation DOES check res.ok (676) and throws, so state is not falsely marked saved (not the VeritaMap silent-state class). But onError:()=>setSaveStatus('idle') shows no error and keeps no dirty flag; edits are lost on reload.",
        "Toast on error ('Changes not saved, retry') and keep dirtyRef true so a later change reattempts. Safe, quick.",
        "OPEN - clean fix ready"),
    (4, "Document Library has no loading or error state: a fetch error is indistinguishable from an empty library.",
        "Reliability/UX", "VeritaScanDocumentLibraryPage.tsx:127,242",
        "MEDIUM",
        "docsQuery has no isLoading/isError handling; docs = docsQuery.data || [] falls back to [] on error, rendering the 'No documents yet' empty state. Same bug class as the fixed VeritaComp empty-state.",
        "Render isLoading (skeleton) and isError (retry message) before the empty-state row. Safe, quick.",
        "OPEN - clean fix ready"),
    (5, "Evidence Library and Inspection Proof are not linked from the scan app; a director may never discover the evidence/coverage half of the product.",
        "Discoverability", "VeritaScanAppPage.tsx (no link to /veritascan/documents or inspection-proof)",
        "MEDIUM",
        "grep of VeritaScanAppPage for documents/inspection-proof/Evidence Library = no link. The two pages only link to each other. Reachable by URL but not surfaced from the scan landing.",
        "Add an 'Evidence Library' button in the VeritaScanAppPage header (verify NavBar doesn't already expose them first).",
        "OPEN"),
    (6, "Document-library Excel does not activate the About sheet on open (wb.views missing); the per-scan workbook does.",
        "Export", "server/veritascanDocumentLibraryExcel.ts",
        "MEDIUM",
        "grep for wb.views/activeTab/firstSheet = none. §6.1 requires the workbook to open with the About sheet active; per-scan handler sets it at routes.ts:15213.",
        "Add wb.views = [{ firstSheet:0, activeTab:0, visibility:'visible' }] before writeBuffer.",
        "OPEN - clean fix ready"),
    (7, "Document-library Excel About sheet lacks the 'not a regulatory determination' disclaimer that §6.2 mandates (per-scan workbook has it).",
        "Export", "server/veritascanDocumentLibraryExcel.ts:113-139",
        "MEDIUM",
        "About has provenance sections (URL pointers, No PHI, attestation) but no disclaimer section; per-scan About has one at routes.ts:15069.",
        "Add a Disclaimer section mirroring the per-scan wording (self-assessment record; accreditation manual governs on conflict).",
        "OPEN"),
    (8, "Em-dash in the public Document Library subtitle (CLAUDE.md §3 public-copy ban).",
        "Copy rule", "VeritaScanDocumentLibraryPage.tsx:251",
        "LOW-MED",
        "'Link external documents ... URLs only - files stay in your SharePoint ...' uses an em-dash in customer-visible prose.",
        "Replace the em-dash with a period/colon. Mechanical.",
        "OPEN - clean fix ready"),
    (9, "Export PDF/Excel handlers fail silently on error (catch only console.error, 'fail silently').",
        "UX", "VeritaScanScanPage.tsx:761-766,802-806",
        "LOW",
        "Both check res.ok and throw, but the catch only logs; the user clicks Export, spinner clears, no file, no message.",
        "Toast on catch (useToast already in the file).",
        "OPEN"),
    (10, "Client plan gate is a blocklist (server gate is a correct allowlist).",
        "Policy", "VeritaScanAppPage.tsx:145",
        "LOW",
        "hasPlanAccess = plan !== 'free' && plan !== 'per_study' (blocklist). CLAUDE.md §8 requires an allowlist; server hasScanAccess (13986) is authoritative so not exploitable, but the two can drift and this likely trips script/audit.py.",
        "Replace with an explicit allowlist mirroring the server plan set.",
        "OPEN"),
    (11, "Neither export locks identity/citation columns while unlocking review-input columns (uniform protection).",
        "Export", "veritascanDocumentLibraryExcel.ts:215; routes.ts:15204",
        "LOW-MED",
        "Both call ws.protect but never set cell.protection.locked=false on a review column. May be an intentional read-only export posture (per-scan About says editing happens in the app).",
        "Either unlock the Notes/review column before protecting, or amend §6.4 to allow a fully-locked export. Your ruling.",
        "OPEN - your ruling"),
    (12, "'needs-review' filter is an unlabeled checkbox with no explanation of the criteria.",
        "UX", "VeritaScanDocumentLibraryPage.tsx:327-330",
        "LOW",
        "Filter surfaces NULL effective/review dates + past-due rows (comment :120), but nothing tells the user why a doc needs review.",
        "Add a tooltip/subtext describing the criteria.",
        "OPEN"),
    (13, "Legacy create route can persist a scan with lab_id = NULL (swallowed try/catch).",
        "Reliability", "server/routes.ts:14041-14064",
        "LOW",
        ".run(activeLab?.id ?? null, ...) inside try{}catch{}; a null-lab scan is invisible to lab-scoped list queries. Orphan risk, no leak.",
        "If activeLab resolves null, reject or log rather than persist a null-lab scan.",
        "OPEN"),
    (14, "Inspection Proof has no error state and an all-red, no-onboarding first run.",
        "UX", "VeritaScanInspectionProofPage.tsx:100-108,222-225",
        "LOW",
        "coverageQuery renders isLoading but no isError; on error every item shows red 'No Evidence Linked'. First run (zero links) is a wall of red with no CTA to the library.",
        "Add an isError branch + a first-run hint linking to the Document Library.",
        "OPEN"),
    (15, "wb.creator inconsistent between the two VeritaScan workbooks ('VeritaAssure' vs 'Perplexity Computer').",
        "Export", "veritascanDocumentLibraryExcel.ts:55; routes.ts:14994",
        "LOW",
        "Per-scan uses 'Perplexity Computer' (matches the PDF author rule); the library workbook uses 'VeritaAssure'.",
        "Standardize to 'Perplexity Computer'.",
        "OPEN"),
    (16, "Em-dash glyph used as the empty-cell placeholder in library table cells; 'CMS surveyor' phrasing; 'Validation Study' doc-type label.",
        "Copy (minor)", "VeritaScanDocumentLibraryPage.tsx:365-367,44; VeritaScanPage.tsx:64",
        "LOW",
        "Placeholder '—' appears in customer-visible cells. 'a Joint Commission, CAP, or CMS surveyor' pairs CMS as a peer surveyor. 'Validation Study' is a doc-classification bucket (likely legit) not a lab action.",
        "Swap placeholder to '-'/'None'; 'CMS surveyor' and 'Validation Study' are your regulatory calls.",
        "OPEN - your call"),
]
PASS_NOTE = ("NO-PHI / URL-pointers-only invariant: VERIFIED INTACT. Schema (db.ts:194 'URL pointers only. "
             "No file content ever lands'), forms collect external_url + metadata only, no file input/upload "
             "endpoint, xlsx + PDF hammer 'No PHI by architecture'. Promise holds.")

SEV_COLOR = {"MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaScan Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaScan - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-10)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

# Pass-note row (green banner)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
pc = ws.cell(row=r, column=1, value="PASS  |  " + PASS_NOTE)
pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
pc.fill = PatternFill("solid", fgColor="E6F2F2")
for ci in range(1, 9):
    ws.cell(row=r, column=ci).border = border
ws.row_dimensions[r].height = 30

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} rows + pass-note")
