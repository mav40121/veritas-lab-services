#!/usr/bin/env python3
# build_veritastock_scorecard_xlsx.py
# VeritaStock module audit scorecard (2026-07-13). 4-lens read-only audit
# (reliability+multi-lab, artifact/§5/§6, copy+invariants, UX) with every
# finding verified against the actual code by the main agent before listing.
# House teal style. Author metadata "Perplexity Computer". Delivered to
# C:\Users\veril\Downloads\VeritaStock_Scorecard.xlsx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.filters import AutoFilter

TEAL = "01696F"; SECT = "E6F2F2"; ALT = "EBF3F8"; WHITE = "FFFFFF"
TEXT = "28251D"; DARK = "0A3A3D"
RED = "A12C7B"; AMBER = "964219"; GREEN = "437A22"; GRAY = "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.properties.creator = "Perplexity Computer"
wb.properties.lastModifiedBy = "Perplexity Computer"

def hcell(ws, cell, text):
    c = ws[cell]; c.value = text
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    c.border = border

# ---- Findings sheet ----
fs = wb.active; fs.title = "Findings"
headers = ["ID", "Severity", "Area", "Location (file:line)", "Finding", "Impact / failure scenario", "Fix direction", "Lens", "Status"]
widths = [6, 9, 16, 34, 52, 60, 46, 8, 18]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    col = openpyxl.utils.get_column_letter(i)
    fs.column_dimensions[col].width = w
    hcell(fs, f"{col}1", h)
fs.row_dimensions[1].height = 28
fs.freeze_panes = "B2"

SEV_FILL = {"HIGH": RED, "MED": AMBER, "LOW": GRAY, "RULING": DARK}

rows = [
 # ID, Sev, Area, Location, Finding, Impact, Fix, Lens, Status
 ("H1","HIGH","Access control","server/routes.ts:6443 (audit-log); :6489/:6504 (consumption-events)",
  "Enterprise audit-log and consumption-events read owner-wide (WHERE owner_user_id / lab_id IN all owner labs), not narrowed to the caller's active lab_members. Sibling rollup/incoming/expired endpoints DO JOIN lab_members ON lm.user_id = ?; these two do not.",
  "A tech granted only the ED stockroom via the product's own /veritastock/team feature calls GET /api/labs/<their-lab>/veritastock/audit-log (labScopeMiddleware passes) and receives the FULL enterprise trail: item names, quantities, who-did-what for every location. Defeats the per-location access model. Intra-tenant, no PHI.",
  "Scope both queries to the intersection of owner labs and the caller's active lab_members (reuse resolveEnterpriseGroup, as the rollup does).","1","Verified"),

 ("H2","HIGH","Multi-lab / artifacts","server/veritabench.ts:925 (reorder PDF), :983 (reorder Excel), :1045 (snap PDF), :1129/:1139 (labels), :1206 (count sheet)",
  "The five legacy download routes query WHERE account_id = ? (every lab the owner holds) while the on-screen list (/api/inventory :886) uses resolveLegacyLabId + WHERE lab_id = ?. Identity is stamped from labs WHERE owner_user_id = ? LIMIT 1 (one arbitrary facility).",
  "A multi-lab owner on the bare /veritastock URL sees the active lab's reorder list on screen, clicks Download PDF/Excel or Snap Order, and gets a document mixing items from EVERY facility they own, headed with one arbitrary facility's name/CLIA. That doc can be sent to a vendor listing another facility's supplies.",
  "Scope the five legacy routes by resolveLegacyLabId(sqlite, req), mirroring the list read they back. Safe /labs/:labId/ variants already exist (:2597+).","1","Verified"),

 ("H3","HIGH","Artifact / Sec 3","server/orderDocument.ts:181, 303, 312, 315, 321 (Reorder PDF); :705, :706, :707 (Snap PDF)",
  "Literal em-dash (U+2014) is emitted as the null/empty placeholder into rendered PDF table cells: Suggested Order, Days Left, After Delivery, Catalog #, Est. Cost (Reorder) and Catalog/Lot/Department (Snap).",
  "Every Reorder and Snap Order PDF a VeritaStock customer downloads contains em-dashes, a Sec 3 NON-NEGOTIABLE / Sec 5 breach on the flagship artifact. Reachable on essentially every order (any item without a catalog number, any zero-shortfall/expiry-driven row). Copy severity, but ubiquitous.",
  "Replace each rendering '—' with a plain '-' or empty string (the Excel builder already uses '' for the same cells). Em-dashes in code comments are fine.","2","Verified"),

 ("H4","HIGH","Copy / pricing","client/src/pages/VeritaStockPage.tsx:1693",
  "The public hero renders a hardcoded 'From $499/yr' tile that does NOT branch on onStock, so it ships on standalone veritastock.com. $499 is the grandfathered legacy price (current Clinic tier is $999), so it is also stale/wrong.",
  "A prospect on either host sees a dollar figure in VeritaStock collateral (violates the no-$-in-collateral posture) AND a stale price that no longer maps to any tier. Wrong-price public claim.",
  "Remove the dollar tile; keep only plan-inclusion language (the adjacent 'All Plans' tile can stay).","3","Verified"),

 ("M1","MED","Reliability (error-as-empty)","VeritaStockPage.tsx:855 (loadItems); VeritaStockAuditTrailPage.tsx:87; SnapOrder:79; Vendors:83; Receiving:110; Enterprise:108",
  "Read-as-empty cluster: if(res.ok) setState with no else + silent catch, so a 500/403 renders the empty/onboarding state. Audit trail case is worst: a failed load shows 'No audit entries yet' on a surveyor-facing trail (false compliance signal). Enterprise incoming case hides pending transfers (badge shows 0).",
  "On a flaky connection a stocked site is told 'No inventory items yet / No audit entries yet', a destination user never sees stock waiting to accept, and a compliance trail reads as 'nothing ever happened'. Same class rated HIGH in five prior modules.",
  "Add an error state + a distinct 'couldn't load, Retry' card; gate every empty/first-time state on a confirmed 2xx. Pattern already in-repo at Enterprise:332-341 and Trends.","1/4","Verified"),

 ("M2","MED","Copy / invariant","VeritaStockPage.tsx:1660 (subtitle), :1681 (surveyor line); ArticleInventoryManagementPage.tsx:305",
  "Standalone-host lab-token leaks: the hero half-branches on onStock (SEO title + badges branch) but the on-page subtitle 'Laboratory Inventory & Reagent Management' and 'former TJC laboratory surveyor' do not, so 'Laboratory' ships on veritastock.com. The resource article's tie-in calls the product 'VeritaStock for laboratory inventory management'.",
  "The standalone multi-site product presents itself as lab-specific, contradicting the positioning; the only permitted 'lab' token in VeritaStock collateral is an email address.",
  "Branch the subtitle on onStock like the SEO title; drop 'laboratory' from the surveyor line ('former TJC surveyor'); reword the article tie-in to 'multi-location clinical supply and inventory management'.","3","Verified"),

 ("M3","MED","Artifact / copy","server/orderDocument.ts:481",
  "The Reorder Excel About-sheet filter banner says 'this workbook does not include the full lab' and is NOT gated on STOCK_DEPLOYMENT, unlike every other lab-token in the file (the sibling inventoryCountExcel.ts:104 gates the same banner).",
  "On the standalone VeritaStock deployment, a filtered reorder workbook's About sheet literally says 'the full lab', a copy-exception leak in a customer artifact.",
  "Apply the sibling ternary: `${STOCK_DEPLOYMENT ? 'every location' : 'the full lab'}`.","2","Verified"),

 ("M4","MED","Reliability (silent mutation)","VeritaStockPage.tsx:1348 (handleDelete); VeritaStockVendorsPage.tsx:105/:508; VeritaStockReceivingPage.tsx:128/:161",
  "Mutations that don't surface failure: handleDelete has a catch but no else (a 403/409/500 closes the dialog silently, item stays); Vendors delete has no res.ok check and the page imports no toast at all; Receiving handleReceive/applyLeadTime have no catch so a network throw re-enables the button with no message.",
  "User deletes an item, dialog vanishes, item is still there, no explanation. On a floor kiosk with spotty wifi a tech taps Receive, the spinner flips back, and there is no confirmation or error, did the stock land or not.",
  "Add else/!res.ok toasts and a network catch (mirror handleReceive at VeritaStockPage:1388, which is correct). Add useToast to the Vendors page.","1/4","Verified"),

 ("M5","MED","Copy / positioning","client/src/pages/HospitalInventoryPage.tsx:41, 43, 71, 135, 149; client/src/lib/faqContent.ts:352; client/src/pages/OperationsPage.tsx:62",
  "The /hospital-inventory landing is built on a competitor-relative cost pitch ('without the six-figure system', 'Enterprise inventory at a fraction of what the big systems cost'), which conflicts with the value-first, no-competitor-relative-discount posture. Plus lab tokens: 'For hospitals, clinics, and labs' (:41), 'Built by lab and quality professionals' (:149), and 'the spreadsheets most labs use today' repeated in faqContent:352 + OperationsPage:62.",
  "The page anchors VeritaStock's value to being cheaper than enterprise systems rather than to capability; and it carries 'lab' tokens across collateral.",
  "Reframe to capability/outcome value (multi-site control, surveyor-ready audit). Swap 'labs' -> 'clinical departments/sites'; 'lab and quality' -> 'clinical and quality'. Positioning half is Michael's call.","3","Verified"),

 ("R1","RULING","Product / memory","memory project_veritastock_costing_roadmap vs VeritaStockPage.tsx:499-528, 2103-2156, 2293, 2374; server/orderDocument.ts:321/377/516/581",
  "The costing-roadmap memory says costing is DEFERRED (quantity-only schema). The shipped code disagrees: in-app Unit Cost input, $-on-hand valuation + inventory-turns + '$X at risk' tiles, ABC-class-by-annual-dollar column, and dollar waste write-offs are all live, and the Order PDF/Excel expose a dollar order total (gated on unit_cost being populated, so it fails safe).",
  "Either the memory is stale or costing shipped without updating it. This decides whether the cost figures in the customer Order artifacts are intended or should be suppressed under the no-$-in-collateral posture.",
  "Michael to rule: (a) costing is now a shipped feature -> update the memory, keep cost in artifacts; or (b) suppress cost in the customer-facing Order PDF/Excel per the deferred posture.","3","Needs ruling"),

 ("L1","LOW","Copy / UI","VeritaStockPage.tsx:1788; VeritaStockEnterprisePage.tsx:126, 220, 221, 582",
  "When a location has no name the UI renders the fallback label 'Lab {id}' (e.g. 'Lab 5'), a lab token used as a location name on a multi-site product.",
  "Cosmetic lab-token leak in the location column/labels.",
  "Default unnamed locations to 'Location N' (or 'Site N').","3","Verified"),

 ("L2","LOW","Reliability (safety)","server/veritastockDemoReset.ts:203, 330",
  "Demo reset issues unscoped DELETEs (DELETE FROM inventory_transfers with no WHERE; DELETE FROM audit_log WHERE module='veritastock' with no lab/owner filter), guarded only by the STOCK_DEPLOYMENT env flag at the entry.",
  "Belt-and-suspenders gap: if STOCK_DEPLOYMENT were ever set on the main DB by misconfig, every customer's transfer history and VeritaStock audit trail would be deleted. Never runs on the main service today.",
  "Add WHERE lab_id IN (DEMO_LABS) / owner scoping so a misconfigured flag cannot cause tenant-wide data loss.","1","Verified"),

 ("L3","LOW","Artifact / consistency","server/orderDocument.ts:437-438; server/inventoryCountExcel.ts:66-67",
  "Excel author metadata is the brand string (wb.creator = 'VeritaStock'/'VeritaAssure') rather than 'Perplexity Computer' as the sibling Excel exports (veritabench.ts:395) and the PDF path use.",
  "Metadata inconsistency; not a black-letter Sec 6 breach (Sec 6 doesn't spell out Excel author) but out of parity.",
  "Set wb.creator / wb.lastModifiedBy = 'Perplexity Computer' in both Excel builders.","2","Verified"),

 ("L4","LOW","Reliability (minor)","server/veritabench.ts:1323",
  "The barcode scan lookup resolves WHERE account_id = ? AND barcode_value = ?, i.e. across all the owner's labs rather than the active lab.",
  "A scanned barcode can resolve to an item in a different facility of the same owner; minor within-tenant scoping looseness on a read.",
  "Scope by resolveLegacyLabId like the list read.","1","Verified"),

 ("L5","LOW","UX polish","VeritaStockPage.tsx:2407 (6 icon buttons/row); Enterprise:469-472 (no loading row) / :471 ('No items match.'); AuditTrail:78 (no debounce); VeritaStockPage:756 (lots dialog error-as-empty); Receiving:282 (h-7 inputs); InventoryKioskPage.tsx:269/362 (dead page 'lab director')",
  "Kiosk ergonomics + polish: six 28px icon-only actions per row (two destructive, adjacent) below the 44px touch target; Enterprise first-load shows 0-value cards with no in-grid loading row and 'No items match.' as the empty state; audit search fires per keystroke; lots dialog hides load errors as 'no lots'; receiving inline inputs cramped on touch; the legacy (redirected, unimported) InventoryKioskPage still carries 'lab director'.",
  "Mis-tap risk on shared kiosks; brief empty/broken first paint; chatty search. All recoverable.",
  "Enlarge/space primary actions; add a loading row + real empty state to Enterprise; debounce audit search; give the lots dialog an error state; leave/retire the dead kiosk page.","4","Verified"),
]

r = 2
for (rid, sev, area, loc, finding, impact, fix, lens, status) in rows:
    vals = [rid, sev, area, loc, finding, impact, fix, lens, status]
    for i, v in enumerate(vals, start=1):
        col = openpyxl.utils.get_column_letter(i)
        c = fs[f"{col}{r}"]; c.value = v
        c.font = Font(name="Calibri", size=10, color=TEXT, bold=(i == 2))
        c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
        c.border = border
        if i == 2:  # severity color
            c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=SEV_FILL.get(sev, GRAY))
        else:
            c.fill = PatternFill("solid", fgColor=(WHITE if r % 2 == 0 else ALT))
    fs.row_dimensions[r].height = 96
    r += 1
fs.auto_filter.ref = f"A1:I{r-1}"

# ---- Summary sheet (sheet 1, active) ----
su = wb.create_sheet("Summary", 0)
su.column_dimensions["A"].width = 120
def s_title(cell, text):
    c = su[cell]; c.value = text
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c.border = border
def s_sect(cell, text):
    c = su[cell]; c.value = text
    c.font = Font(name="Calibri", bold=True, size=12, color=DARK)
    c.fill = PatternFill("solid", fgColor=SECT)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    c.border = border
def s_body(cell, text):
    c = su[cell]; c.value = text
    c.font = Font(name="Calibri", size=11, color=TEXT)
    c.alignment = Alignment(vertical="top", horizontal="left", indent=1, wrap_text=True)

su["A1"].value = None
s_title("A1", "VeritaStock Module Audit Scorecard")
su.row_dimensions[1].height = 30
s_body("A2", "Audit date 2026-07-13. Method: 4 read-only lens agents (reliability + multi-lab isolation; generated artifacts vs Sec 5/Sec 6; compliance copy + product invariants; UX dead-ends), then every finding verified against the actual code by the main agent before listing. Surfaces: ~43 inventory routes (server/veritabench.ts + routes.ts), 11 client pages incl. the 2836-line VeritaStockPage, Order PDF/Excel, count Excel, Intacct CSV, kiosk PIN, cross-site transfers, consumption ledger.")
su.row_dimensions[2].height = 90
s_sect("A3", "Counts: 4 HIGH, 5 MED, 4 LOW, 1 ruling (see Findings tab)")
s_sect("A4", "Bottom line")
s_body("A5", "Security posture is genuinely strong: NO cross-tenant (cross-owner) IDOR, transfer send/accept/reject are atomic with membership + active-location gate, kiosk PIN uses pbkdf2 + timingSafeEqual + lockout + JWT-only labId, and the consumption ledger is append-only and HIPAA-free. The real defects are (1) two cross-LOCATION reads that ignore per-location grants [H1], (2) five artifact download routes that scope by account_id and bleed a multi-lab owner's facilities into one document [H2], (3) em-dashes rendered into every Order/Snap PDF [H3], and (4) a stale/prohibited public price [H4], plus the recurring error-as-empty and silent-mutation classes and the standalone-host lab-token leaks.")
su.row_dimensions[5].height = 118
s_sect("A6", "Confirmed CLEAN (covered, no defect)")
s_body("A7", "No cross-tenant IDOR. Transfer send/batch/accept/reject: shared-owner + member validation, atomic sqlite.transaction(), active-location 409 gate enforced server-side. Kiosk PIN: pbkdf2+timingSafeEqual, 5-attempt/15-min lockout, labId taken from JWT only, kiosk/user JWTs non-interoperable. Consumption ledger: INSERT-only, no PHI. Lots: FEFO, on-hand stays authoritative. Mutation guard resolveInventoryItemForMutation: account_id owner OR active lab_members, 403/404 no existence leak. Lab-identity resolution in exports uses raw labs/users SELECT (no getUserById person-name bug). Count Excel is Sec 6-compliant; Intacct CSV fails safe (named preflight before build). VeritaStockLandingPage (veritastock.com front door) and the DemoPage section are fully clean. Filters actually filter; destructive actions are confirm-gated. Most mutation handlers surface both success and failure.")
su.row_dimensions[7].height = 150
s_sect("A8", "Needs your ruling")
s_body("A9", "R1: the costing-roadmap memory says costing is DEFERRED / quantity-only, but in-app costing (unit cost, $ valuation, ABC class, waste $) and a dollar order total in the Order artifacts are fully shipped. Reconcile the memory, and decide whether cost belongs in the customer-facing Order PDF/Excel. M5 also has a positioning half: whether to keep the competitor-relative 'six-figure system' pitch on /hospital-inventory or reframe to value-first.")
su.row_dimensions[9].height = 74
su.sheet_view.showGridLines = False
fs.sheet_view.showGridLines = False

import os
out = os.path.join(os.path.expanduser("~"), "Downloads", "VeritaStock_Scorecard.xlsx")
wb.save(out)
print("wrote", out)
print("rows:", len(rows))
