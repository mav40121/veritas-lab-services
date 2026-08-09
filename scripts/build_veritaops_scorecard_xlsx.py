#!/usr/bin/env python3
"""Build the VeritaOps module audit scorecard (internal deliverable for Michael).

Last module of the 17-module 4-lens sweep. VeritaOps = the CPRT (Cost Per
Reportable Test) calculator, CLSI GP11-A basis. Public demo /demo/operations
(DemoCprtPage); app /veritaops-app + /labs/:labId/veritaops-app.

4-lens read-only audit (reliability/multi-lab, export/PDF vs Sec5/6, math/
compliance-copy/data-truth, UX), every finding independently re-verified by the
main agent against current code (server/veritaops.ts, client VeritaOpsAppPage.tsx,
client DemoCprtPage.tsx, client hooks/useActiveLabId.ts, client App.tsx).

Main-agent verifications / recalibrations vs the sub-agents:
  - NO HIGH findings. VeritaOps is the cleanest module of the sweep: every
    lab-scoped route has authMiddleware + labScopeMiddleware and constrains
    lab_id = req.scope.labId on read AND :id mutation (no IDOR); the lab-scoped
    PDF reads identity from the labs row by labId (the recurring getUserById
    multi-lab bug is ABSENT on the scoped path); computeCprt math is sound and
    compares against the lab's OWN entered costs, so there is NO fabricated
    peer/industry-benchmark class (contrast VeritaBench).
  - 5 MED confirmed: /veritaops-app missing from LAB_SCOPABLE_PATHS (lab-switch
    can't re-scope the URL), the legacy POST omitting lab_id (orphaned invisible
    study, masked by the LegacyWorkspaceRedirect), the error-as-empty list load
    (2-lens corroborated), the PDF-popup false-success (Gate-3 step-8 class), and
    the legacy PDF stamping owner identity not the active lab.
  - 3 LOW confirmed: silent delete failure, the truncated CLIA-missing string,
    and em-dash null-placeholder glyphs in the in-app table (NOT in any PDF).
  - 3 subagent findings DROPPED on verification (adversarial pass): the GP11-A
    title suffix was already correct (shipped #1143); the demo CTA is NOT a
    dead-end (it links to /pricing, DemoCprtPage:290); the demo "read-only
    tooltip / numeric-0 friction" findings mis-cited lines (DemoCprtPage has no
    disabled buttons or numeric inputs -- it is a read-only presentation page).

Usage:  python scripts/build_veritaops_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaOps_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaOps_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 20)]

ROWS = [
    (1, "Multi-lab: the VeritaOps page is missing from the lab-switcher's scopable-path list, so switching labs while on VeritaOps does NOT rewrite the URL to the lab-scoped route. Every other module's -app path is listed; VeritaOps was omitted from the sync.",
        "Multi-lab", "client/src/hooks/useActiveLabId.ts:10-29 (LAB_SCOPABLE_PATHS) vs App.tsx:508 (lab-scoped route exists)",
        "MEDIUM",
        "LAB_SCOPABLE_PATHS lists /veritascan-app, /veritamap-app, /veritatrack-app, /veritacomp-app, /veritastaff-app, /veritalab-app, /veritapolicy-app, /veritaqc-app, etc., but NOT /veritaops-app. App.tsx:508 DOES define `<Route path=\"/labs/:labId/veritaops-app\">`, so the lab-scoped route exists and the omission is a sync gap (the array's own comment says 'Keep this in sync with the lab-scoped Route list in App.tsx'). Effect: isLabScopablePath('/veritaops-app') returns false, so the LabSwitcher updates the server default + memberships cache but leaves the URL on the unscoped /veritaops-app instead of prefixing /labs/:labId/. The page still works (the unscoped route resolves the active lab server-side), but the switch does not cleanly re-scope the URL the way it does for every other module.",
        "Add \"/veritaops-app\" to LAB_SCOPABLE_PATHS. One-line, safe.",
        "OPEN - MED, clean 1-line fix"),
    (2, "Data-loss shape: the LEGACY POST /api/veritaops/studies INSERT writes account_id but NO lab_id, while every SELECT filters WHERE lab_id = ?. A study created through the legacy route gets a NULL lab_id and is invisible in every list, including the one that just created it.",
        "Reliability/Multi-lab", "server/veritaops.ts:169-201 (INSERT, no lab_id) vs 138 / 313 / 323 (SELECT ... WHERE lab_id = ?)",
        "MEDIUM",
        "The legacy INSERT column list (170-179) is account_id, test_name, loinc, ... notes, created_at, updated_at -- lab_id is absent, so the row lands with lab_id NULL. The legacy list (138), the lab-scoped list (313), and the lab-scoped :id read (323) all filter WHERE lab_id = ?, so a NULL-lab_id study never appears anywhere. In practice this is MASKED: LegacyWorkspaceRedirect funnels a lab member to /labs/:labId/veritaops-app, whose POST (the lab-scoped INSERT) sets lab_id from req.scope.labId. The residual is a user with zero lab membership (can't be redirected), a direct API call, or a race before the redirect resolves -- any of which silently drops the study.",
        "Stamp lab_id on the legacy INSERT (resolve the owner's active/home lab as the GET does), or retire the legacy POST to a 410/redirect. Fixes with #5.",
        "OPEN - MED (latent data-loss)"),
    (3, "Error-as-empty on the studies list load: loadStudies does `if (res.ok) setStudies(...)` with an empty catch and no else, so a 500/403 renders the empty / first-time state as if the lab has no cost studies. (Corroborated independently by the reliability and UX lenses.)",
        "Reliability/UX", "client/src/pages/VeritaOpsAppPage.tsx:657-662 (loadStudies)",
        "MEDIUM",
        "`const loadStudies = useCallback(async () => { try { const res = await fetch(listUrl,...); if (res.ok) setStudies(await res.json()); } catch {} finally { setLoading(false); } })`. On a non-2xx or a network throw, studies stays [] and loading flips false, so the page shows the zero-studies state -- indistinguishable from a genuinely empty lab. A director reads it as 'no CPRT studies on file' rather than 'the load failed'. Same class the sweep fixed in VeritaLab / VeritaQC / VeritaResponse; VeritaOps never got the error card + Retry. NOTE: handleSave (669-691) is the CORRECT pattern in the same file (checks res.ok, else-toasts the server error, catch-toasts) -- copy it.",
        "Add an error state; render a distinct 'Couldn't load studies, Retry' card; only show the empty state after a confirmed 2xx-empty response.",
        "OPEN - MED, clean fix ready"),
    (4, "PDF download reports success even when the browser blocked the popup: handleDownloadPdf correctly checks res.ok on the token fetch, but then ignores the window.open() return and fires the 'PDF generated' toast unconditionally. A blocked popup = success toast with no PDF.",
        "Reliability/UX", "client/src/pages/VeritaOpsAppPage.tsx:705-725 (handleDownloadPdf)",
        "MEDIUM",
        "The token fetch is guarded (710-713 toast a destructive error on !res.ok), so a SERVER failure is handled. But `window.open(\\`${API_BASE}/api/pdf/${token}\\`, '_blank')` (718) is called after two awaits -- outside the direct user-gesture callstack -- so popup blockers routinely block it, and the very next line toasts `PDF generated for ${study.test_name}` (719) regardless of whether the window opened. The user sees success and no document. This is exactly the Gate-3 step-8 / PR #286 browser-only class: a server-side verify script passes (the token IS minted) while the actual click yields nothing. Script-invisible.",
        "Capture the window.open handle; if null/undefined, toast a 'popup blocked -- allow popups or click here' fallback with the token URL instead of a success toast.",
        "OPEN - MED (browser-only, Gate-3 step-8)"),
    (5, "The LEGACY PDF route stamps the account-owner's identity, not the active lab: it resolves labName / CLIA from the users row by ownerId, so a multi-lab owner's secondary-lab PDF prints the owner's lab name + CLIA, and falls back to the PERSON's name/email when clia_lab_name is null. The lab-scoped PDF route is clean.",
        "Export/Multi-lab", "server/veritaops.ts:285-293 (legacy PDF) vs 436-438 (lab-scoped PDF reads the labs row -- clean)",
        "MEDIUM",
        "Legacy route: `ownerRow = SELECT clia_lab_name, clia_number, name, email FROM users WHERE id = ?` (ownerId); labName = ownerRow.clia_lab_name || ownerRow.name || 'Laboratory'; preparedBy = ownerRow.name || ownerRow.email. So for a secondary lab hosted under a multi-lab owner, the legacy PDF header shows the OWNER's lab identity, and when clia_lab_name is null it prints the person's name as the lab. The LAB-SCOPED route (436-446) does this RIGHT -- reads lab_name/clia_number from the labs row by req.scope.labId, per its own comment 'so multi-lab users get the correct CLIA / lab name.' Same masked-by-redirect residual as #2: members are funneled to the scoped route, so the legacy path is reachable only off the redirect. Conditional/edge version of the VeritaResponse identity bug.",
        "Have the legacy PDF resolve identity from the labs row like the scoped route (or retire the legacy PDF path with the legacy POST). Fixes with #2.",
        "OPEN - MED (latent, legacy path only)"),
    (6, "Silent delete failure: handleDelete guards `if (res.ok) { toast; reload }` with no else and an empty catch, so a 404/500 does nothing, the confirm dialog still closes, and the user assumes the study was deleted.",
        "Reliability/UX", "client/src/pages/VeritaOpsAppPage.tsx:693-702 (handleDelete)",
        "LOW",
        "`const res = await fetch(itemUrl(id), {method:'DELETE'}); if (res.ok) { toast('Study deleted'); loadStudies(); } } catch {} finally { setDeleteTarget(null); }`. No else branch, empty catch. A server rejection leaves the study in place but closes the dialog with no error, so the deletion silently no-ops. Low-frequency (delete is rare and the list would still show the row on next load), but it is the same silent-failure class as #3/#4 on the write path.",
        "Add an else toast (read err.error) and a catch toast; keep the dialog open or re-surface the row on failure.",
        "OPEN - LOW"),
    (7, "The PDF's CLIA-missing text is truncated to 'Not on file' rather than the Sec-5 canonical 'CLIA: Not on file - enter in account settings'. Minor wording gap in the report header for a lab that hasn't entered its CLIA.",
        "Export/Sec 5", "server/veritaops.ts:291 (legacy) + 445 (lab-scoped) -- cliaNumber: labRow?.clia_number || 'Not on file'",
        "LOW",
        "Sec 5 (PDF Requirements) says: show 'CLIA: Not on file - enter in account settings' if missing. Both VeritaOps PDF routes pass cliaNumber = <clia> || 'Not on file' -- the short form, which the header then renders. Not a compliance defect (the field is present and honest), just a deviation from the standard's exact missing-CLIA string, so the customer isn't told where to fix it.",
        "Pass the full 'Not on file - enter in account settings' fallback string (matching the other modules) in both PDF routes.",
        "OPEN - LOW (Sec-5 wording)"),
    (8, "Em-dash null-placeholder glyphs in the in-app cost table: six cells render a literal em-dash for an empty/N-A value. These are in-app UI placeholders, NOT prose and NOT in any generated PDF (the PDF is clean), so it's a house-style consistency nit rather than a Sec-3 public-artifact breach.",
        "Copy/Style", "client/src/pages/VeritaOpsAppPage.tsx:535,572,580,853,856,859",
        "LOW",
        "deltaLabel = '\\u2014' (535); cell fallbacks aAnnual/bAnnual != null ? fmt : '\\u2014' (572,580); L3/L4/annual columns render '\\u2014' when include_capital/include_overhead is 0 or volume is 0 (853,856,859). All are placeholder glyphs for 'no value', which is a common UI convention; none reach a PDF, marketing page, or export. Sec 3's em-dash ban targets public-facing artifacts (PDFs, website, collateral). Flagged for consistency only -- swap to a hyphen or 'n/a' if you want strict house style in the app UI too.",
        "Optional: replace the '\\u2014' placeholders with a hyphen or 'n/a' for house-style consistency. No customer-facing artifact impact.",
        "OPEN - LOW (optional style)"),
]

PASS_NOTES = [
    "NO IDOR anywhere in VeritaOps. Every lab-scoped route (list, :id read, PUT, DELETE, PDF) carries authMiddleware + labScopeMiddleware and constrains the row by `lab_id = req.scope.labId` (SELECT ... WHERE id = ? AND lab_id = ?, DELETE ... WHERE id = ? AND lab_id = ?); the mutation resolver returns 403 'You don't have access to this study's lab' (not a 404) for a foreign lab. hasOpsAccess gates the plan on every route. There is no cross-lab UNIQUE collision (the table has no cross-tenant unique index like the VeritaBench pi_entries bug). Cross-account and cross-lab boundaries both hold on the production (lab-scoped) paths.",
    "The CPRT math (computeCprt) is sound and HONEST about its comparison basis. It builds the four cost layers (L1 direct reagent, L2 + QC/cal, L3 + capital, L4 + overhead) from the lab's OWN entered costs and volume, and compares studies against each other -- there are NO hardcoded peer/industry benchmark bands and NO fabricated 'how you compare to peer labs' claim (the class that made VeritaBench a HIGH). The demo (DemoCprtPage) presents a worked example, not a peer aggregate. No unsourced-statistic or fabricated-capability finding for this module.",
    "The LAB-SCOPED PDF is multi-lab-safe and Sec-5 compliant: it resolves lab_name / clia_number from the LABS row by req.scope.labId (server/veritaops.ts:436-438, with the comment 'so multi-lab users get the correct CLIA / lab name'), so the recurring getUserById identity bug is ABSENT on the scoped path (it survives only on the legacy path, #5). Footer carries Page X of Y; author metadata is Perplexity Computer; TM (not R) throughout. The identity issue is confined to the redirect-masked legacy route.",
    "The absence of a director signature on the CPRT PDF is CORRECT, not a Sec-5 violation. CPRT is an internal cost-accounting worksheet (a management/operations artifact), not a compliance determination that requires laboratory-director-or-designee sign-off. This mirrors the Sec-5 internal-use exemption granted to VeritaScan and VeritaLab certificate reports; forcing a signature block here would misrepresent an operational worksheet as a regulatory record.",
    "The happy path and the SAVE path are handled well: handleSave (VeritaOpsAppPage:669-691) checks res.ok, else-toasts the server error message, and catch-toasts a network error -- the correct pattern. Deletes go through a confirm (AlertDialog / deleteTarget). The systemic gaps are narrow and known: the READ path (list load, #3) and two WRITE edges (PDF popup #4, delete #6), plus the two legacy-route residuals (#2, #5). Nothing in the module fabricates data, leaks across accounts, or miscomputes cost.",
    "VERIFICATION DROPPED 3 subagent findings (adversarial pass, so the scorecard doesn't over-call): (a) the CLSI GP11-A title suffix was flagged MED but is CORRECT ('Basic Cost Accounting for Clinical Laboratory Services; Approved Guideline', already shipped #1143 -- confirmed against the CLSI catalog, no phantom 'Laboratory' word missing); (b) the public demo CTA 'Start a free trial and run your first CPRT study' was flagged as a dead-end but LINKS to /pricing (DemoCprtPage:290) -- refuted; (c) the demo 'read-only tooltip missing / numeric-0 friction' items mis-cited lines beyond the file's length and don't anchor (DemoCprtPage has no disabled buttons and no numeric inputs -- it is a read-only presentation page). Reporting these as dropped so the confirmed set stands on verified ground.",
    "Excel Standard (Sec 6) is N/A: VeritaOps ships no Excel export (the CPRT deliverable is the PDF). No em-dash / auto-filter / About-sheet / identity-header findings apply. The only forbidden-URL surface (the PDF + demo page) is clean; the demo CTA points to /pricing (canonical), no /#/ or localhost.",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaOps Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaOps - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-31)   [CPRT / CLSI GP11-A; module 17 of 17 -- cleanest of the sweep, 0 HIGH]"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 108
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
for sev in ("HIGH", "MEDIUM", "LOW-MED", "LOW"):
    print(f"  {sev}: {sum(1 for x in ROWS if x[4]==sev)}")
