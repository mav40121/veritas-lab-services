"""
VeritaPolicy Phase 2 QA/QC sweep.

Runs automated checks on all 96 policy templates:
1. Structural completeness (required JSON fields)
2. Accreditor reference leak (must be 0)
3. Copyright/copy-rule compliance (em-dash, banned terms)
4. Master-list coverage (every master-list policy_id has a template)
5. Policy name match (template name == master-list name)
6. CFR citation format sanity
7. Token placeholder presence
8. Required phrasing (medical director or designee; verification not validation)

Outputs:
  C:/Users/veril/OneDrive/Desktop/Lab/Verita Products/VeritaPolicy/Phase2_QA_Report.xlsx
"""

from __future__ import annotations
import json
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = Path(__file__).parent.parent / "server" / "policyTemplates" / "data"
MASTER_LIST_TS = Path(__file__).parent.parent / "server" / "veritapolicyMasterList.ts"
OUTPUT = Path(r"C:\Users\veril\OneDrive\Desktop\Lab\Verita Products\VeritaPolicy\Phase2_QA_Report.xlsx")

TEAL = "01696F"
FAIL_FILL = "FCE4EC"
WARN_FILL = "FFF9C4"
OK_FILL = "E8F5E9"

# Required fields in every template
REQUIRED_FIELDS = ["policy_id", "slug", "policy_name", "section", "purpose",
                   "cfr_text_blocks", "scope", "policy_statements",
                   "procedure_steps", "definitions"]

# Required tokens (at least these should appear)
REQUIRED_TOKENS = ["<<LAB_NAME>>"]
RECOMMENDED_TOKENS = ["<<CLIA_NUMBER>>", "<<DIRECTOR_NAME>>", "<<EFFECTIVE_DATE>>"]

# Accreditor leak patterns
ACCREDITOR = re.compile(
    r"\bAABB\b|\bTJC\b|\bCOLA\b|Joint Commission|College of American Pathologists|"
    r"\bGEN\.\d|\bTRM\.\d|\bCOM\.\d|\bDC\.\d{2}|\bQSA\.\d|\bLD\.\d|\bIM\.\d|"
    r"\bHR\.\d{2}|\bEC\.\d{2}|\bPER \d|\bPRE \d|\bLIS \d|\bORG \d"
)

# Em-dash detection (excluding code-comment use; templates are customer-facing)
EM_DASH = re.compile(r"—|—")

# Banned terms per CLAUDE.md §3
BANNED = re.compile(r"EP Evaluator|CAMLAB|LabVine Learning", re.I)

# "Medical director" without "or designee" is a soft fail per CLAUDE.md §5
MD_WITHOUT_DESIGNEE = re.compile(r"\bmedical director\b(?!\s+or designee)", re.I)
LAB_DIRECTOR_BARE = re.compile(r"\blaboratory director\b(?!\s+or designee|\s+qualifications|\s+is responsible)", re.I)

# Validation vs verification for lab activities
VALIDATION_FOR_LAB = re.compile(r"\bmethod validation\b|\bvalidation suite\b", re.I)

# CFR citation format sanity
CFR_FORMAT_GOOD = re.compile(r"\b(?:21|29|42|45|10)\s*CFR\s*\d+(\.\d+)?")

def load_master_list() -> list[dict]:
    text = MASTER_LIST_TS.read_text(encoding="utf-8")
    start = text.find("VERITAPOLICY_MASTER_LIST")
    assign_idx = text.find("= [", start)
    open_idx = assign_idx + 2
    depth = 0
    end_idx = -1
    for i in range(open_idx, len(text)):
        if text[i] == "[": depth += 1
        elif text[i] == "]":
            depth -= 1
            if depth == 0: end_idx = i; break
    return json.loads(text[open_idx:end_idx + 1])

def walk_strings(node, path=""):
    """Yield (path, string) for every string in a nested JSON structure."""
    if isinstance(node, str):
        yield (path, node)
    elif isinstance(node, list):
        for i, v in enumerate(node):
            yield from walk_strings(v, f"{path}[{i}]")
    elif isinstance(node, dict):
        for k, v in node.items():
            yield from walk_strings(v, f"{path}.{k}" if path else k)

def qa_one(path: Path, master_by_id: dict) -> list[dict]:
    """Return a list of finding dicts: {file, severity, category, location, detail}."""
    findings = []
    try:
        j = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [{"file": path.name, "severity": "FAIL", "category": "structure",
                 "location": "(file)", "detail": f"JSON parse error: {e}"}]

    pid = j.get("policy_id", "?")
    fname = path.name

    # 1. Required fields
    for f in REQUIRED_FIELDS:
        if f not in j:
            findings.append({"file": fname, "severity": "FAIL", "category": "structure",
                             "location": f, "detail": "missing required field"})
        elif f in ("cfr_text_blocks", "policy_statements", "procedure_steps") and not j.get(f):
            findings.append({"file": fname, "severity": "FAIL", "category": "structure",
                             "location": f, "detail": "empty required list"})

    # 2. Accreditor leak
    for loc, s in walk_strings(j):
        for m in ACCREDITOR.finditer(s):
            findings.append({"file": fname, "severity": "FAIL", "category": "copyright",
                             "location": loc, "detail": f"accreditor ref: {m.group(0)!r}"})

    # 3. Em-dash
    for loc, s in walk_strings(j):
        if EM_DASH.search(s):
            findings.append({"file": fname, "severity": "WARN", "category": "copy_rule",
                             "location": loc, "detail": "em-dash in customer-facing content"})

    # 4. Banned terms
    for loc, s in walk_strings(j):
        for m in BANNED.finditer(s):
            findings.append({"file": fname, "severity": "FAIL", "category": "copy_rule",
                             "location": loc, "detail": f"banned term: {m.group(0)!r}"})

    # 5. Medical director without "or designee" (skip CFR verbatim text blocks)
    for loc, s in walk_strings(j):
        if "cfr_text_blocks" in loc and "verbatim" in loc:
            continue
        if MD_WITHOUT_DESIGNEE.search(s):
            findings.append({"file": fname, "severity": "WARN", "category": "phrasing",
                             "location": loc, "detail": "'medical director' without 'or designee'"})

    # 6. Validation for lab activities
    for loc, s in walk_strings(j):
        if "cfr_text_blocks" in loc and "verbatim" in loc:
            continue
        if VALIDATION_FOR_LAB.search(s):
            findings.append({"file": fname, "severity": "WARN", "category": "phrasing",
                             "location": loc, "detail": "'validation' for lab activity (use 'verification')"})

    # 7. Token placeholders (purpose + scope + policy_statements should contain <<LAB_NAME>>)
    all_body = " ".join(s for _, s in walk_strings(j))
    for tok in REQUIRED_TOKENS:
        if tok not in all_body:
            findings.append({"file": fname, "severity": "WARN", "category": "tokens",
                             "location": "(body)", "detail": f"missing required token {tok}"})

    # 8. Master-list coverage
    if pid in master_by_id:
        master = master_by_id[pid]
        if j.get("policy_name", "").strip() != master["policy_name"].strip():
            findings.append({"file": fname, "severity": "WARN", "category": "name_match",
                             "location": "policy_name",
                             "detail": f"template name '{j.get('policy_name','')}' != master '{master['policy_name']}'"})
    else:
        findings.append({"file": fname, "severity": "FAIL", "category": "master_list",
                         "location": "policy_id",
                         "detail": f"policy_id {pid} not in master list"})

    # 9. CFR citation format
    for blk in j.get("cfr_text_blocks", []):
        cite = blk.get("citation", "")
        if not CFR_FORMAT_GOOD.search(cite):
            findings.append({"file": fname, "severity": "WARN", "category": "cfr_format",
                             "location": "cfr_text_blocks.citation",
                             "detail": f"unusual CFR cite format: {cite!r}"})

    return findings

def consolidation_analysis(master_rows: list[dict]) -> list[dict]:
    """Group master-list policies by overlapping CFR citations; identify
    clusters that could potentially be served by a single combined policy."""
    # Build (policy_id, list-of-CFR-sections) for each row
    pol_cfr = []
    for r in master_rows:
        cfrs = re.findall(r"42 CFR \d+(?:\.\d+(?:\([a-z0-9]+\))?)?|45 CFR \d+(?:\.\d+(?:\([a-z0-9]+\))?)?|21 CFR \d+(?:\.\d+(?:\([a-z0-9]+\))?)?|29 CFR \d+(?:\.\d+(?:\([a-z0-9]+\))?)?|10 CFR \d+(?:\.\d+(?:\([a-z0-9]+\))?)?",
                          r.get("cfr_citations", ""))
        pol_cfr.append({"policy_id": r["policy_id"], "policy_name": r["policy_name"],
                        "section": r["section"], "cfrs": set(cfrs)})

    candidates = []
    # Pair-wise overlap >= 3 CFR sections OR Jaccard >= 0.6
    for i, a in enumerate(pol_cfr):
        for b in pol_cfr[i + 1:]:
            overlap = a["cfrs"] & b["cfrs"]
            union = a["cfrs"] | b["cfrs"]
            if not union: continue
            jaccard = len(overlap) / len(union)
            if len(overlap) >= 3 or jaccard >= 0.6:
                candidates.append({
                    "policy_a_id": a["policy_id"],
                    "policy_a_name": a["policy_name"],
                    "policy_a_section": a["section"],
                    "policy_b_id": b["policy_id"],
                    "policy_b_name": b["policy_name"],
                    "policy_b_section": b["section"],
                    "shared_cfrs": "; ".join(sorted(overlap)),
                    "shared_count": len(overlap),
                    "jaccard": round(jaccard, 2),
                })
    candidates.sort(key=lambda c: (-c["shared_count"], -c["jaccard"]))
    return candidates

def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=TEAL)

def main():
    master_rows = load_master_list()
    master_by_id = {r["policy_id"]: r for r in master_rows}
    print(f"Loaded {len(master_rows)} master-list rows")

    files = sorted(DATA_DIR.glob("*.json"))
    print(f"Found {len(files)} template JSON files")

    all_findings = []
    by_file_status = {}
    for f in files:
        findings = qa_one(f, master_by_id)
        all_findings.extend(findings)
        fails = sum(1 for x in findings if x["severity"] == "FAIL")
        warns = sum(1 for x in findings if x["severity"] == "WARN")
        status = "FAIL" if fails else ("WARN" if warns else "OK")
        by_file_status[f.name] = {"status": status, "fails": fails, "warns": warns}

    # Check master-list templates that are MISSING
    missing = [r for r in master_rows if not any(j_pid == r["policy_id"]
                                                  for j_pid in [json.loads(f.read_text(encoding='utf-8')).get('policy_id') for f in files])]
    for r in missing:
        all_findings.append({"file": "(missing)", "severity": "FAIL", "category": "master_list_gap",
                             "location": f"policy_id {r['policy_id']}",
                             "detail": f"master-list policy '{r['policy_name']}' has no template"})

    consol = consolidation_analysis(master_rows)

    # Write xlsx
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "VeritaPolicy Phase 2 QA Report"
    ws["A1"].font = Font(size=16, bold=True, color=TEAL)
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    n_files = len(files)
    n_fail = sum(1 for v in by_file_status.values() if v["status"] == "FAIL")
    n_warn = sum(1 for v in by_file_status.values() if v["status"] == "WARN")
    n_ok = n_files - n_fail - n_warn
    n_findings = len(all_findings)
    n_fail_findings = sum(1 for x in all_findings if x["severity"] == "FAIL")
    n_warn_findings = sum(1 for x in all_findings if x["severity"] == "WARN")
    summary = [
        ("Template files", n_files),
        ("Master-list rows", len(master_rows)),
        ("Files with FAIL", n_fail),
        ("Files with WARN only", n_warn),
        ("Files OK", n_ok),
        ("Total findings", n_findings),
        ("  FAIL findings", n_fail_findings),
        ("  WARN findings", n_warn_findings),
        ("Consolidation candidate pairs", len(consol)),
    ]
    for i, (k, v) in enumerate(summary, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12

    # Findings sheet
    ws = wb.create_sheet("Findings")
    cols = ["file", "severity", "category", "location", "detail"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        style_header(cell)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}1"
    for i, fnd in enumerate(all_findings, start=2):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=i, column=j, value=fnd.get(c, ""))
        fill = FAIL_FILL if fnd["severity"] == "FAIL" else WARN_FILL
        for j in range(1, len(cols) + 1):
            ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=fill)
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + j)].width = min(max(len(c) + 3, 14), 80)

    # File status sheet
    ws = wb.create_sheet("File Status")
    cols = ["file", "status", "fails", "warns"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c); style_header(cell)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}1"
    for i, (fname, st) in enumerate(sorted(by_file_status.items()), start=2):
        ws.cell(row=i, column=1, value=fname)
        ws.cell(row=i, column=2, value=st["status"])
        ws.cell(row=i, column=3, value=st["fails"])
        ws.cell(row=i, column=4, value=st["warns"])
        fill = {"FAIL": FAIL_FILL, "WARN": WARN_FILL, "OK": OK_FILL}[st["status"]]
        for j in range(1, len(cols) + 1):
            ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=fill)
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + j)].width = 40 if j == 1 else 12

    # Consolidation candidates
    ws = wb.create_sheet("Consolidation Candidates")
    cols = ["policy_a_id", "policy_a_name", "policy_a_section",
            "policy_b_id", "policy_b_name", "policy_b_section",
            "shared_cfrs", "shared_count", "jaccard"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c); style_header(cell)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}1"
    for i, cand in enumerate(consol, start=2):
        for j, c in enumerate(cols, 1):
            ws.cell(row=i, column=j, value=cand.get(c, ""))
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + j)].width = 32 if "name" in c else 14

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nWrote {OUTPUT}")
    print(f"\nSummary: {n_files} files, {n_fail} FAIL, {n_warn} WARN, {n_ok} OK")
    print(f"Findings: {n_fail_findings} FAIL + {n_warn_findings} WARN = {n_findings}")
    print(f"Consolidation candidate pairs: {len(consol)}")

if __name__ == "__main__":
    main()
