#!/usr/bin/env python3
"""
import-lisa-milford-inventory.py

One-time import of Lisa Veri's Hematology Inventory Jan-Jun 2026 WorkDay.xlsx
into VeritaStock for her Milford lab (lab_id=4, UMass Memorial Health -
Milford Regional Medical Center).

Source spreadsheet columns and mapping:
  Description    -> item_name
  Number         -> catalog_number
  Cost           -> notes ("Unit cost: $X.XX")
  Department     -> department (Heme, Coag, Urines pass through)
  Vendor         -> vendor (with typo fixes: Hoogic->Hologic, "Werfen "->Werfen)
  Min. on Hand   -> reorder_point
  Inventory X/Y  -> most recent non-null becomes quantity_on_hand
  Order N        -> dropped (no time-series support in inventory_items schema)

ADMIN_SECRET is pulled from Railway env at run time per CLAUDE.md §12;
never written to a file, never echoed in chat.

Run with:
  python scripts/import-lisa-milford-inventory.py

Idempotency: this script is NOT idempotent. Re-running creates duplicate
inventory_items rows. Run once.
"""

import json
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

SOURCE_FILE = Path.home() / "Downloads" / "Hematology Inventory Jan-Jun 2026 WorkDay.xlsx"
TARGET_LAB_ID = 4
PROD_URL = "https://www.veritaslabservices.com/api/admin/import-inventory"

RAILWAY_TOKEN = "e1a9b567-9c6f-48da-bbba-9fef54a88c82"
PROJECT_ID = "29c628f1-7860-4fca-8fee-227159bb86e8"
ENV_ID = "cd669f7c-23f3-434c-895d-ca40ac504e91"
SERVICE_ID = "170f5560-8cf0-4341-9c87-294062ebedd1"

VENDOR_FIXES = {
    "Hoogic": "Hologic",
    "Werfen ": "Werfen",
}


def pull_admin_secret():
    payload = json.dumps({
        "query": (
            f'query {{ variables('
            f'projectId: "{PROJECT_ID}", '
            f'environmentId: "{ENV_ID}", '
            f'serviceId: "{SERVICE_ID}") }}'
        )
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://backboard.railway.com/graphql/v2",
        data=payload,
        headers={
            "Authorization": f"Bearer {RAILWAY_TOKEN}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())
    return data["data"]["variables"]["ADMIN_SECRET"]


def parse_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]

    inventory_col_indices = [
        i for i, h in enumerate(headers)
        if isinstance(h, str) and h.startswith("Inventory ")
    ]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        description = row[1]
        if not description:
            continue

        number = row[2]
        cost = row[3]
        dept = row[4]
        vendor = row[5]
        min_on_hand = row[6]

        quantity = 0
        for i in reversed(inventory_col_indices):
            v = row[i]
            if v is not None and isinstance(v, (int, float)):
                quantity = int(v)
                break

        if isinstance(vendor, str):
            vendor_clean = VENDOR_FIXES.get(vendor, vendor.strip())
        else:
            vendor_clean = None

        if isinstance(cost, (int, float)):
            notes = f"Unit cost: ${cost:.2f}"
        else:
            notes = None

        items.append({
            "item_name": str(description).strip(),
            "catalog_number": str(number) if number is not None else None,
            "department": str(dept).strip() if isinstance(dept, str) and dept.strip() else None,
            "vendor": vendor_clean if vendor_clean else None,
            "quantity_on_hand": quantity,
            "reorder_point": int(min_on_hand) if isinstance(min_on_hand, (int, float)) else 0,
            "notes": notes,
        })
    return items


def main():
    if not SOURCE_FILE.exists():
        print(f"ERROR: source file not found: {SOURCE_FILE}", file=sys.stderr)
        sys.exit(1)

    print("--- Lisa's Milford Hematology Inventory import ---")
    print(f"Source: {SOURCE_FILE}")
    print(f"Target: lab_id={TARGET_LAB_ID}")
    print()

    print("[1/3] Pulling ADMIN_SECRET from Railway env...")
    admin_secret = pull_admin_secret()
    print("      ok")

    print("[2/3] Parsing xlsx...")
    items = parse_xlsx(SOURCE_FILE)
    print(f"      parsed {len(items)} items")
    for it in items[:3]:
        dept = (it["department"] or "-")[:8]
        print(
            f"        - {it['item_name'][:38]:38s}  "
            f"dept={dept:8s} qty={it['quantity_on_hand']:4d} "
            f"reorder={it['reorder_point']:4d}"
        )
    print(f"        ... ({len(items) - 3} more)")

    print(f"[3/3] POSTing to {PROD_URL} ...")
    payload = json.dumps({
        "secret": admin_secret,
        "labId": TARGET_LAB_ID,
        "items": items,
    }).encode("utf-8")
    req = urllib.request.Request(
        PROD_URL,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read())
    except urllib.error.HTTPError as e:
        print(f"      HTTP {e.code}: {e.read().decode('utf-8', errors='replace')}")
        sys.exit(1)

    print()
    if result.get("ok"):
        print(f"PASS - inserted {result.get('inserted')} items into lab_id={TARGET_LAB_ID}")
        if result.get("errors"):
            print(f"  errors ({len(result['errors'])}):")
            for e in result["errors"][:10]:
                print(f"    - {e}")
        sys.exit(0)
    else:
        print(f"FAIL - {result}")
        sys.exit(1)


if __name__ == "__main__":
    main()
