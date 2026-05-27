"""
VeritaPolicy Phase 0 source-triage.

Walks Michael's P&P storehouse, extracts text from .docx / .doc / .pdf /
.xlsx, fuzzy-matches each document against the 96-row VERITAPOLICY_MASTER_LIST,
flags PII / employer markers, and emits a multi-sheet xlsx that drives
Phase 2 authoring.

Run:
  python scripts/veritapolicy_phase0_triage.py

Output:
  C:/Users/veril/OneDrive/Desktop/Lab/Verita Products/VeritaPolicy/Phase0_Triage.xlsx

Requires: python-docx, openpyxl, pdfplumber, pywin32 (for .doc files via
Word COM automation on Windows).
"""

from __future__ import annotations
import os
import re
import sys
import time
import json
import hashlib
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────────────────────
# Config
# ───────────────────────────────────────────────────────────────────────────
SOURCE_ROOT = Path(r"C:\Users\veril\OneDrive\Desktop\Lab\Documents to be moved\P and P")
OUTPUT_DIR  = Path(r"C:\Users\veril\OneDrive\Desktop\Lab\Verita Products\VeritaPolicy")
OUTPUT_XLSX = OUTPUT_DIR / "Phase0_Triage.xlsx"
MASTER_LIST_TS = Path(__file__).parent.parent / "server" / "veritapolicyMasterList.ts"

# Skip patterns: lock files, partials, hidden, common junk
SKIP_PREFIXES = ("~$", ".~$", "._")
SKIP_SUFFIXES = (".tmp", ".bak", ".old", ".partial.doc", ".partial.docx")

# PII / employer markers we want to flag for sanitization.
# Geographic / org names common to Michael's prior employer history.
PII_PATTERNS = [
    (re.compile(r"\bLMC\b"),                "LMC"),
    (re.compile(r"\bMass(achusetts)? General\b", re.I), "Mass General"),
    (re.compile(r"\bUMass\b"),              "UMass"),
    (re.compile(r"\bMilford\b", re.I),       "Milford"),
    (re.compile(r"\bSusan Giuliano\b", re.I), "Susan Giuliano"),
    (re.compile(r"\bHallmark Health\b", re.I), "Hallmark Health"),
    (re.compile(r"\bMelroseWakefield\b", re.I), "MelroseWakefield"),
    (re.compile(r"\bMelrose-Wakefield\b", re.I), "Melrose-Wakefield"),
    (re.compile(r"\bLawrence Memorial\b", re.I), "Lawrence Memorial"),
    (re.compile(r"\bMichael Veri\b", re.I), "Michael Veri"),
    (re.compile(r"\bLisa Veri\b", re.I), "Lisa Veri"),
    (re.compile(r"CLIA\s*[#:]?\s*\d{2}D\d{7}", re.I), "CLIA number"),
]

# ───────────────────────────────────────────────────────────────────────────
# Master list loader (parses the TS file's JSON-like array)
# ───────────────────────────────────────────────────────────────────────────
def load_master_list() -> list[dict]:
    """Parse veritapolicyMasterList.ts and pull out the 96 policy rows."""
    text = MASTER_LIST_TS.read_text(encoding="utf-8")
    # Find '= [' that initializes the array literal. The bare '[' after
    # 'VERITAPOLICY_MASTER_LIST' would match the TypeScript type annotation
    # `VeritaPolicyMasterRow[]`, so anchor on '= [' instead.
    start = text.find("VERITAPOLICY_MASTER_LIST")
    assign_idx = text.find("= [", start)
    if assign_idx < 0:
        raise RuntimeError("Could not find array literal in master list TS file")
    open_idx = assign_idx + 2  # the '['
    depth = 0
    end_idx = -1
    for i in range(open_idx, len(text)):
        if text[i] == "[":
            depth += 1
        elif text[i] == "]":
            depth -= 1
            if depth == 0:
                end_idx = i
                break
    array_text = text[open_idx:end_idx + 1]
    rows = json.loads(array_text)
    return rows

MASTER_ROWS = load_master_list()
print(f"Loaded {len(MASTER_ROWS)} master-list policies.")

# ───────────────────────────────────────────────────────────────────────────
# Text extractors
# ───────────────────────────────────────────────────────────────────────────
WORD_APP = None
def get_word():
    global WORD_APP
    if WORD_APP is None:
        import win32com.client
        WORD_APP = win32com.client.Dispatch("Word.Application")
        WORD_APP.Visible = False
        WORD_APP.DisplayAlerts = False
    return WORD_APP

def shutdown_word():
    global WORD_APP
    if WORD_APP is not None:
        try:
            WORD_APP.Quit()
        except Exception:
            pass
        WORD_APP = None

def extract_docx(path: Path) -> str:
    import docx
    try:
        d = docx.Document(str(path))
        parts = []
        for p in d.paragraphs:
            t = p.text.strip()
            if t:
                parts.append(t)
        return "\n".join(parts)
    except Exception as e:
        return f"[docx extract failed: {e}]"

def extract_doc(path: Path) -> str:
    """Use Word COM to read .doc; converts to text via Range.Text."""
    try:
        app = get_word()
        doc = app.Documents.Open(str(path), ReadOnly=True, AddToRecentFiles=False)
        try:
            text = doc.Range().Text
        finally:
            doc.Close(SaveChanges=False)
        return text or ""
    except Exception as e:
        return f"[doc extract failed: {e}]"

def extract_pdf(path: Path) -> str:
    import pdfplumber
    try:
        parts = []
        with pdfplumber.open(str(path)) as pdf:
            # Cap at first 6 pages to keep triage fast; policies are usually short.
            for page in pdf.pages[:6]:
                t = page.extract_text() or ""
                parts.append(t)
        return "\n".join(parts)
    except Exception as e:
        return f"[pdf extract failed: {e}]"

def extract_xlsx(path: Path) -> str:
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        parts = []
        for sheet in wb.worksheets:
            parts.append(f"## {sheet.title}")
            for row in sheet.iter_rows(values_only=True, max_row=80):
                row_text = " | ".join(str(c) if c is not None else "" for c in row)
                if row_text.strip(" |"):
                    parts.append(row_text)
        wb.close()
        return "\n".join(parts)
    except Exception as e:
        return f"[xlsx extract failed: {e}]"

def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".doc":
        return extract_doc(path)
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".xlsx":
        return extract_xlsx(path)
    return ""

# ───────────────────────────────────────────────────────────────────────────
# Classifier: fuzzy-match document text against master-list policy_name +
# description. Returns top-K candidate (policy_id, score) tuples.
# ───────────────────────────────────────────────────────────────────────────
STOPWORDS = {
    "policy","laboratory","specimen","procedure","procedures","laboratorys",
    "review","reviewed","approved","approval","testing","results","result",
    "ensure","include","including","with","this","that","from","into",
    "shall","must","when","each","every","other","other's","other's",
    "section","supervisor","staff","manager","director","department",
    "test","tests","tested","data","number","numbers","date","dates",
    "form","forms","line","lines","report","reports","record","records",
    "make","made","take","taken","used","uses","using","year","years",
    "documented","documentation","required","requirements","requirement",
    "performance","verification","performed","performing","verify",
    "lab","labs","quality","management","control","controls",
    "monitor","monitoring","ongoing","required","analyze","analyzed",
    "patient","patients",
}
def tokenize(s: str, drop_stopwords: bool = True) -> list[str]:
    toks = re.findall(r"[A-Za-z]{4,}", s.lower())
    if drop_stopwords:
        toks = [t for t in toks if t not in STOPWORDS]
    return toks

# Precompute IDF over master-list rows so common-across-all-policies tokens
# (lab vocabulary) don't dominate similarity scores. Doc frequency = how
# many master rows mention this token. IDF = log(N / df).
import math
MASTER_TOKEN_SETS: list[tuple[str, set[str], str]] = []
DF: dict[str, int] = defaultdict(int)
for r in MASTER_ROWS:
    blob = " ".join([r.get("policy_name", ""), r.get("description", ""), r.get("section", ""), r.get("notes", "")])
    toks = set(tokenize(blob))
    MASTER_TOKEN_SETS.append((r["policy_id"], toks, r["policy_name"]))
    for t in toks:
        DF[t] += 1
N_MASTER = len(MASTER_ROWS)
IDF: dict[str, float] = {t: math.log((N_MASTER + 1) / (df + 1)) + 1 for t, df in DF.items()}

def score_doc_vs_master(doc_tokens: set[str], master_tokens: set[str]) -> float:
    """Sum of IDF weights of overlap tokens, normalized so a perfect
    match against the master row returns ~1.0 and a single common-word
    hit returns near 0."""
    if not doc_tokens or not master_tokens:
        return 0.0
    overlap = doc_tokens & master_tokens
    if not overlap:
        return 0.0
    overlap_idf = sum(IDF.get(t, 0) for t in overlap)
    master_idf  = sum(IDF.get(t, 0) for t in master_tokens) or 1.0
    return overlap_idf / master_idf

def classify(doc_text: str, doc_filename: str, top_k: int = 3) -> list[tuple[str, str, float]]:
    """Return top-K (policy_id, policy_name, score) candidates."""
    # Weight filename heavier by repeating it. Cap doc body to keep tokenize fast.
    blob = (doc_filename + " ") * 6 + doc_text[:6000]
    doc_tokens = set(tokenize(blob))
    if not doc_tokens:
        return []
    scored = []
    for pid, mtok, pname in MASTER_TOKEN_SETS:
        s = score_doc_vs_master(doc_tokens, mtok)
        # Filename contains the policy name (case-insens), strong bonus.
        if pname and len(pname) > 10 and pname.lower() in doc_filename.lower():
            s += 0.6
        # SequenceMatcher on filename vs policy name as a secondary signal.
        if pname:
            sim = SequenceMatcher(None, doc_filename.lower(), pname.lower()).ratio()
            if sim > 0.55:
                s += 0.3 * sim
        scored.append((pid, pname, s))
    scored.sort(key=lambda x: x[2], reverse=True)
    return scored[:top_k]

# Strip XML-illegal control chars (Word COM can emit \x00..\x08, \x0b, \x0c,
# \x0e..\x1f). Keep \t \n \r.
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def sanitize_for_xlsx(v):
    if isinstance(v, str):
        v = _ILLEGAL_XLSX.sub(" ", v)
        # openpyxl also caps at 32767 chars per cell.
        if len(v) > 32000:
            v = v[:32000] + "...[truncated]"
    return v

# ───────────────────────────────────────────────────────────────────────────
# PII / employer detection
# ───────────────────────────────────────────────────────────────────────────
def detect_pii(text: str) -> list[str]:
    hits = []
    for pat, label in PII_PATTERNS:
        if pat.search(text or ""):
            hits.append(label)
    return sorted(set(hits))

# ───────────────────────────────────────────────────────────────────────────
# Triage classification: keep / dedupe / skip
# ───────────────────────────────────────────────────────────────────────────
def classify_row(path: Path, text: str, top_match_score: float) -> tuple[str, str]:
    """Return (status, reason). status in {keep, review, skip}."""
    name = path.name
    folder = str(path.parent.relative_to(SOURCE_ROOT))
    if name.startswith(SKIP_PREFIXES) or name.endswith(SKIP_SUFFIXES):
        return ("skip", "temp/lock file")
    if "Need to Complete" in folder:
        return ("skip", "incomplete draft folder")
    if not text or text.startswith("["):
        return ("skip", "no text extracted")
    word_count = len(text.split())
    if word_count < 80:
        return ("skip", f"too short ({word_count} words)")
    if top_match_score >= 0.22:
        return ("keep", f"strong master-list match (score={top_match_score:.3f})")
    if top_match_score >= 0.08:
        return ("review", f"weak master-list match (score={top_match_score:.3f})")
    return ("review", "no clear master-list match; manual review needed")

# ───────────────────────────────────────────────────────────────────────────
# Walk + process
# ───────────────────────────────────────────────────────────────────────────
def walk_source():
    out = []
    t0 = time.time()
    n = 0
    for path in sorted(SOURCE_ROOT.rglob("*")):
        if not path.is_file():
            continue
        n += 1
        ext = path.suffix.lower()
        size = path.stat().st_size
        relpath = path.relative_to(SOURCE_ROOT).as_posix()
        # Skip patterns
        if path.name.startswith(SKIP_PREFIXES) or path.name.endswith(SKIP_SUFFIXES):
            out.append({
                "relpath": relpath, "filename": path.name, "ext": ext, "size_kb": round(size / 1024, 1),
                "status": "skip", "reason": "temp/lock file",
                "top1_id": "", "top1_name": "", "top1_score": 0,
                "top2_id": "", "top2_name": "", "top2_score": 0,
                "top3_id": "", "top3_name": "", "top3_score": 0,
                "pii_markers": "", "word_count": 0, "preview": "",
            })
            continue
        if ext not in (".docx", ".doc", ".pdf", ".xlsx"):
            out.append({
                "relpath": relpath, "filename": path.name, "ext": ext, "size_kb": round(size / 1024, 1),
                "status": "skip", "reason": f"unsupported ext {ext}",
                "top1_id": "", "top1_name": "", "top1_score": 0,
                "top2_id": "", "top2_name": "", "top2_score": 0,
                "top3_id": "", "top3_name": "", "top3_score": 0,
                "pii_markers": "", "word_count": 0, "preview": "",
            })
            continue
        text = extract_text(path)
        word_count = len(text.split())
        candidates = classify(text, path.name)
        c1 = candidates[0] if len(candidates) > 0 else ("", "", 0.0)
        c2 = candidates[1] if len(candidates) > 1 else ("", "", 0.0)
        c3 = candidates[2] if len(candidates) > 2 else ("", "", 0.0)
        status, reason = classify_row(path, text, c1[2])
        pii = detect_pii(text)
        preview = (text or "")[:240].replace("\n", " | ").strip()
        out.append({
            "relpath": relpath, "filename": path.name, "ext": ext, "size_kb": round(size / 1024, 1),
            "status": status, "reason": reason,
            "top1_id": c1[0], "top1_name": c1[1], "top1_score": round(c1[2], 3),
            "top2_id": c2[0], "top2_name": c2[1], "top2_score": round(c2[2], 3),
            "top3_id": c3[0], "top3_name": c3[1], "top3_score": round(c3[2], 3),
            "pii_markers": ", ".join(pii),
            "word_count": word_count,
            "preview": preview,
        })
        if n % 25 == 0:
            print(f"  [{n}] processed: {relpath[:80]} ({time.time() - t0:.1f}s elapsed)")
    print(f"Total files walked: {n}. Elapsed: {time.time() - t0:.1f}s")
    return out

# ───────────────────────────────────────────────────────────────────────────
# Coverage rollup: which policy_ids have at least one strong match?
# ───────────────────────────────────────────────────────────────────────────
def coverage_rollup(rows: list[dict]) -> list[dict]:
    by_pid: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        if r["status"] == "keep" and r["top1_id"]:
            by_pid[r["top1_id"]].append(r)
        if r["status"] == "review" and r["top1_id"] and r["top1_score"] >= 0.04:
            by_pid[r["top1_id"]].append(r)
    out = []
    for m in MASTER_ROWS:
        pid = m["policy_id"]
        matches = by_pid.get(pid, [])
        out.append({
            "policy_id": pid,
            "policy_name": m["policy_name"],
            "section": m.get("section", ""),
            "candidate_count": len(matches),
            "best_score": max((mm["top1_score"] for mm in matches), default=0),
            "candidates": "; ".join(f"{mm['filename']} ({mm['top1_score']:.2f})" for mm in matches[:5]),
            "gap": "GAP" if not matches else "",
        })
    return out

# ───────────────────────────────────────────────────────────────────────────
# Excel writer
# ───────────────────────────────────────────────────────────────────────────
TEAL = "01696F"
ALT  = "EBF3F8"
GAP_FILL = "FCE4EC"
KEEP_FILL = "E8F5E9"
REVIEW_FILL = "FFF9C4"
SKIP_FILL = "EEEEEE"

def style_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def write_sheet(wb, name, headers, rows, freeze="B2", row_color_key=None):
    ws = wb.create_sheet(name)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    for i, r in enumerate(rows, 2):
        for j, h in enumerate(headers, 1):
            v = sanitize_for_xlsx(r.get(h, ""))
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Calibri", size=10, color="28251D")
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if row_color_key:
            color = row_color_key(r)
            if color:
                for j in range(1, len(headers) + 1):
                    ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[i].height = 30
    # auto-width with a cap
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(h) + 3, 14), 60)
    return ws

def main():
    print(f"Walking {SOURCE_ROOT}")
    rows = walk_source()
    shutdown_word()
    cov = coverage_rollup(rows)
    keepers = [r for r in rows if r["status"] == "keep"]
    review  = [r for r in rows if r["status"] == "review"]
    skipped = [r for r in rows if r["status"] == "skip"]
    summary = {
        "Total files walked": len(rows),
        "Keepers (strong master-list match)": len(keepers),
        "Needs Manual Review": len(review),
        "Skipped": len(skipped),
        "Master-list policies covered": sum(1 for c in cov if c["candidate_count"] > 0),
        "Master-list policies gap": sum(1 for c in cov if c["candidate_count"] == 0),
        "Files flagged with PII markers": sum(1 for r in rows if r["pii_markers"]),
    }
    print("Summary:", summary)

    wb = openpyxl.Workbook()
    # Strip default sheet
    wb.remove(wb.active)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws["A1"] = "VeritaPolicy Phase 0 Triage"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=TEAL)
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    for i, (k, v) in enumerate(summary.items(), start=4):
        ws.cell(row=i, column=1, value=k).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=i, column=2, value=v).font = Font(name="Calibri", size=11)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18

    cols = ["relpath", "filename", "ext", "size_kb", "word_count", "status", "reason",
            "top1_id", "top1_name", "top1_score",
            "top2_id", "top2_name", "top2_score",
            "top3_id", "top3_name", "top3_score",
            "pii_markers", "preview"]

    def color_for(r):
        return {"keep": KEEP_FILL, "review": REVIEW_FILL, "skip": SKIP_FILL}.get(r["status"])

    write_sheet(wb, "All Files", cols, rows, row_color_key=color_for)
    write_sheet(wb, "Recommended Keepers", cols, keepers, row_color_key=color_for)
    write_sheet(wb, "Needs Manual Review", cols, review, row_color_key=color_for)
    write_sheet(wb, "Skip List", cols, skipped, row_color_key=color_for)

    cov_cols = ["policy_id", "policy_name", "section", "candidate_count", "best_score", "candidates", "gap"]
    def gap_color(r):
        return GAP_FILL if r["gap"] == "GAP" else None
    write_sheet(wb, "Master-List Coverage", cov_cols, cov, row_color_key=gap_color)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
