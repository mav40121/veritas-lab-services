#!/usr/bin/env python3
"""
build_parking_lot_xlsx.py

Parses PARKING_LOT.md and produces a styled PARKING_LOT.xlsx alongside it.
One sheet "Parking Lot" with columns: Section, ID, Title, Status, Effort,
Importance, Description, Source, Notes.

STANDING RULE (memory feedback_parking_lot_xlsx_sync): every time
PARKING_LOT.md is edited, re-run this script so the Excel mirror stays
in sync. Commit both files together.

Run: python scripts/build_parking_lot_xlsx.py
Output: PARKING_LOT.xlsx at the repo root (next to PARKING_LOT.md).
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_MD = REPO_ROOT / "PARKING_LOT.md"
OUTPUT_XLSX = REPO_ROOT / "PARKING_LOT.xlsx"

# CLAUDE.md §6 brand colors.
TEAL = "01696F"
TEAL_TINT = "E6F2F2"
ALT_ROW = "EBF3F8"
TEXT = "28251D"
BORDER = "D0D0D0"
WHITE = "FFFFFF"


def parse_parking_lot(md_path: Path) -> list[dict]:
    """Return one dict per entry across OPEN / CLOSED / RETIRED sections."""
    text = md_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    section = None  # current section: "OPEN", "CLOSED", "RETIRED"
    entries: list[dict] = []
    current: dict | None = None
    buffer: list[str] = []

    def flush():
        nonlocal current, buffer
        if current is None:
            return
        body = "\n".join(buffer).strip()
        # Pull bold-labeled fields out of the body.
        fields = parse_fields(body)
        # "What" / "What:" or unlabeled lead paragraph becomes description.
        description = fields.get("What") or fields.get("STATUS") or first_paragraph(body)
        current.update({
            "Status": fields.get("Status", ""),
            "Effort": fields.get("Effort", ""),
            "Importance": fields.get("Importance", ""),
            "Description": description,
            "Source": fields.get("Source", ""),
            "Notes": fields.get("Pre- vs post-COLA", "") or fields.get("Fix shape", ""),
        })
        entries.append(current)
        current = None
        buffer = []

    for line in lines:
        if line.strip() == "## OPEN":
            section = "OPEN"
            continue
        if line.strip().startswith("## CLOSED"):
            section = "CLOSED"
            continue
        if line.strip().startswith("## NOT CARRIED OVER"):
            section = "RETIRED"
            continue
        m = re.match(r"^###\s+([A-Z]?\d+)\.\s+(.+)$", line)
        if m:
            flush()
            current = {
                "Section": section or "?",
                "ID": m.group(1),
                "Title": m.group(2).strip(),
            }
            buffer = []
            continue
        if current is not None:
            buffer.append(line)

    flush()
    return entries


def parse_fields(body: str) -> dict:
    """Extract **Label:** value pairs from a body. Values can span multiple lines
    until the next **Label:** or end of body."""
    out: dict[str, str] = {}
    # Match **Label:** at start of a line (after optional whitespace).
    pattern = re.compile(r"^\s*\*\*([^*][^*]*?):\*\*\s*(.*)$", re.MULTILINE)
    matches = list(pattern.finditer(body))
    for i, m in enumerate(matches):
        label = m.group(1).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
        value = (m.group(2) + "\n" + body[start:end]).strip()
        # Collapse any leftover **bold** wrappers in value for cleanliness.
        out[label] = value
    return out


def first_paragraph(body: str) -> str:
    """Fallback when no **What:** field is present."""
    for para in body.split("\n\n"):
        if para.strip():
            return para.strip()
    return ""


def shorten(text: str, max_chars: int = 800) -> str:
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3].rstrip() + "..."


def write_xlsx(entries: list[dict], out_path: Path) -> None:
    """One sheet per section (OPEN / CLOSED / RETIRED). OPEN opens first so
    triage view is the default when Michael opens the file."""
    wb = Workbook()
    # Drop the default empty sheet; we create our own named sheets below.
    default_sheet = wb.active
    wb.remove(default_sheet)

    by_section: dict[str, list[dict]] = {}
    for e in entries:
        by_section.setdefault(e.get("Section", "?"), []).append(e)

    # Stable ordering: OPEN first (active triage), then CLOSED audit trail,
    # then RETIRED rejections. Any unexpected section appended last.
    preferred = ["OPEN", "CLOSED", "RETIRED"]
    extras = [s for s in by_section.keys() if s not in preferred]
    sheet_order = [s for s in preferred if s in by_section] + extras

    for section_name in sheet_order:
        write_section_sheet(wb, section_name, by_section[section_name])

    # Set the OPEN sheet as the active tab on file open.
    if "OPEN" in wb.sheetnames:
        wb.active = wb.sheetnames.index("OPEN")

    wb.save(out_path)


def write_section_sheet(wb: Workbook, section_name: str, entries: list[dict]) -> None:
    """Write a single section's entries to its own sheet with shared styling."""
    ws = wb.create_sheet(title=section_name)

    # Section column dropped: each sheet IS the section.
    columns = [
        ("ID", 8),
        ("Title", 60),
        ("Status", 28),
        ("Effort", 14),
        ("Importance", 14),
        ("Description", 80),
        ("Source", 40),
        ("Notes", 50),
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=TEAL)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=10, color=TEXT)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, e in enumerate(entries, start=2):
        values = [
            e.get("ID", ""),
            shorten(e.get("Title", ""), 200),
            shorten(strip_md(e.get("Status", "")), 200),
            shorten(strip_md(e.get("Effort", "")), 80),
            shorten(strip_md(e.get("Importance", "")), 80),
            shorten(strip_md(e.get("Description", "")), 1200),
            shorten(strip_md(e.get("Source", "")), 400),
            shorten(strip_md(e.get("Notes", "")), 400),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=ALT_ROW)
        ws.row_dimensions[row_idx].height = 90

    ws.freeze_panes = "B2"
    if entries:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(entries) + 1}"


def strip_md(text: str) -> str:
    """Strip markdown formatting for cleaner cell content."""
    if not text:
        return ""
    # Drop **bold** markers.
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    # Drop inline backticks.
    text = text.replace("`", "")
    # Collapse multiple spaces but preserve line breaks.
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def main():
    if not SOURCE_MD.exists():
        print(f"ERROR: {SOURCE_MD} not found", file=sys.stderr)
        sys.exit(1)
    entries = parse_parking_lot(SOURCE_MD)
    sections = {}
    for e in entries:
        sections.setdefault(e.get("Section", "?"), []).append(e)
    print(f"Parsed {len(entries)} entries:")
    for s, lst in sections.items():
        print(f"  {s:8s}  {len(lst)} items")
    write_xlsx(entries, OUTPUT_XLSX)
    size = OUTPUT_XLSX.stat().st_size
    print(f"Wrote {OUTPUT_XLSX} ({size:,} bytes)")


if __name__ == "__main__":
    main()
