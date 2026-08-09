#!/usr/bin/env python3
"""Build the VeritaResponse module audit scorecard (internal deliverable for Michael).

4-lens read-only audit (reliability/multi-lab, export/PDF vs Sec 5/6, compliance-copy/
data-truth, UX), each finding independently verified by the main agent against current
code. Main-agent recalibrations vs the sub-agents:
  - Lens 1 called the PDF lab-identity defect a multi-lab-only MED ("owner's default lab
    stamped"). Lens 2 proved it is WORSE and always-on: getUserById is a Drizzle TYPED
    select over the users schema object, which declares no lab_name/clia_number, so the
    builder's user.lab_name / user.clia_number are BOTH undefined on every render -> the
    PDF prints the logged-in PERSON's name and always "CLIA: Not on file", single-lab and
    multi-lab alike. Verified: 11 working callsites (VeritaScan 15359, VeritaComp 21370,
    ...) read identity via raw `SELECT clia_number, clia_lab_name FROM users`; only the 5
    VeritaResponse PDF routes use the typed getUserById. Elevated to HIGH (#1).
  - Lens 4 called the finding-detail "not found on 500" a HIGH; recalibrated to MED (#4):
    it renders an error-ish state, not a false-CLEAN one, so it is less dangerous than the
    list error-as-empty (#2) which paints "zero findings / nothing overdue".
  - Confirmed HIGH: the always-on PDF identity break (#1), the list error-as-empty (#2),
    and the create silent-success (#3).
  - CLEAN confirmed: NO cross-lab IDOR (userCanAccessFinding is a real per-lab membership
    guard; GET list is lab_id-scoped) -- unlike the sibling VeritaTrack /worklist, this
    module has no security HIGH; every due-date cadence NUMBER is regulatorily correct;
    zero em-dashes in any public copy or rendered PDF; the POC editor + completeness UX is
    among the strongest in the suite.

Usage:  python scripts/build_veritaresponse_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaResponse_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaResponse_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 20)]

ROWS = [
    (1, "Every generated Plan-of-Correction PDF (CMS-2567, CAP, TJC, COLA, AABB) prints the logged-in PERSON's name where the lab name belongs and ALWAYS shows 'CLIA: Not on file', even when the lab has a CLIA on record. These are compliance documents submitted to CMS / CAP / TJC with the wrong facility identity and no CLIA.",
        "PDF/Data-truth", "server/pdfReport.ts:5744-5746 (+ class: 5931,6119,6309,6498) fed by routes.ts:18647,18708,18766,18823,18882",
        "HIGH",
        "The builders read identity as `user?.lab_name || user?.name` and `user?.clia_number || user?.cliaNumber`, but the route feeds them `storage.getUserById(dataUserId)`, a Drizzle TYPED select `db.select().from(users)` that returns ONLY the columns declared in the users schema object (schema.ts:6-19: id,email,name,plan,...). That object has NO lab_name/clia_number (those are on the LABS table, schema.ts:92-93), so user.lab_name / user.clia_number / user.cliaNumber are ALL undefined -> labName falls through to user.name (the person) and clia falls to the 'Not on file' string. Always-on, single-lab too. Proof it is unique to this module: 11 other PDF routes (VeritaScan 15359, VeritaComp 21370, CUMSUM 21531, ...) read identity via raw `SELECT clia_number, clia_lab_name FROM users`; only VeritaResponse uses the typed getUserById. Visible contradiction: licenseCtxFromReq resolves the ACTIVE lab for the footer license band, so the PDF footer shows the correct lab while the body header shows the person + no CLIA.",
        "In each of the 5 PDF routes, resolve identity from the finding's lab (labs row via resolveActiveLabForRequest, or finding.lab_id) OR use raw `SELECT clia_number, clia_lab_name FROM users`, and fix the builder key to clia_lab_name. Fix first, ahead of everything else.",
        "OPEN - HIGH (compliance-doc identity)"),
    (2, "A failed findings load renders 'No findings yet' -- a 500/403/network blip is indistinguishable from an empty board, and the summary tiles read '0 Overdue / 0 due within 7 days'. For a tool whose whole job is not missing accreditor deadlines, this hides real overdue deficiencies.",
        "Reliability/UX", "client/src/pages/VeritaResponseAppPage.tsx:123-134 (fetchData) + 326 (empty state) + 214-225 (counts)",
        "HIGH",
        "fetchData does `const res = await fetch(...); const data = await res.json(); setFindings(Array.isArray(data)?data:[])` with NO res.ok check and a `catch { // silent fail }`. A 403/500 error body (or HTML) is non-array -> findings=[] -> the table renders 'No findings yet. Click New Finding to record one.' and openCount/overdueCount computed from [] show 0. Same error-as-empty class fixed in VeritaTrack/VeritaLab/VeritaScan this sweep. (The detail page's LinkagePanel + vcLink share the shape at lower stakes -- #18.)",
        "throw / set an error flag on !res.ok, and render a distinct error card with Retry (mirror VeritaTrackAppPage 1080-1092), separate from the empty state.",
        "OPEN - HIGH, clean fix ready"),
    (3, "New Finding silently 'succeeds' on a failed POST: the dialog clears + closes exactly as on success, so the director believes a finding and its due-date clock were logged when they were not. If the backend is down, the refetch also fails (#2) and paints 'No findings yet', fully masking the loss.",
        "Reliability/UX", "client/src/pages/VeritaResponseAppPage.tsx:144-174 (handleCreate)",
        "HIGH",
        "handleCreate does `await fetch(POST /findings, ...)` with NO res.ok check, then unconditionally clears every field, setShowCreate(false), and awaits fetchData(). A 400 (validation) or 500 leaves the finding uncreated while the UI is identical to success. In a deadline tracker this is a missed-citation risk. The module imports no toast system, so there is no error surface at all.",
        "Guard on res.ok; on failure keep the dialog open and surface body.error (siblings use useToast variant:destructive).",
        "OPEN - HIGH, clean fix ready"),
    (4, "Opening a finding during a transient 500 renders 'This finding may have been deleted, or you do not have access to it' -- false and alarming for an active finding due in days, with no Retry except a full manual reload.",
        "Reliability/UX", "client/src/pages/VeritaResponseFindingPage.tsx:365-380 (fetchFinding) + 635-647 (not-found card)",
        "MEDIUM",
        "fetchFinding does `if (!res.ok) { setFinding(null); return; }` and `catch { setFinding(null); }`, so a 500/503/network drop lands in the same not-found card as a real 404/403. No status branch, no retry. Less dangerous than #2 (this shows an error-ish state, not a false-CLEAN one), so MED not HIGH, but it misattributes cause.",
        "Branch on res.status: distinct 'couldn't load, try again' error card with Retry for 5xx/network; reserve the not-found copy for 404/403.",
        "OPEN - MED, clean fix ready"),
    (5, "The running footer on all 5 POC PDFs is missing the brand line and 'Page X of Y': a multi-page Plan of Correction has no page numbers and no per-page 'VeritaAssure | VeritaResponse | Confidential' footer. This is the exact bug already found and fixed for VeritaPolicy.",
        "PDF/Sec 5", "server/pdfReport.ts:5897,6084,6276,6462,6657 (all pass \"\" as baseFooter)",
        "MEDIUM",
        "Each generator calls `applyLicenseToPuppeteer(html, \"\", licenseCtx)` with an EMPTY baseFooter, so licenseAugmentedFooterTemplate returns only the license band (no brand line, no <span class=pageNumber>/totalPages). CLAUDE.md Sec 5 requires the footer on every page. The in-body .footer-note appears once at the end, not as a running footer. VeritaPolicy fixed the identical bug (see comment pdfReport.ts:5680-5683; it now passes VERITAPOLICY_FOOTER_TEMPLATE); VeritaResponse never got the same treatment.",
        "Define a VERITARESPONSE_FOOTER_TEMPLATE (hyphen, not em-dash) with the brand line + Page X of Y and pass it as baseFooter in all 5 calls.",
        "OPEN - clean fix ready"),
    (6, "Wrong regulatory citation printed IN the CMS-2567 PDF (and in-app): 'State Operations Manual section 7314'. SOM Sec 7314 is 'Category 1 remedies', which has nothing to do with a Plan of Correction; the POC-content sections are Sec 7304.4 / Sec 7317.",
        "Data-truth/PDF", "server/pdfReport.ts:5843 + 5883 (rendered) + client VeritaResponseFindingPage.tsx:759 (+ comment 5720)",
        "MEDIUM",
        "The PDF column sub-head reads '(5 POC elements per SOM section 7314)' and the footer 'the five Plan of Correction elements required by the State Operations Manual section 7314 (42 CFR 493)'. Lens-3 pulled the CMS SOM Chapter 7 PDF: Sec 7314 is titled 'Special Procedures for ... Category 1 Remedies ...'; the POC sections are Sec 7304.4 ('plan of correction requirements') and Sec 7317 ('Acceptable Plan of Correction'). Note SOM Ch 7 is the SNF/NF chapter, so for a CLIA lab the cleaner authority is 42 CFR Part 493 Subpart R + the CMS-2567 form instructions. The five ELEMENTS themselves and the 10-day / 14-day windows are all correct -- only the section number is wrong.",
        "Drop the specific SOM number and cite '42 CFR 493 and the CMS-2567 form instructions' (safest), or Sec 7317 if a SOM cite is wanted. Confirm the target citation before shipping.",
        "OPEN - MED (data-truth)"),
    (7, "Findings list + detail go STALE after a lab switch: the fetch effect keys only on plan access, not the active lab, so switching /labs/10 -> /labs/11 keeps showing lab 10's findings and lab 10's overdue counts under the lab 11 URL until a manual Refresh.",
        "Multi-lab", "client/src/pages/VeritaResponseAppPage.tsx:136-142 (useEffect deps [hasPlanAccess]) ; mirror on FindingPage",
        "MEDIUM",
        "activeLabId comes from the URL (useActiveLabId). The lab switcher rewrites the labId param; wouter keeps the same component mounted, so findings state persists and findingsApi's recompute never triggers a refetch because the effect deps are [hasPlanAccess] only. Not a cross-tenant leak (a later action sends the correct X-Active-Lab-Id), but the displayed compliance data belongs to the wrong lab. Same class as task #107.",
        "Add activeLabId to the effect dependency array (and to the detail page's fetch effect).",
        "OPEN - clean fix ready"),
    (8, "Delete (both the list-row delete and the detail-page delete) treats a failed DELETE as success: the row refetch runs / the page navigates to the list unconditionally, with no pending state and no error surface. A 403 seat/plan gate or 404 reads as a successful delete.",
        "Reliability/UX", "client/src/pages/VeritaResponseAppPage.tsx:176-182 (list) + VeritaResponseFindingPage.tsx:439-446 (detail)",
        "MEDIUM",
        "List handleDelete: `await fetch(DELETE)` with no res.ok, then fetchData(). Detail handleDelete: `await fetch(DELETE)` then navigate(...) unconditionally. Both are correctly wrapped in ConfirmDialog (good), but ConfirmDialog calls onConfirm synchronously and closes, so no pending state is possible, and a failed delete is indistinguishable from a real one (the row simply reappears on refetch, or the user lands on the list assuming it worked).",
        "Check res.ok; only refetch/navigate on success, else surface a destructive toast; disable the confirm control while pending.",
        "OPEN - clean fix ready"),
    (9, "Effectiveness checkpoints (30/60/90) run on a native window.prompt + alert, and the 'Not effective' branch REOPENS the finding -- a material state change -- through that browser prompt with no styled confirmation. A failed checkpoint LOAD is swallowed to 'no checks yet'. window.prompt is also suppressed in sandboxed/embedded previews, dead-ending the flow.",
        "UX/Reliability", "client/src/pages/VeritaResponseFindingPage.tsx:239-276 (prompt 258, reopen 261, alerts 253/274, load catch 239-244)",
        "MEDIUM",
        "handleCheckpoint uses `window.prompt(status==='effective' ? ... : 'What recurred at ${days} days? (this reopens the finding)', '')` and `alert(data.error || 'Could not record checkpoint.')`. The 'not effective' path re-opens a closed finding via a native prompt. load() does `catch { /* leave as-is */ }`, so a failed GET leaves checks null and the panel shows 'Start 30/60/90 day monitoring' as if none exist (error-as-empty). Diverges from the module's own dialog/checklist UX and the sibling toast pattern.",
        "Replace prompt/alert with a small dialog + toast; give the checks query a real error state distinct from 'no checks yet'.",
        "OPEN"),
    (10, "The 'cross-links to your most recent VeritaCheck study for the cited standard' claim overstates the endpoint: it fires ONLY when standard_ref matches 493.xxx (returns null for CAP item IDs / TJC RFIs) and, when it does fire, returns the most-recent study of ANY specialty scoped by user_id -- not one matched to the cited standard, and a multi-lab owner can be shown another lab's study.",
        "Data-truth/Multi-lab", "server/routes.ts:18942-18948 (WHERE user_id, LIMIT 1) ; copy VeritaAssurePage.tsx:129, DemoLabPage.tsx:1387/1395",
        "MEDIUM",
        "The veritacheck-link handler: for a 493 match it does `SELECT ... FROM studies WHERE user_id = ? ORDER BY date DESC LIMIT 1` (dataUserId = owner). The code's own comment (18901-18911) says 'most-recent-study, not specialty-matched'. So (a) 'for the cited standard' implies a match the code does not perform, (b) the link is absent for CAP/TJC findings, and (c) it is user_id-scoped, so a multi-lab owner viewing a lab-B finding can be shown lab-A's study with a deep link that opens the other lab's study (same owner, so not a tenant leak, but wrong-lab evidence in a surveyor context).",
        "Reword marketing to 'surfaces your most recent VeritaCheck study when the finding cites a 42 CFR 493 standard'; scope the study lookup to finding.lab_id.",
        "OPEN"),
    (11, "COLA and AABB findings get a hard red 'Overdue / past their deadline' treatment that contradicts the module's OWN copy: COLA is labeled 'consultative; no hard deadline' and AABB 'lab CAPA timing varies', yet dueDateForFinding stamps a concrete due_date (COLA 30 / AABB 45) that turns the card red and counts it as overdue.",
        "Data-truth/UX", "client/src/pages/VeritaResponseAppPage.tsx:44-49 (deadlineNote) + 214-223 (overdue counting) + 301-302 ; server routes.ts:18387-18388 (offsets)",
        "MEDIUM",
        "offsets = { CAP:30, TJC:60, CMS:10, AABB:45, COLA:30, Other:30 }. The COLA deadlineNote says 'COLA is consultative; no hard deadline. Soft target of 30 days' and AABB's says the 45 days is the FDA reportable-event window ('lab CAPA timing varies'). But the list counts overdue/due-soon purely off due_date and warns 'N finding(s) past their deadline. Submit or request an extension.' So a soft/parallel-obligation clock is rendered identically to a hard CMS/CAP/TJC deadline -- an internal contradiction. (The numbers themselves are correct; the UI TREATMENT is the issue.)",
        "Treat COLA (and AABB accreditation) due_date as a soft/'target' state with distinct styling + wording, not the red 'Overdue / past their deadline' used for CMS/CAP/TJC.",
        "OPEN"),
    (12, "'Renders the federal CMS-2567 PDF' implies the official government form; the real CMS-2567 is issued BY the surveyor with deficiencies pre-printed. The artifact itself is honest ('FORM CMS-2567 (compatible)', 'mirrors the structure of CMS Form 2567') -- the marketing drops that qualifier.",
        "Data-truth/Copy", "client/src/pages/VeritaAssurePage.tsx:129 ; DemoLabPage.tsx:1387,1394 (vs pdfReport.ts:5809,5883)",
        "MEDIUM",
        "Marketing: 'Renders the federal CMS-2567 Plan of Correction PDF with all 5 POC elements labeled.' In-PDF: line 5809 prints 'FORM CMS-2567 (compatible)' and 5883 'This document mirrors the structure of CMS Form 2567.' Internal contradiction between the honest artifact and the marketing claim; the class Michael flags as an overstated capability.",
        "Reword to 'Renders a CMS-2567-compatible Plan of Correction PDF' to match the in-PDF framing.",
        "OPEN"),
    (13, "An 'Other'-accreditor finding, or a CAP/TJC/COLA/AABB finding on a lab whose accreditor flag was later dropped (legacy data the dropdown still supports), renders the full editor but NO PDF download card and no message saying why -- a dead-end for producing the response document.",
        "UX", "client/src/pages/VeritaResponseFindingPage.tsx:732,794,852,912,969 (cards gated on accreditor===X && labAllowedAccreditors.has(X))",
        "MEDIUM",
        "Each download card is gated `finding.accreditor === 'X' && labAllowedAccreditors.has('X')` (CMS gated on accreditor only). There is no card for accreditor==='Other', and a CAP finding on a now-unflagged lab renders no card at all, with no explanation. The per-accreditor gating is otherwise handled WELL (the module does NOT show all 5 buttons at once -- that lens item is satisfied); the gap is only the Other/legacy edges.",
        "For 'Other', show a neutral 'no standardized form for this accreditor' note; for the legacy-flag case, still render the matching card (or an explanatory line).",
        "OPEN"),
    (14, "The 'Laboratory Director or Designee Review' signature block is in natural document flow with no page-break-inside:avoid, so on a long deficiency narrative / long POC it can spill or orphan onto page 2 (Sec 5 requires the signature on page 1 of compliance documents).",
        "PDF/Sec 5", "server/pdfReport.ts:5857,6044,6236,6422,6617 (.director-block, all 5)",
        "LOW",
        "`.director-block { margin-top: 14px; ... }` with no break protection. For a short finding it lands on page 1 (rule satisfied); for a long two-column POC it can push past. It is NOT deliberately placed on its own page (the worst case Sec 5 targets), so this is a robustness gap, not a definite defect -- consistent with the natural-flow approach in sibling modules.",
        "Add `page-break-inside: avoid;` to .director-block so it never splits or orphans past page 1.",
        "OPEN"),
    (15, "CLIA-fallback wording drifts between the 5 builders (CMS uses a hyphen, the other four a comma); and when a real CLIA is present the header prints the bare number with no 'CLIA:' label prefix.",
        "PDF/Copy", "server/pdfReport.ts:5746 (CMS hyphen) vs 5933,6121,6311,6500 (comma)",
        "LOW",
        "CMS: 'CLIA: Not on file - enter in account settings'; CAP/TJC/COLA/AABB: 'CLIA: Not on file, enter in account settings'. Both correctly avoid the em-dash. Trivial cosmetic drift; standardize when fixing #1.",
        "Standardize the fallback string across all 5 builders (and prefix the real value with 'CLIA:').",
        "OPEN"),
    (16, "POC elements 2 and 3 are collapsed into one field ('2 + 3. Identify Others Affected and Prevent Recurrence'). CMS treats element 2 (how the facility identifies OTHERS potentially affected) and element 3 (systemic measures to prevent recurrence) as distinct; merging risks labs under-addressing element 2, a commonly cited POC weakness.",
        "PDF/Substance", "server/pdfReport.ts:5736,5850 ; client VeritaResponseFindingPage.tsx:737,1187",
        "LOW",
        "The label '2 + 3. Identify Others Affected and Prevent Recurrence' maps two distinct CMS POC elements to one field. The marketing 'all 5 POC elements labeled' stays technically true (all five ARE labeled), so this is substance polish, not a false claim.",
        "Consider splitting into two labeled sub-fields, or add helper text prompting the 'who else was affected' analysis explicitly.",
        "OPEN"),
    (17, "The CAP anchor is labeled 'Inspection date' but CAP's 30-day clock runs from receipt of the Inspection Summation Report. Anchoring on the earlier inspection date yields a due date at/before the true deadline (conservative/safe), so no missed-deadline risk -- just imprecise labeling.",
        "Data-truth/Copy", "client/src/pages/VeritaResponseAppPage.tsx:45 (anchorLabel 'Inspection date', +30)",
        "LOW",
        "CAP requires the POC response within 30 days of receiving the Inspection Summation Report, not the physical inspection date. The +30 offset is right; only the anchor LABEL is imprecise. Because the summation report post-dates the inspection, the computed due date is on/before the real deadline (safe direction).",
        "Rename the anchor to 'Summation report received' (still +30).",
        "OPEN"),
    (18, "Supplementary panels fail silently to error-as-empty: a failed links fetch reads as 'No links yet', and a failed VeritaCheck lookup makes the whole cross-link card vanish with no indication it errored.",
        "Reliability/UX", "client/src/pages/VeritaResponseFindingPage.tsx:154-166 (LinkagePanel catch) + 455-469 (vcLink)",
        "LOW",
        "LinkagePanel: `catch { /* leave null */ }` -> total===0 -> 'No links yet.' vcLink: on !res.ok/catch stays null and the entire VeritaCheck card is omitted. Low stakes (informational), same class as #2/#4.",
        "Add a lightweight 'couldn't load links' inline state.",
        "OPEN"),
    (19, "The list-row Delete is an icon-only button (Trash2) with no aria-label/title, so screen-reader users hear an unlabeled button on a destructive control. (The detail-page delete correctly has visible 'Delete' text.)",
        "UX/a11y", "client/src/pages/VeritaResponseAppPage.tsx:373-379",
        "LOW",
        "`<Button variant='ghost' ...><Trash2 size={13} /></Button>` with no aria-label. Minor a11y gap on a destructive action.",
        "Add aria-label='Delete finding' (and a title).",
        "OPEN"),
    (20, "The detail page never shows the cadence rule where the user acts (e.g. 'CAP: 30 days from inspection' lives only in the create dialog), and when due_date is null the due-date alert is fully hidden -- so a finding whose clock has not started shows no clock and no 'set an anchor date to start the clock' prompt.",
        "UX", "client/src/pages/VeritaResponseFindingPage.tsx:693 (alert only if due_date) + 722-726",
        "LOW",
        "The detail alert shows 'anchored on inspection date ...' but never the cadence note (deadlineNote exists only in the create dialog, AppPage 44-50). With due_date null the alert is hidden entirely, so a not-yet-started clock is invisible. Also minor: Refresh swaps the whole page for the full-screen spinner rather than an in-place indicator.",
        "Surface the accreditor deadlineNote on the detail alert; show a 'clock not started -- set the anchor date' hint when due_date is null.",
        "OPEN"),
    (21, "Stale comment + latent null-lab_id in the legacy POST: the comment claims 'reads key on user_id, so nothing hides', but GET /api/findings was since changed to WHERE lab_id. The INSERT omits lab_id and a swallowed follow-up UPDATE sets it to activeLab?.id ?? null; a null leaves the finding invisible to every list while still reachable by id.",
        "Reliability/Multi-lab", "server/routes.ts:18465-18472",
        "LOW",
        "The legacy create INSERTs without lab_id, then `try { UPDATE findings SET lab_id = ? } catch {}` with `activeLab?.id ?? null`. If resolveActiveLabForRequest returns null the finding is written lab_id NULL and never appears in the legacy or lab-scoped list (both filter WHERE lab_id). Low likelihood (resolver falls back to default lab), but the invariant the comment relies on no longer holds.",
        "Set lab_id in the initial INSERT (as the lab-scoped POST at 19142 already does); update the stale comment.",
        "OPEN"),
]

PASS_NOTES = [
    "NO cross-lab IDOR (the class that made VeritaTrack's /worklist a security HIGH is NOT present here). Every /api/findings/:id route -- GET :id, PUT, DELETE, all 5 PDFs, all completeness routes, veritacheck-link -- goes through userCanAccessFinding (routes.ts:11365-11377), which fetches by PK then requires row.user_id === dataUserId OR an ACTIVE lab_members row for row.lab_id (a real per-lab membership guard that fails CLOSED to 404). GET /api/findings is lab_id-scoped via resolveLegacyLabId; the /api/labs/:labId/findings* routes use labScopeMiddleware + WHERE id=? AND lab_id=?. No confirmed cross-tenant read/write/export leak. The multi-lab defects that DO exist (#1 identity, #7 staleness, #10 veritacheck-link) are same-owner wrong-lab issues, not tenant leaks.",
    "Every due-date cadence NUMBER is regulatorily correct (verified against primary sources): CMS-2567 = 10 calendar days from receipt (CMS SOM Sec 7317/7304.4); CAP = 30 days; TJC ESC = 60 days from posting of the survey findings report; AABB = 45 days framed as the FDA reportable-event window (21 CFR 606.171); COLA = 30-day soft target. The CMS-only 14-day public-release note (QSO-25-19-ALL, June 2025) is real and correctly attached to CMS only, NOT to the AO findings. The lone data-truth defect is the SOM SECTION NUMBER (#6), never the numbers or the anchors.",
    "Em-dash ban is CLEAN -- the usual top finding of this sweep is genuinely absent. Grep of the whole VeritaResponse PDF section and all public copy found ZERO em-dashes in rendered output (the only two are internal code comments at pdfReport.ts:6296 and FindingPage:340). TM not (R) throughout; module rendered as VeritaResponse(TM); no 'validate/validation' misuse (validate* identifiers are form-completeness checks); no EP Evaluator / CAMLAB / LabVine / dated-manual references; module count 'seventeen' correct on HomePage.",
    "PDF STRUCTURE is faithful to Sec 5 on the things that usually break: a single 'Laboratory Director or Designee Review' block with Accepted / Not accepted checkboxes + Print Name/Initials/Signature/Date, on the same page as the finding (the #1 defect is a broken IDENTITY read, not a layout or hardcode problem); the 5 POC elements are correctly labeled and map to a real CMS-2567 Plan-of-Correction structure; author metadata 'Perplexity Computer' on every generator; no URLs in the PDF (canonicalization N/A); no hardcoded lab name/CLIA and no stray Riverside/22D0999999. No Excel export exists in this module, so Sec 6 is N/A.",
    "The POC editor + completeness UX is among the STRONGEST in the suite: the 5 POC elements are discrete labeled fields with explanatory placeholders; a per-accreditor completeness checklist shows green/red per element client-side; download buttons are DISABLED until the minimum floor is met and show a 'Rendering...' pending state; per-accreditor gating means only the finding's own accreditor card renders (all-5-buttons anti-pattern avoided); the save bar has a real idle/saving/saved/error state machine (the one mutation done right); due-date math guards null + isNaN so there is no 'NaN days'/'Invalid Date'; both deletes are wrapped in ConfirmDialog; the plan gate uses the correct explicit-allowlist style. The systemic weakness is READ/WRITE FAILURE HANDLING (#2/#3/#4/#8/#9/#18), not the core workflow.",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaResponse Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaResponse - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-12)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 72
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
for sev in ("HIGH", "MEDIUM", "LOW"):
    print(f"  {sev}: {sum(1 for x in ROWS if x[4]==sev)}")
