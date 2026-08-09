#!/usr/bin/env python3
"""Build the VeritaQC module audit scorecard (internal deliverable for Michael).

4-lens read-only audit (reliability/multi-lab, export/PDF vs Sec5/6, compliance-copy/
data-truth, UX), each finding independently verified by the main agent against current
code. Main-agent recalibrations vs the sub-agents:
  - The Westgard back-dated-entry bug (QC1) is CONFIRMED by reading the evaluator:
    `i = sdis.length - 1` evaluates the last DATE-ordered result, but the just-inserted
    result only sorts last when its result_date is the latest, so a back-dated entry
    scores a different (in-control) point -> false accept + baseline poison. HIGH.
  - Two lenses independently flagged the daily-review green "all clear" on a failed
    load (QC2) with identical line numbers; confirmed HIGH (affirmative green success on
    a swallowed error is worse than a neutral empty state).
  - The wrong CFR (QC4) is eCFR-verified: 493.1256(d) is the daily-control/IQCP clause,
    not corrective action (493.1282); and it CONTRADICTS the module's own article/FAQ.
    Kept MED (wrong citation persisted in a finding; Michael's regulatory ruling), not
    HIGH, matching how the VeritaResponse SOM-7314 citation was scored.
  - CLEAN confirmed: multi-lab isolation is strong (every route uses req.scope.labId,
    no VeritaTrack /worklist gap); the monthly PDF does NOT have the VeritaResponse
    getUserById identity bug (raw `SELECT lab_name, clia_number FROM labs`) and has a
    real footer with Page X of Y; the client WRITE path checks res.ok everywhere; 6 of 7
    Westgard rule definitions are correct.

Usage:  python scripts/build_veritaqc_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaQC_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaQC_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 20)]

ROWS = [
    (1, "The Westgard evaluator scores the WRONG result on any back-dated or out-of-order QC entry: it always evaluates the latest-dated result, not the one just entered. Back-entering a missed out-of-control QC produces a false 'clean', and the flyer enters the accepted baseline and corrupts future SD.",
        "Reliability/Correctness", "server/routes.ts:2646-2667 (evaluateWestgardForLot); insert 2731-2737",
        "HIGH",
        "history is `ORDER BY result_date ASC, id ASC`, then `const i = sdis.length - 1` (2667) evaluates the LAST element = the latest-dated result. The just-inserted result (max id) only sorts last when its result_date is >= every existing date. result_date is a free field (String(result_date), 2735; no server validation). Back-enter a missed QC (yesterday's date) with a 1-3s flyer value: it sorts into the middle, i points at the latest in-control point (z~=0), NO rule fires, requires_corrective_action=false, and the flyer is stored accepted_for_reporting=1 (2732) -> false accept + baseline poison. Every rule (1-3s single-point through N-x/N-T windows) is anchored on the wrong point.",
        "Anchor to the actual candidate: `const i = ids.indexOf(newResultId); if (i < 0) return [];` so single-point + windows end at the entered result. Fix first. (Pair with QC7 date validation.)",
        "OPEN - HIGH (crown jewel)"),
    (2, "The Daily Review page renders an affirmative green 'all clear' (checkmark, 0 rejections, 0 missing corrective actions) when the QC load FAILS. A director can sign off a month believing QC was clean while the fetch actually errored.",
        "Reliability/UX", "client/src/pages/VeritaQCDailyReviewPage.tsx:163-172 (fetch) + 370-380 (render) + 312-316 (tiles)",
        "HIGH",
        "load() does `if (res.ok) { setResults(...) }` with NO else and a `catch { console.error }`. On any non-2xx (500/502/403) or network throw, results stays [], loading clears, and the render hits `groups.length === 0` -> a green CheckCircle2 with 'No results match these filters.' The summary tiles (totalRejections, missingCA) compute to 0 from the empty array. This is the recurring error-as-empty class, amplified here into an affirmative green success affordance on a QC surface. Same class fixed in VeritaLab/VeritaTrack/VeritaResponse this sweep.",
        "Add a loadError state; on !res.ok/catch render a distinct destructive error card with Retry (mirror VeritaLabAppPage:459-468), never the green empty state.",
        "OPEN - HIGH, clean fix ready"),
    (3, "The QC entry page's results table shows 'No results logged for this lot yet' on a failed fetch, and on a failed LOT SWITCH keeps the PREVIOUS lot's rows under the new lot's header (one lot's QC values mislabeled as another's).",
        "Reliability/UX", "client/src/pages/VeritaQCAppPage.tsx:172-189 (loadResults) + 673 (empty render)",
        "MEDIUM",
        "loadResults does `if (res.ok) { setResults(...) }`, no else, `catch { console.error }`, and never clears results before the new fetch. First-load failure -> 'No results logged for this lot yet' (hides real history). Switching lots when the new fetch fails leaves the prior lot's rows displayed under the now-selected lot's 'Recent results (last 20)' header. Same-tenant, but a QC data misattribution.",
        "Track a per-query error; clear results on lot change before the fetch; render a distinct error + Retry instead of the empty state.",
        "OPEN - clean fix ready"),
    (4, "The QC->VeritaResponse escalation persists the wrong CFR into the finding: standard_ref is hardcoded '42 CFR 493.1256(d)' and labeled 'the CLIA QC corrective-action requirement', but 493.1256(d) is the daily-control-count / equivalent-testing (IQCP) clause. Corrective action is 493.1282, which the module's OWN article + FAQ already cite.",
        "Data-truth", "server/routes.ts:3249 (INSERT) + comment 3201 + toast VeritaQCDailyReviewPage.tsx:125/450",
        "MEDIUM",
        "The INSERT hardcodes `'42 CFR 493.1256(d)'` as the finding standard_ref (3249) and the comment (3201) calls it 'the CLIA QC corrective-action requirement.' eCFR: 493.1256(d) requires performing control procedures / daily two-level controls / EQC -- NOT corrective action. The CLIA corrective-action standard is 493.1282 (493.1282(b)(2) on control failures). The module's public article (ArticleQCTestingIntoCompliancePage:85,130) and FAQ (faqContent:124) BOTH correctly cite 493.1282, so this persisted finding contradicts the app's own published citation and mis-frames the CAPA's regulatory basis on a surveyor-facing record. Michael's regulatory ruling; recommendation is well-supported.",
        "Change the escalation standard_ref to 42 CFR 493.1282 (optionally also cite 493.1256 for the control procedure breached); fix the comment + toast. Your call on the exact cite.",
        "OPEN - your regulatory ruling"),
    (5, "IDOR: a corrective action's qc_rule_violation_id is stored without any ownership check, and escalate-to-response reads it unscoped, so a user can bake ANOTHER lab's Westgard rule metadata (rule_code, severity, detail) into their own finding. Low blast radius (no PHI, no lot identity).",
        "Security/Multi-lab", "server/routes.ts:3169 (raw insert) + 3225 (unscoped read)",
        "MEDIUM",
        "POST /qc/corrective-actions validates qc_result_id against the lab (3157) but inserts `qc_rule_violation_id ? Number(...) : null` raw (3169). qc_rule_violations has NO lab_id column (db.ts:5038); it is only safe when reached via a lab-scoped qc_result_id. On escalate, `SELECT rule_code, severity, detail FROM qc_rule_violations WHERE id = ?` (3225, no lab scope) bakes those into the finding description/severity (3253-3254) and returns them (3275). A user can pass a foreign, guessable small-int violation id -> cross-tenant read of another lab's Westgard metadata. The ctx (analyte/lot/value) correctly comes from the caller's own result, so no lot-identity leak.",
        "In the CA insert, verify `SELECT 1 FROM qc_rule_violations WHERE id=? AND qc_result_id=?` before storing (else null it), so the violation must belong to the same lab-scoped result.",
        "OPEN - MED (security)"),
    (6, "R-4s fires on the range alone (any two consecutive single-level results spanning >4 SD), without the canonical requirement that one point exceed +2s and the other -2s. Canonical R-4s is a WITHIN-RUN across-level rule; this is an across-run single-level adaptation that over-flags valid runs.",
        "Data-truth/Correctness", "server/routes.ts:2682-2685; description pdfQCMonthly.ts:165",
        "MEDIUM",
        "`if (i >= 1 && Math.abs(z - sdis[i-1]) > 4)` -> R-4s rejection, detail 'range ...SD across zero'. It never verifies the two points straddle the mean (e.g. z_prev=-2.2, z=+1.9 spans 4.1>4 and fires a REJECTION even though +1.9 is inside +2s; canonical R-4s would not reject). Direction is non-dangerous (false reject, forces unwarranted rejection), but it deviates from Westgard/CLSI C24 (R-4s is within-run across the two levels of one run). The user-facing description honestly matches the code (no mislabel), so this is a definitional-fidelity call on the intended QC scheme.",
        "Either require sdis[i]>2 && sdis[i-1]<-2 (or vice-versa), or relabel as a single-level range/delta rule. Your ruling on the intended within-run vs single-level scheme.",
        "OPEN - your ruling"),
    (7, "POST /qc/results does not validate result_value finiteness or result_date format server-side (unlike the sibling control-lots POST). A direct API call with result_value='1e999'->Infinity or a non-ISO date is accepted, breaking the date-ordered evaluation (compounds QC1) and the PDF month window.",
        "Reliability", "server/routes.ts:2719 (presence-only guard) + 2734-2735 (Number/String unvalidated)",
        "MEDIUM",
        "The guard only checks presence, then inserts Number(result_value) and String(result_date). The control-lots POST enforces Number.isFinite (2816-2821); this handler doesn't. Infinity/NaN values and non-ISO dates ('04/15/2026') enter qc_results, sort wrong under ORDER BY result_date, and corrupt Westgard windows + the PDF's month-string comparison. Browser form guards it, but the API is the source of truth.",
        "Reject 400 unless Number.isFinite(Number(result_value)) and result_date matches ^\\d{4}-\\d{2}-\\d{2}$.",
        "OPEN - clean fix (pair with QC1)"),
    (8, "A control-lot load failure renders the 'No control lots yet for this lab. Add your first control lot.' onboarding state, so a transient error makes the tech think their lots vanished and invites re-creating existing lots.",
        "Reliability/UX", "client/src/pages/VeritaQCAppPage.tsx:152-170 (loadLots) + 498-512 (render)",
        "MEDIUM",
        "loadLots `catch { console.error }` with no error state; on failure lots stays [] -> the onboarding empty state with an 'Add your first control lot' CTA. A 409 would block a duplicate server-side, but the UX is alarming and masks the whole entry surface. Same error-as-empty class as QC2/QC3.",
        "Distinguish load-error from truly-empty; show error + Retry, keep the onboarding copy only for a confirmed 2xx-empty response.",
        "OPEN"),
    (9, "A QC result can be logged against a retired or on-hold control lot: the lot dropdown lists every status and Submit is not gated on lot status, polluting the QC record for a lot no longer in use.",
        "Data quality/UX", "client/src/pages/VeritaQCAppPage.tsx:534-539 (dropdown) + 660 (Submit)",
        "MEDIUM",
        "The lot <Select> maps lots.map(...) with no status filter (appends [status] but stays selectable); Submit is disabled only on submitting/isReadOnly. handleRetireLot slides off the retired lot (388-390) to prevent this, but the user can re-select it and log a result with no warning/disable.",
        "When selectedLot.status !== 'active', show an inline warning and disable Submit (or require explicit override).",
        "OPEN"),
    (10, "Expired control lots are not flagged: expiration renders as plain grey text with no comparison to today and no EXPIRED badge, and Submit stays enabled, so QC can be logged against a lot that expired last month with zero visual cue.",
        "Data quality/UX", "client/src/pages/VeritaQCAppPage.tsx:548-551",
        "MEDIUM",
        "`Exp: {selectedLot.expiration_date || 'n/a'}` -- no date comparison, no badge, Submit enabled regardless.",
        "Compare expiration_date to today; render a red EXPIRED badge and warn (or block) on Submit.",
        "OPEN"),
    (11, "Dead-end triage: the Daily Review flags 'missing corrective action' on a rejection but offers NO way to file one. A CA can only be filed via the forced modal at entry time, so any imported or interrupted rejection is stranded -- visible to the reviewer with no remediation path a surveyor then sees unresolved.",
        "UX/Workflow", "client/src/pages/VeritaQCDailyReviewPage.tsx:458-461 ('missing' badge, no action)",
        "MEDIUM",
        "The missing-CA row renders only a red 'missing' badge; there is no 'File corrective action' control on the page. CAs are created only through VeritaQCAppPage's entry-time modal; the 'Escalate to VeritaResponse' button appears only for results that ALREADY have a CA. The missing_ca status filter exists to triage these, but 'Open lot in entry page' only lets you log a NEW result, not attach a CA to the existing one.",
        "Add a 'File corrective action' action on missing-CA rows in the daily review that opens the CA modal bound to that qc_result_id.",
        "OPEN"),
    (12, "The monthly QC attestation can be filed without opening the PDF and with unresolved out-of-control runs, in one un-confirmed click; the past-attestations table does not show WHO signed.",
        "Data quality/UX", "client/src/pages/VeritaQCDailyReviewPage.tsx:551-553 (button) + 229-257 (handler) + 560-579 (table)",
        "MEDIUM",
        "The File-attestation button is enabled as soon as a lot is picked; handleFileAttestation posts attestation_acknowledged:true with no precondition, no check for outstanding missing-CA rejections in the period, and no ConfirmDialog, despite copy that says 'Review the PDF, then file.' The past-attestations table shows Period/Filed/Notes but not the signing user. The director's monthly sign-off is the compliance artifact; 11 sibling pages use ConfirmDialog, neither QC page does.",
        "Warn/confirm before filing when the lot+period has missing-CA rejections; surface the signing user in the past-attestations table.",
        "OPEN"),
    (13, "The corrective-action modal has no escape hatch if the CA save keeps failing: Esc/overlay/X are blocked while a result is pending and there is no Cancel button, so a tech whose CA save fails repeatedly (server down) is trapped, recoverable only by a hard reload they cannot know is safe.",
        "UX/Reliability", "client/src/pages/VeritaQCAppPage.tsx:732 (onOpenChange block) + 785-789 (single button)",
        "MEDIUM",
        "onOpenChange returns early while caForResultId is set (blocks dismissal); the footer has only 'File corrective action' (no Cancel, unlike the Add-Lot dialog at 899-905). The QC result itself already persisted (POST at 216 before the modal opened), so the forcing traps the user on a transient CA-save failure. The forcing intent is good; the no-exit-on-repeated-failure is the gap.",
        "After N failed CA saves, allow dismissal with a 'saved as missing corrective action, resolve from Daily Review' message (needs QC11's remediation path).",
        "OPEN"),
    (14, "The monthly QC review PDF sign-off is a 'MONTHLY REVIEW ATTESTATION' block (Reviewer/Title/Date/Acknowledged) rather than the Sec 5 canonical 'LABORATORY DIRECTOR OR DESIGNEE REVIEW' block with Accepted / Not accepted checkboxes and a blank Signature line. Defensible as a recurring-review shape, but it deviates from the house standard.",
        "PDF/Sec 5", "server/pdfQCMonthly.ts:232-256 (ackBox)",
        "MEDIUM",
        "Header is 'MONTHLY REVIEW ATTESTATION' (234); fields are Reviewer (pre-filled name)/Title/Date/Acknowledged (YES). No Accepted/Not accepted checkboxes, no explicit blank Signature line. Sec 5 specifies one 'LABORATORY DIRECTOR OR DESIGNEE REVIEW' block with Accepted/Not accepted + Print Name/Initials/Date. The block IS on page 1 (page-1 rule satisfied). The attestation shape ('I reviewed all runs; CAs documented; issues escalated') is arguably more appropriate for a recurring review than an accept/reject verdict.",
        "Add an explicit blank Signature line; either adopt the house Accepted/Not-accepted header or get your sign-off that the attestation shape is the intended QC-review deviation.",
        "OPEN - your ruling"),
    (15, "The Daily Review summary counters ('Rejection-rule fires in window', 'Missing corrective action') are computed from the server-FILTERED results, so applying an unrelated status filter silently changes the headline metric's meaning and undercounts.",
        "Data-truth", "client/src/pages/VeritaQCDailyReviewPage.tsx:312-316",
        "LOW",
        "totalRejections/missingCA derive from `results`, which the server already filtered by statusFilter. With statusFilter='missing_ca', 'Rejection fires' counts only rejections whose CA is unfiled -> undercount of the true window total.",
        "Compute the summary from an unfiltered (status=any) fetch, or relabel to reflect the active filter.",
        "OPEN"),
    (16, "A constant baseline (SD=0, e.g. an instrument reporting the same rounded integer) suppresses ALL Westgard rules, so a subsequent wild value is not flagged (false accept). Rare edge, dangerous direction.",
        "Reliability", "server/routes.ts:2661 (if (sd === 0) return [])",
        "LOW",
        "When the accepted baseline is all-identical, sd=0 and the evaluator returns [] with no evaluation. mfr_mean/mfr_sd exist on the lot but are never used by the evaluator as a fallback.",
        "When sd=0 but n>=threshold, fall back to the lot's mfr_sd or flag any nonzero deviation.",
        "OPEN"),
    (17, "The monthly PDF labels the reviewer role 'medical director or designee' where 'laboratory director or designee' is correct (QC review is a CLIA laboratory-director responsibility); the document contradicts its own narrative which says 'laboratory director'.",
        "PDF/Copy", "server/pdfQCMonthly.ts:240 + routes.ts:3132 + footer 283",
        "LOW",
        "ack-label 'Reviewer (medical director or designee)' (240) and reviewerTitle 'Medical director or designee' (3132), vs the narrative/attestation body which say 'laboratory director or designee'. The hard Sec 5 rule (never bare 'medical director') is satisfied ('or designee' present); this is precision/consistency.",
        "Standardize reviewer-role labels/titles to 'laboratory director or designee'.",
        "OPEN"),
    (18, "The monthly PDF pre-fills the reviewer name (logged-in user) under a hardcoded 'Medical director or designee' title, so if a bench tech generates the PDF to preview it prints '[Tech] / director' -- a soft misattribution on a compliance record.",
        "PDF", "server/routes.ts:3119 (reviewer name) + 3131-3132 (hardcoded title)",
        "LOW",
        "reviewerName = reviewer?.name (the downloader), reviewerTitle hardcoded to the director title regardless of the user's actual role. Corrected at wet-signing, but a misattribution vector on the generated form.",
        "Leave reviewer name/title blank for wet signature, or pull the actual title from the user record.",
        "OPEN"),
    (19, "The monthly QC review PDF cites no QC control-procedures regulation (42 CFR 493.1256); sibling compliance PDFs tie the document to its CFR basis.",
        "PDF/Copy", "server/pdfQCMonthly.ts:258-285 (no CFR anywhere)",
        "LOW",
        "The narrative + attestation cite Westgard/CLSI EP23-A but not the governing 42 CFR 493.1256 (control procedures). Not required by Sec 5 (bold-CFR is VeritaCheck-specific), so LOW.",
        "Optionally add a footnote referencing 42 CFR 493.1256 (control procedures) as the review basis.",
        "OPEN"),
    (20, "Westgard bias/trend notation is non-standard in the PDF glossary: '10-x' / '7-T' with hyphens vs classic '10x' / '7T'. Core rules (1-3s, 1-2s, 2-2s, R-4s, 4-1s) are labeled correctly.",
        "Copy/Cosmetic", "server/pdfQCMonthly.ts:167-170",
        "LOW",
        "Bias matched `^(\\d+)-x$` rendered '10-x'; trend `^(\\d+)-T$` rendered '7-T'. Internally consistent with the evaluator's rule_code, but a surveyor may not recognize '10-x' as the standard '10x'.",
        "Optionally normalize display labels to canonical Westgard notation ('10x', '8x', '7T').",
        "OPEN"),
    (21, "CUMSUM casing drift: the CumSum page uses the deliberate brand spelling 'CUMSUM' in headings but mixed-case 'CumSum' in the running-statistic labels (e.g. 'ACCEPT - |CumSum| <= 7.0 sec').",
        "Copy", "client/src/pages/CumsumPage.tsx:302,315,414,419,422,459,521 (CumSum) vs 241,257,287,457 (CUMSUM)",
        "LOW",
        "House rule keeps the deliberate spelling CUMSUM. The deliberate misspelling is preserved (no correct-spelling 'CUSUM' leaks anywhere user-facing), but the page mixes CUMSUM (brand) with CumSum (metric label). Cosmetic only.",
        "Standardize the metric label to CUMSUM (or confirm the brand-vs-metric distinction is intentional).",
        "OPEN"),
    (22, "'CLSI C24 supports lab-set bias_consecutive_count and trend_consecutive_count' is a mild overstatement (3 places). C24 broadly endorses selecting rules to the lab's quality requirements but does not specifically endorse configurable bias/trend counts.",
        "Copy/Data-truth", "client/src/pages/DemoQcPage.tsx:351 + VeritaAssurePage.tsx:96 + VeritaQCDailyReviewPage.tsx:497",
        "LOW",
        "Defensible as paraphrase of C24's rule-selection guidance, flagged as calibration only (not a fabrication).",
        "Soften to 'consistent with CLSI C24 rule-selection guidance' or similar.",
        "OPEN"),
    (23, "The demo QC reviewer's title reads 'Medical Director' (not 'or designee') in a demo attestation.",
        "Copy/Demo", "client/src/pages/DemoQcPage.tsx:330",
        "LOW",
        "A fictional reviewer's personal title (Dr. Sarah Mitchell, MD). The demo's attestation paragraph and the real PDF field label are both compliant; cosmetic only.",
        "Optionally use 'Laboratory Director' or add 'or designee' for consistency.",
        "OPEN"),
    (24, "Secondary Daily Review panels swallow load errors: a failed loadLots leaves the monthly-review lot dropdown silently empty; a failed loadPastReviews silently shows no prior attestations.",
        "Reliability/UX", "client/src/pages/VeritaQCDailyReviewPage.tsx:174-188 + 190-204",
        "LOW",
        "Both `catch { console.error }` only. Lower stakes than QC2/QC3, but a director could conclude 'no lots to review' / 'no prior attestations' on a fetch failure.",
        "Surface a small inline error on these secondary panels.",
        "OPEN"),
    (25, "The entry-page recent-results table omits the rule-detail tooltip and the rejection row-highlight that the Daily Review has, so on the entry page the tech sees only the bare rule code with no hover explanation and no row-level red.",
        "UX/Consistency", "client/src/pages/VeritaQCAppPage.tsx:700-707 (no title, no row tint)",
        "LOW",
        "Entry table renders `<Badge>{v.rule_code}</Badge>` with no title={v.detail}; the daily review adds title={v.detail} (422) and red-tints needsCA rows (412). The in-the-moment CA modal does show detail, so this is a consistency gap.",
        "Add title={v.detail} to the entry-page rule badges and red-tint rows carrying a rejection.",
        "OPEN"),
    (26, "a11y: the Daily Review 'accepted' checkmark has no aria-label, while the same icon on the entry page correctly uses aria-label='accepted'. Neither QC page uses ConfirmDialog on Retire-lot / File-attestation, diverging from the sweep's sibling pattern.",
        "a11y/Consistency", "client/src/pages/VeritaQCDailyReviewPage.tsx:468 (no aria-label)",
        "LOW",
        "Screen-reader users get no label on the daily-review accepted indicator; minor. Retire/attestation lack the ConfirmDialog 11 sibling pages use.",
        "Add aria-label='accepted'; consider a confirm on Retire/File-attestation for parity.",
        "OPEN"),
]

PASS_NOTES = [
    "Multi-lab isolation is STRONG -- unlike the VeritaTrack /worklist gap, EVERY VeritaQC route chains labScopeMiddleware and consistently reads/writes with the validated req.scope.labId (qc/results GET+POST, qc/lots, control-lots POST/PATCH, qc/recent, period-reviews GET/POST/PDF, corrective-actions, all import endpoints), with defensive lot/result ownership re-checks before each write. The escalate-to-response finding is written with lab_id = req.scope.labId (3251), the caId is validated to the lab (3209), and it is idempotent (409 with the existing finding id). The one isolation defect is QC5 (low-blast IDOR on the unvalidated qc_rule_violation_id, no PHI).",
    "The monthly QC review PDF is CLEAN of the two high-severity bug classes found in the sibling VeritaResponse audit: (1) lab identity is resolved correctly from the labs table via a raw `SELECT lab_name, clia_number FROM labs WHERE id = ?` (routes.ts:3062), NOT the Drizzle-typed storage.getUserById path that broke VeritaResponse -- so it prints the real lab name + CLIA with a proper 'CLIA: Not on file' hyphen fallback; and (2) the footer is a real FOOTER_TEMPLATE carrying Page X of Y on every page (pdfQCMonthly.ts:73-79 + 308-310), NOT the empty-string bug. Author metadata 'Perplexity Computer'; TM not (R); zero em-dashes in PDF output (the only one is a JSDoc comment); signature/attestation block on page 1; no URLs. No Excel export exists for VeritaQC, so Sec 6 is N/A.",
    "The client WRITE path is CLEAN -- every mutation (submit result, file CA, add/retire lot, escalate, PDF download, monthly attestation) checks res.ok and only reports success on a real 2xx, with pending states and success + destructive-error toasts; Add-Lot handles the 409 duplicate distinctly and escalate handles the 409 already-linked. No false-success save path. The module's systemic weakness is entirely in the READ path (QC2/QC3/QC8/QC24 error-as-empty), not silent-save. Plan gate uses the required explicit allowlist on both pages.",
    "The Westgard rule DEFINITIONS are 6-of-7 correct in both computed logic and user-facing description: 1-3s (|z|>3 reject), 1-2s (|z|>2 WARNING, correctly NOT an auto-reject), 2-2s (two consecutive same-side >2s), 4-1s (four consecutive same-side >1s), N-x bias, N-T trend -- with the PDF glossary matching the evaluator (no description-vs-logic contradictions). The baseline correctly excludes the candidate point and uses sample SD (n-1). Only R-4s deviates (QC6). All other CFR citations are eCFR-verified correct (493.1256 control procedures, 493.1282 corrective actions, 493.1105 retention); the Levey-Jennings / SDI / sigma math is correct; the deliberate CUMSUM spelling holds (QC21 is casing-only); and there is NO fabricated capability -- QC entry is honestly manual and there is no false 'auto-imports from the analyzer/LIS' claim (the import pipeline is an explicit manual mapping step).",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaQC Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaQC - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-12)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 86
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
for sev in ("HIGH", "MEDIUM", "LOW"):
    print(f"  {sev}: {sum(1 for x in ROWS if x[4]==sev)}")
