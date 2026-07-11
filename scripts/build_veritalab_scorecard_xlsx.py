#!/usr/bin/env python3
"""Build the VeritaLab module audit scorecard (internal deliverable for Michael).

Findings from a 4-lens read-only audit (reliability/multi-lab, export/PDF/xlsx,
compliance-copy/data-truth, UX), each verified by the main agent against current
code. Main-agent recalibrations vs the sub-agents:
  - The state-registry DATASET itself is sound: 51 rows (50 states + DC), 10 marked
    licensure yes/exempt, conservative, invents no agency names/URLs. The findings
    are in the COPY + citations layered on top, not the data. Sub-agent "all 50
    states" fear did NOT reproduce.
  - Lens-1/lens-4 both flagged the Excel export multi-lab bleed and the CMS-116
    download false-success; deduped to single rows.
  - Two "verify-before-ship" data rows (FL licensure, personnel-license states) are
    Michael's regulatory call, NOT unilateral code/data changes.

Usage:  python scripts/build_veritalab_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaLab_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaLab_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 22)]

ROWS = [
    (1, "CMS-116 issued-cert wire-back never sets lab_id, so the auto-created CLIA cert lands with NULL lab_id and is INVISIBLE on the roster it was built to populate.",
        "Multi-lab/Reliability", "routes.ts:25172-25190 (INSERT); roster read WHERE lab_id at :24883",
        "HIGH",
        "The wire-back INSERT column list is (user_id, cert_type, cert_name, cert_number, issuing_body, issued_date, expiration_date, lab_director, notes, is_auto_populated, is_active, created_at, updated_at) with NO lab_id, and VALUES starts with Number(lab.owner_user_id). The very next log line prints lab_id=${labId} (false; labId is never inserted). The certificate roster GET reads WHERE lab_id = ?, so the freshly-issued CLIA cert the CMS-116 flow promises to add never appears. The manual lab-scoped add-cert INSERT (:24901) DOES set lab_id, proving the column exists and the omission is the bug.",
        "Add lab_id to the wire-back INSERT column list + labId to VALUES. Same class as the VeritaStaff bulk-commit tier2_lab_id fix. Clean.",
        "OPEN - HIGH, fix ready"),
    (2, "A failed certificate-list fetch renders the 'No certificates yet' empty state, so a transient error looks identical to an empty roster. A director with certs sees 'none' and may re-add duplicates.",
        "Reliability/UX", "client/src/pages/VeritaLabAppPage.tsx:141-153 (loadCertificates)",
        "HIGH",
        "loadCertificates only setCertificates(data) inside if(res.ok); the catch does console.error and nothing else; there is no isError/error state. On any non-2xx or network failure the list stays [] and the page shows 'No certificates yet'. Same shape as the VeritaStaff roster + VeritaScan/VeritaMap empty-vs-error fixes already shipped this sweep.",
        "Add a distinct error state ('couldn't load certificates, retry') separate from the empty state, with a retry control. Same class as the VeritaStaff #6 fix.",
        "OPEN - HIGH, clean fix ready"),
    (3, "Certificate Excel export is scoped WHERE user_id, not lab_id, so a multi-lab owner's export bleeds every lab's certs into one workbook.",
        "Multi-lab/Export", "routes.ts:25396 (export query)",
        "MEDIUM",
        "The cert Excel export selects lab_certificates WHERE user_id = ?, ignoring the active lab. For a multi-lab owner (one owner_user_id across labs) the download mixes all labs' certificates into a single file with one lab's identity header, breaking the per-lab surveyor deliverable. Same class as the VeritaPolicy/VeritaStaff multi-lab re-key already shipped.",
        "Scope the export to the active lab (resolveActiveLabForRequest + WHERE lab_id). Same class as the VeritaPolicy/VeritaStaff fixes.",
        "OPEN - clean fix ready"),
    (4, "Documents modal never resets its list before fetching, so on a failed load it renders the PREVIOUS cert's documents under the newly opened cert. Download/Delete then act on the wrong cert's doc IDs.",
        "Reliability/UX", "VeritaLabAppPage.tsx:259-270 (openDocuments)",
        "MEDIUM",
        "openDocuments(certId) sets docCertId + docLoading, fetches, then setDocuments only inside if(res.ok); it never clears documents=[] on entry and has no error branch. Opening cert A (loads its docs), then opening cert B when B's fetch fails, leaves A's documents visible under B's modal. The row-level Download/Delete buttons operate on those stale doc IDs -> action against the wrong certificate's evidence.",
        "setDocuments([]) at the top of openDocuments + add an error state. Prevents cross-cert document actions.",
        "OPEN - clean fix ready"),
    (5, "Reminder-cadence copy is wrong: the app how-to card and the demo say reminders fire at '90, 60, and 30 days', but the real scheduler fires at 9 months, 6 months, 3 months, 30 days, and expiration. Misstates months as days AND drops 2 of 5 reminders on a real capability.",
        "Data-truth/Copy", "VeritaLabAppPage.tsx:396,400; DemoLabPage.tsx:1267,1275 (vs scheduler routes.ts:24733-24739)",
        "MEDIUM",
        "The scheduler builds reminders [9month, 6month, 3month, 30day, expiration] (routes.ts:24733-24739). The public VeritaLabPage already states this correctly (9/6/3 month + 30 day). But the in-app card whatItDoes + the tip ('emails you 90, 60, and 30 days') and the demo (both the paragraph :1267 and the how-it-works :1275) say '90, 60, and 30 days'. Customer-facing understatement of a shipped feature, and internally inconsistent with the public page.",
        "Change the app card + demo copy to the real 9-month / 6-month / 3-month / 30-day / expiration cadence (match the public page). Copy-only.",
        "OPEN - clean fix ready"),
    (6, "Stale 'Phase 4' placeholder copy tells the customer PDF generation is not built yet, while the working Download PDF button ships a live, implemented generator. Also leaks internal roadmap language.",
        "Data-truth/Copy", "Cms116FormTab.tsx:590-593,612 (vs endpoint routes.ts:25207 -> cms116Pdf.ts:474)",
        "MEDIUM",
        "Cms116FormTab renders 'the printed PDF (Phase 4) carries the signature block' (:590-593) and 'PDF generation lands in Phase 4' (:612). The same component ships a working Download PDF button (:357) hitting a live route (routes.ts:25207 -> generateCms116Pdf, cms116Pdf.ts:474). So the copy both (a) states a shipped feature isn't built and (b) exposes internal 'Phase 4' roadmap language to customers.",
        "Rewrite the two copy blocks to describe the shipped Download-PDF flow; drop 'Phase 4'.",
        "OPEN - clean fix ready"),
    (7, "Demo misattributes the CLIA auto-populate source: it claims CLIA is 'Auto-populated from your CMS-116 application', but CLIA is pulled from the account's CLIA lookup at signup. The CMS-116 -> tracker wire-back is a separate (deferred) path.",
        "Data-truth/Copy", "DemoLabPage.tsx:1274 (vs signup lookup VeritaLabAppPage.tsx:451)",
        "MEDIUM",
        "DemoLabPage:1274 'Auto-populated CLIA from your CMS-116 application.' The real mechanism auto-populates the CLIA cert from the account's CLIA number at signup, not from a CMS-116 filing. Stating a data flow that doesn't exist (and conflating it with the deferred wire-back) misrepresents how the product works to a prospect.",
        "Reword to 'Auto-populated from your CLIA verification at signup' (or drop the source clause).",
        "OPEN - clean fix ready"),
    (8, "Em-dash in the certificate Excel About sheet violates the public-facing em-dash ban.",
        "Export", "routes.ts:25459 (aboutBody with \\u2014)",
        "MEDIUM",
        "The cert Excel About paragraph contains a literal em-dash (\\u2014: '...on file in VeritaLab \\u2014 CLIA...'). CLAUDE.md Sec 3 bans em-dashes in every customer-facing artifact including Excel About paragraphs. Same class as the VeritaScan/VeritaPolicy About em-dash fixes.",
        "Replace the em-dash with a comma or colon.",
        "OPEN - clean fix ready"),
    (9, "Wrong CFR citation on the CMS-116 specialty/subspecialty section: cites '42 CFR 493.5' (which is test categories BY COMPLEXITY), where the specialty/subspecialty categories live in 42 CFR 493.2 / Subpart I.",
        "Compliance citation", "Cms116FormTab.tsx:496 (rendered), :92 (comment)",
        "MEDIUM",
        "Rendered UI at :496 'Mirrors the categories in 42 CFR 493.5.' and comment :92 'non-waived testing per 42 CFR 493.5'. 493.5 is 'Categories of tests by complexity' (waived/moderate/high); the specialty and subspecialty categories are defined at 493.2 and enumerated in Subpart I (493.801 et seq.). The checkbox list content itself is accurate; only the citation is misapplied.",
        "Change the citation to 42 CFR 493.2 / Subpart I. Log to the Master Citation Index for the canonical anchor.",
        "OPEN - clean fix + Master Citation Index note"),
    (10, "The State Registry empty state leaks an internal admin API to the customer: it tells them an administrator can run 'POST /api/admin/seed-state-registry'.",
        "UX/Copy", "StateRegistryTab.tsx:139-142",
        "MEDIUM",
        "When the per-state catalog isn't seeded, the empty-state renders 'An administrator can run the seed via POST /api/admin/seed-state-registry' in a code block. That's an internal ops endpoint exposed to a customer, and it reads as a broken/unfinished product. (On prod the table is seeded, so this is a defensive/edge surface, but the copy should never ship an admin endpoint.)",
        "Replace with a customer-appropriate 'state registry is being finalized, contact info@veritaslabservices.com' message.",
        "OPEN - clean fix ready"),
    (11, "Florida licensure row is likely stale/false: it says a Florida lab needs a state license, but Florida deregulated clinical-lab licensure circa 2018 and defers to CLIA. If so, a FL lab is told it needs a license it does not. VERIFY BEFORE SHIP.",
        "Data-truth (regulatory)", "server/stateRegistryData.ts:91-101 (FL, licensure_required 'yes', Fla. Stat. Ch. 483)",
        "MEDIUM",
        "The FL row sets licensure_required 'yes' cited to 'Florida Statutes Chapter 483', last_verified 2026-05-28. Florida's clinical-lab (and lab-personnel) licensure was repealed ~2018, deferring to CLIA. Certainty is not absolute (FL is commonly-cited), so this is a verify-before-ship flag on a data row, not a unilateral change. Changing state-requirement DATA is your regulatory call.",
        "Michael verifies FL's current statute; if deregulated, set licensure_required 'no' with the CLIA-only note. Do NOT change the data row without your ruling.",
        "OPEN - your ruling (regulatory)"),
    (12, "The 37 'no' rows are silent on personnel licensure that several of those states actually have (e.g. GA, HI, LA, MT, ND, NV, TN, WV), while CA/FL/MD correctly flag personnel-licensure as a separate axis. Inconsistent treatment of the same fact class.",
        "Data-truth (regulatory)", "server/stateRegistryData.ts:60 (CLIA_ONLY_NOTE) applied flat to 'no' rows",
        "MEDIUM",
        "CLIA_ONLY_NOTE ('No separate state laboratory license is required. The federal CLIA certificate from CMS is the operating credential.') is applied verbatim to states that DO license laboratory PERSONNEL. The CA/FL/MD rows call out personnel licensure as a distinct requirement; the flat 'no' rows do not, so a Nevada or Louisiana lab could miss a real personnel-license obligation. Same fact class handled two ways.",
        "Michael rules on which 'no' states carry a personnel-licensure note; add a personnel-license line to those rows. Regulatory call; do not backfill without your list.",
        "OPEN - your ruling (regulatory)"),
    (13, "CMS-116 Download PDF can report success even when the browser popup is blocked or the pre-save failed, so the user believes a PDF was produced when it wasn't. Read-only enforcement is also inconsistent (Save disabled, Download PDF not).",
        "UX/PDF", "Cms116FormTab.tsx:292-317 (handleDownloadPdf), :357",
        "LOW-MED",
        "handleDownloadPdf triggers the PDF open/save without asserting the new tab actually opened or the PUT that persists the draft succeeded; a popup-blocked or failed pre-save still reads as success. Separately, Save is disabled in read-only but Download PDF is not, so a read-only user can trigger a write path. Same shape as the VeritaPolicy PDF-token race the Gate-3 browser rule exists for.",
        "Assert the popup opened / handle the blocked case; gate Download PDF by the same read-only flag as Save. Browser-verify (Gate 3 step 8).",
        "OPEN"),
    (14, "CMS-116 tab is a dead-end when no active lab is resolved: the form is fully editable and Save is enabled, but Save/Download then toast 'No active lab'. No upfront guard.",
        "UX", "Cms116FormTab.tsx (editable form; Save/Download toast 'No active lab')",
        "LOW-MED",
        "With activeLabId null the entire CMS-116 form renders editable and the buttons are active; the user fills it out, clicks Save, and only then gets 'No active lab', losing the work. This overlaps with #1 (the wire-back) but is a distinct UX guard gap.",
        "Guard the tab up front when no lab is resolved (banner + disabled inputs), rather than failing on Save.",
        "OPEN"),
    (15, "COLA and 'board certifications' are advertised as tracked cert types but are absent from the CERT_TYPES enum (they fall under 'Other'); the public page omits COLA while the demo/app card include it.",
        "Copy", "VeritaLabAppPage.tsx:61-72 (enum), :396; DemoLabPage.tsx:1267,1273",
        "LOW",
        "CERT_TYPES = clia/cap/tjc/state_license/lab_director_license/vendor_agreement/other. The demo (:1267,:1273) and app card (:396) list 'COLA' and 'board certifications' as tracked types; neither is a first-class enum value, so they'd be logged as 'Other'. The public VeritaLabPage omits COLA, so the surfaces disagree. Minor overstatement.",
        "Either add COLA (and a board-cert type) to the enum, or align the copy to the actual enum. Your call.",
        "OPEN - your ruling"),
    (16, "'method validation' appears in a customer-facing registry note (NY row), against the labs-verify / manufacturers-validate house rule.",
        "Copy (house style)", "server/stateRegistryData.ts:156 (NY row)",
        "LOW",
        "The NY row renders 'Test menu and method validation reviewed at permit issuance...'. CLAUDE.md Sec 3 bans 'method validation' describing the lab's activity (use 'method verification / performance verification'). NY CLEP itself uses 'validation' in its own docs, so this is a house-style flag, not a factual error.",
        "Reword to 'method verification' (or attribute the term to NY CLEP if kept). Low priority.",
        "OPEN"),
    (17, "last_verified '2026-05-28' is stamped on all 37 bulk-generated 'no' states, which renders to customers as a per-state verification date that didn't happen.",
        "Data-truth", "server/stateRegistryData.ts:53 (TODAY) -> :328",
        "LOW",
        "The 37 'no' rows share the TODAY constant (2026-05-28) as last_verified. The header comment defines the field as 'authoring or null' (:29), but the value still renders as a verification timestamp in the State Registry tab, implying individual review of each state.",
        "Either null last_verified on the bulk 'no' rows or relabel the column ('catalog authored'). Low priority.",
        "OPEN"),
    (18, "Certificate empty-state promises passive auto-population 'when you refresh' with no refresh control, so a new user waits for something that won't happen on its own.",
        "UX/Copy", "VeritaLabAppPage.tsx (cert empty-state copy)",
        "LOW",
        "The 'No certificates yet' state tells the user CLIA will auto-populate on refresh, but offers no refresh/reload affordance and the auto-populate is a one-time signup-time action. Combined with #2 (error masquerading as empty) this compounds the confusion.",
        "Add a reload/refresh action or reword to describe the actual one-time signup behavior. Pairs with #2.",
        "OPEN"),
    (19, "Exempt-state CFR anchor cites 42 CFR 493.551 (the opening section of Subpart E); the exempt-state approval provisions run 493.553-493.575.",
        "Compliance citation", "server/stateRegistryData.ts:57 (exempt-state note)",
        "LOW",
        "The exempt-state designation (NY, WA) cites 493.551. That's the general/opening section; the exempt-state approval detail is 493.553-575. Defensible as a subpart anchor but worth tightening. The 'only NY and WA hold this designation' claim is correct.",
        "Tighten the anchor to 493.553-575 (or keep 493.551 as the subpart anchor by choice). Master Citation Index.",
        "OPEN - Master Citation Index"),
    (20, "Two CMS-116 fields (Director NPI, CLIA Director ID) may not exist on the official CMS-116 form; confirm before the wet-signed PDF represents them as form fields.",
        "Compliance (verify)", "Cms116FormTab.tsx:577-578",
        "LOW",
        "The form collects 'Director NPI' (:577) and 'CLIA Director ID (if previously assigned)' (:578). The official CMS-116 (OMB 0938-0581, correctly cited at :339) collects director name/credentials but, to the auditor's knowledge, does not label a Director-NPI field. Both are hedged/optional so low risk, but the printed PDF shouldn't present a field the federal form lacks.",
        "Verify against the current CMS-116; drop or relabel the two fields if not on the official form. Your call.",
        "OPEN - your ruling (verify vs form)"),
]

PASS_NOTES = [
    "State-registry DATASET is sound: exactly 51 rows (50 states + DC), 10 marked licensure yes/exempt (CA, FL, MD, MA, NJ, OR, PA, RI + exempt NY, WA), 4 unknown, 37 no. No duplicates, none missing; it does NOT claim all 50 states license labs and invents no agency names or URLs (all authority_url null). The core accuracy discipline holds; the findings are in the copy + citations layered on top.",
    "CMS-116 PDF is fully shipped and compliant: real endpoint (routes.ts:25207 -> cms116Pdf.ts:474 generateCms116Pdf) with a proper FOOTER_TEMPLATE + Page X of Y (NOT the empty-footer class seen in VeritaPolicy/VeritaPT). No separate certificate-report PDF exists to break.",
    "CLIA cert-type enum is exactly correct (Waiver / PPM / Compliance / Accreditation; no invented types); the CMS-116 specialty/subspecialty checkboxes match the CLIA categories; the OMB control number (0938-0581) is correctly cited.",
    "Copy hygiene clean: accreditor references are CAP/TJC/COLA/AABB only (no ACHC, no 'CMS surveyor'; 'former TJC surveyor' used); all product marks use TM not (R); no dated-manual references; no Mayo/TODO/placeholder leakage.",
    "Destructive actions use ConfirmDialog (delete cert / delete document both confirm); CMS-116 field bindings are correct (no form/state field mismatch); State Registry filter + external links have correct deps.",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaLab Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaLab - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-10)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 46
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
print("HIGH:", sum(1 for x in ROWS if x[4]=="HIGH"),
      "| MEDIUM:", sum(1 for x in ROWS if x[4]=="MEDIUM"),
      "| LOW-MED:", sum(1 for x in ROWS if x[4]=="LOW-MED"),
      "| LOW:", sum(1 for x in ROWS if x[4]=="LOW"))
