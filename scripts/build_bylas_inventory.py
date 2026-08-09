#!/usr/bin/env python3
"""
build_bylas_inventory.py - transcribe the Bylas (SCAHC Clarence Wesley, lab 6)
"Lab Request Form" into VeritaStock inventory items, export a review xlsx, and
optionally load them.

Source is a scanned 1-page form (no text layer), vision-transcribed. QTY/REF-LOT
columns on the form are blank, so every item loads at on_hand = 0 (Clarence
Wesley counts them in via the kiosk afterward). Items flagged verify=True are
ones I'm not fully confident reading off the scan.

  --xlsx    write the review workbook (default)
  --commit  POST to /api/admin/import-inventory for lab 6 (needs ADMIN_SECRET env)
"""
import os, sys, re, json, datetime, urllib.request

LAB_ID = 6
TAG = f"Bylas order form import {datetime.date.today().isoformat()}"

# ---- Section 1: Supply Chain Warehouse Stock  (number, description, issue) ----
WAREHOUSE = [
    ("10008", "Tape, Silk Hypoallergenic 1 in x 10 yd", "BX/12"),
    ("12734", "Syringe 10 mL Disposable Luer Lock", "BX/200"),
    ("10965", "Blue Cap Vacuette Tube Coagulation Sodium Citrate 3.5 mL", "PK/50"),
    ("10902", "Red Cap Vacuette Tube Serum Sep Clot Activator 3.5 mL", "PK/50"),
    ("10904", "Lavender Cap Vacuette Tube K2EDTA Sterile 4 mL", "PK/50"),
    ("10906", "Green Cap Vacuette Tube Lithium Heparin Sep Sterile 3.5 mL", "PK/50"),
    ("10472", "Band-Aid Medtoons 3/4 in x 3 in", "BX/100"),
    ("10641", "Tourniquet Blue Latex-Free 1 in", "PK/100"),
    ("10095", "Holder, Vacutainer One Use", "BG/250"),
    ("10096", "Blood Collection Set Vacutainer Safety-Lok 21G x 3/4in x 12in", "BX/50"),
    ("10097", "Blood Collection Set Vacutainer Safety-Lok 23G x 3/4in x 12in", "BX/50"),
    ("10098", "Adapter BD Multiple Sample Vacutainer, Luer", "BX/100"),
    ("10099", "Needle BD Blood Collection Eclipse 22G x 1-1/4in (Straights)", "BX/48"),
    ("10151", "Purple Top Sani-Cloth, Disposable", "EA/1"),
    ("10178", "Urine Collection Kit Midstream, Sterile 120 mL", "EA/1"),
    ("10180", "Glove, Nitrile, Powder Free - Large", "BX/100"),
    ("10181", "Glove, Nitrile, Powder Free - Medium", "BX/100"),
    ("10182", "Glove, Nitrile, Powder Free - Small", "BX/100"),
    ("10183", "Glove, Nitrile, Powder Free - X-Large", "BX/100"),
    ("10199", "Gauze Sponges Non-Sterile 8 Ply 4 in x 4 in", "PK/200"),
    ("10201", "Gauze Sponges Sterile Latex-Free 2 in x 2 in", "PK/200"),
    ("10202", "Alcohol Prep Pad Sterile, Medium", "BX/200"),
    ("10193", "Paper, Copy 8-1/2 in x 11 in", "CS/500"),
    ("10214", "Band-Aid Flex-Fabric 1 in x 3 in", "BX/100"),
    ("10297", "Hand Sanitizer (8 oz)", "EA/1"),
    ("10265", "AA Battery, Duracell", "PK/4"),
    ("10266", "AAA Battery, Duracell", "PK/4"),
    ("10308", "Tissue (Facial)", "BX/1"),
    ("10426", "Applicator, Cotton Tipped Sterile 6 in", "BX/100"),
    ("10912", "Heel Warmer Infant Gel Pack", "BX/25"),
    ("10915", "Pipette, Transfer 7 mL Graduated to 3 mL", "BX/500"),
    ("10924", "Band-Aid Curad 2 in x 4 in", "BX/50"),
    ("11184", "Needle Blood Collection Eclipse 21G x 1-1/4 in (Straights)", "BX/48"),
    ("11392", "Self-Adherent Coban Colors 2 in x 5 yd", "BX/10"),
    ("12755", "Chucks Underpad Protection Plus Blue 23 in x 36 in", "PK/10"),
]

# ---- Section 2: SCAHC Lab Stock  (description, ref/lot, verify?) ----
LABSTOCK = [
    ("Strep Kit", "", False),
    ("Flu Kit", "", False),
    ("COVID Kit", "", False),
    ("RSV Kit", "", False),
    ("hCG Kit", "", False),
    ("Multistix", "", False),
    ("Label (Refrigerate)", "", False),
    ("Label (Frozen)", "", False),
    ("Urine Test Strips", "", False),
    ("Lavender Pediatric Tube", "", False),
    ("Brown Tube", "", False),
    ("24 Hr Urine Container", "", False),
    ("Transport Vial", "", False),
    ("Transport Cap", "", False),
    ("Kimwipes, Small", "", False),
    ("Slide Holder", "", False),
    ("BreathTek", "", False),
    ("Aptima Vaginal", "", False),
    ("Aptima Urine", "", False),
    ("Glucola", "", False),
    ("QL Bags", "", True),
    ("A1C QC", "", False),
    ("WW Wash (for Tosoh)", "", False),
    ("MAS QC for Vitros", "LOT 26111", False),
    ("MAS QC for Vitros", "LOT 26112", False),
    ("MAS QC for Vitros", "LOT 26113", False),
    ("Collection Swab BD Max", "", False),
    ("Desiccant for Vitros", "REF 125-0237", False),
    ("POS QC COVID Kit", "", True),
    ("POS RSV QC Swab", "", False),
    ("Cellpack for XN-450 10 L", "", False),
    ("SLS Sulfolyser", "", False),
    ("#1 Reagent for Tosoh", "", False),
    ("#2 Reagent for Tosoh", "", False),
    ("Lysercell WDF (SL BX) Sysmex", "", False),
    ("Occult Blood QC", "", False),
    ("Macro Pipet Tips (Lg Pipets)", "", False),
    ("Hemoccult QC", "", False),
    ("MAS Bilirubin QC", "26091", False),
    ("MAS Bilirubin QC", "26092", False),
    ("MAS Bilirubin QC", "26093", False),
    ("Performance Verifier", "87245", False),
    ("Performance Verifier", "84211", True),
]

# ---- Section 3: Direct Order Items  (ref, description) ----
DIRECT = [
    ("TFT501", "Lancet, Tenderfoot (Toddler) Depth 2 mm x 13 mm"),
    ("367285", "Blood Collection Sets (25G x 3/4in x 12in)"),
]

UOM = {"BX": "box", "PK": "pack", "CS": "case", "BG": "bag", "EA": "each"}

def parse_issue(issue):
    m = re.match(r"([A-Z]{2})/(\d+)", issue or "")
    if not m:
        return ("each", 1)
    return (UOM.get(m.group(1), "each"), int(m.group(2)))

def classify(name):
    n = name.upper()
    if any(k in n for k in ("QC", "VERIFIER", "CONTROL", "OCCULT", "HEMOCCULT")):
        return "Control"
    if any(k in n for k in ("REAGENT", "CELLPACK", "SULFOLYSER", "LYSERCELL", "WASH", "GLUCOLA", "DESICCANT")):
        return "Reagent"
    if any(k in n for k in ("KIT", "APTIMA", "BREATHTEK", "MULTISTIX", "STRIP")):
        return "Reagent"
    return "Supply"

def build():
    rows = []
    for num, desc, issue in WAREHOUSE:
        unit, per = parse_issue(issue)
        rows.append(dict(section="Warehouse Stock", item_name=desc, catalog_number=num,
                         order_unit=unit, usage_unit="each" if per > 1 else unit, count_unit=unit,
                         units_per_order_unit=per, units_per_count_unit=per,
                         category=classify(desc), issue=issue, verify=False))
    for desc, ref, verify in LABSTOCK:
        rows.append(dict(section="Lab Stock", item_name=desc, catalog_number=(ref or None),
                         order_unit="each", usage_unit="each", count_unit="each",
                         units_per_order_unit=1, units_per_count_unit=1,
                         category=classify(desc), issue="", verify=verify))
    for ref, desc in DIRECT:
        rows.append(dict(section="Direct Order", item_name=desc, catalog_number=ref,
                         order_unit="each", usage_unit="each", count_unit="each",
                         units_per_order_unit=1, units_per_count_unit=1,
                         category=classify(desc), issue="", verify=False))
    return rows

def to_payload(rows):
    out = []
    for r in rows:
        note = TAG + (f". Stock# {r['catalog_number']}" if r["section"] == "Warehouse Stock" and r["catalog_number"] else "")
        out.append({
            "item_name": r["item_name"], "catalog_number": r["catalog_number"],
            "quantity_on_hand": 0, "reorder_point": 0, "category": r["category"],
            "unit": r["usage_unit"], "order_unit": r["order_unit"], "usage_unit": r["usage_unit"],
            "count_unit": r["count_unit"], "units_per_order_unit": r["units_per_order_unit"],
            "units_per_count_unit": r["units_per_count_unit"], "status": "active", "notes": note,
        })
    return out

def write_xlsx(rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bylas Inventory"
    hdr = ["Section", "Stock# / REF / LOT", "Item Name", "Form Issue", "Order Unit", "Usage Unit", "Count Unit", "Units/Pack", "Category", "On Hand", "VERIFY?"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="01696F")
    for r in rows:
        ws.append([r["section"], r["catalog_number"] or "", r["item_name"], r["issue"],
                   r["order_unit"], r["usage_unit"], r["count_unit"], r["units_per_order_unit"],
                   r["category"], 0, "REVIEW" if r["verify"] else ""])
    for col, w in zip("ABCDEFGHIJK", [16, 18, 52, 10, 11, 11, 11, 10, 12, 9, 9]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(path)

def main():
    rows = build()
    flagged = [r for r in rows if r["verify"]]
    print(f"Bylas items: {len(rows)}  (Warehouse {len(WAREHOUSE)}, Lab Stock {len(LABSTOCK)}, Direct {len(DIRECT)})")
    from collections import Counter
    print("category dist:", dict(Counter(r["category"] for r in rows)))
    print("unit dist:", dict(Counter(r["usage_unit"] for r in rows)))
    print(f"flagged to verify ({len(flagged)}):", [r["item_name"] for r in flagged])
    xlsx = r"C:\Users\veril\Downloads\Bylas_Inventory_Review.xlsx"
    write_xlsx(rows, xlsx)
    print("Review workbook:", xlsx)
    if "--commit" in sys.argv:
        sec = os.environ.get("ADMIN_SECRET", "")
        if not sec:
            print("ADMIN_SECRET not set; cannot commit."); return
        body = json.dumps({"secret": sec, "labId": LAB_ID, "items": to_payload(rows)}).encode()
        req = urllib.request.Request("https://www.veritaslabservices.com/api/admin/import-inventory",
                                     data=body, headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=60) as r:
            res = json.loads(r.read().decode())
        print("IMPORT RESULT:", {k: res.get(k) for k in ("ok", "inserted")}, "errors:", len(res.get("errors", [])))

if __name__ == "__main__":
    main()
