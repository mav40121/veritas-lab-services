#!/usr/bin/env python3
"""Build the VeritaStaff module audit scorecard (internal deliverable for Michael).

Findings from a 4-lens read-only audit (reliability/multi-lab, export/PDF/xlsx,
compliance-copy/data-truth, UX), each verified by the main agent against current
code. Main-agent recalibrations vs the sub-agents:
  - Lens 3 called the "General Supervisor tied to high complexity" a HIGH defect
    vs CLAUDE.md's "Waived = General Supervisor". Recalibrated: the CODE is
    regulatorily correct (CLIA GS 493.1461 IS a high-complexity role; PPM has its
    own GS provisions). This is a CLAUDE.md-Sec-5 imprecision to clarify, NOT a
    code bug -> recorded as a ruling row, not a fix.
  - Lens 4's "sign out" defect CONFIRMED as a security HIGH (token key mismatch).
  - Lens 1's bulk-import wrong-lab CONFIRMED (no lab-scoped bulk route exists).

Usage:  python scripts/build_veritastaff_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaStaff_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaStaff_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 22)]

ROWS = [
    (1, "Staff Portal 'Sign out' is a no-op: it removes the wrong localStorage key, so a shared kiosk never actually logs the tech out. The next person is silently authenticated as the previous one.",
        "Security", "client/src/pages/StaffPortalPage.tsx:128 (vs :87); lib/auth.ts:4,51",
        "HIGH",
        "signOut() does localStorage.removeItem('auth_token'), but the portal reads getItem('veritas_token') (:87) and the app writes 'veritas_token' (lib/auth TOKEN_KEY). 'auth_token' does not exist, so the real token survives; window.location.href='/login' redirects but a return to /staff-access silently re-auths as the prior employee, who can sign policies/competencies + adjust inventory under that identity. Defeats the kiosk-isolation premise in the file header.",
        "signOut() should call clearAuth() (lib/auth.ts:51, removes TOKEN_KEY + USER_KEY) instead of removing 'auth_token'. One-line, clean.",
        "OPEN - HIGH, fix ready"),
    (2, "Bulk import writes to the WRONG LAB on a multi-lab owner. Uploading Lab B's roster inserts every employee into Lab A, silently.",
        "Multi-lab", "routes.ts (bulk-preview/commit/template resolve staff_labs WHERE user_id); VeritaStaffAppPage.tsx:1962/1987/2015",
        "HIGH",
        "The VeritaStaff bulk trio (/api/staff/employees/bulk-preview|bulk-commit|template) exists ONLY in the legacy user_id-scoped form; there is no /api/labs/:labId/staff/employees/bulk-* route (VeritaComp's bulk IS lab-scoped; VeritaStaff's is not). The client always calls the un-scoped route (never passes activeLabId). The handler resolves the lab via SELECT id FROM staff_labs WHERE user_id = ? -> for a multi-lab owner (one staff_labs row per lab, same user_id) this returns an arbitrary (first) row. The preview's duplicate check reads that same wrong lab. Same defect stamps the template lab-name.",
        "Add lab-scoped /api/labs/:labId/staff/employees/bulk-preview|commit|template resolving by tier2_lab_id (mirror the VeritaComp bulk routes), and point the client at them when activeLabId is set. Same class as the VeritaPolicy multi-lab fix.",
        "OPEN - HIGH"),
    (3, "Lab-scoped employee edit silently drops the Staff Portal access toggles (Can adjust inventory / Can view audit). Multi-lab accounts always hit this path.",
        "Multi-lab", "routes.ts:22039-22055 (vs legacy 21795-21816)",
        "MEDIUM",
        "PUT /api/labs/:labId/staff/employees/:id destructures the body WITHOUT canAdjustInventory/canViewAudit and its UPDATE omits can_adjust_inventory/can_view_audit entirely. The legacy PUT (21799,21815-16) DOES persist them, and EmployeeDialog always sends both. On a multi-lab account (activeLabId set -> lab-scoped PUT), toggling access ON, clicking Update, shows 'Employee updated' but the grant never writes; the tech still cannot adjust inventory.",
        "Add canAdjustInventory/canViewAudit to the destructure + UPDATE on the lab-scoped PUT, matching the legacy handler. Clean fix; pairs with #1 (both are the kiosk-access model).",
        "OPEN - clean fix ready"),
    (4, "Employee create/update is non-atomic: DELETE staff_roles + re-insert loop with no transaction, so a mid-loop throw leaves the employee with ZERO roles.",
        "Reliability", "routes.ts:22056-22059 (lab-scoped); 21786+ (legacy)",
        "MEDIUM",
        "On edit, DELETE FROM staff_roles WHERE employee_id (22057) commits, then the role INSERT loop (22059) can throw (malformed specialtyNumber / crash). Old roles gone, new never landed, response 500s. Position-descriptions + instruments PUTs DO wrap in sqlite.transaction() (22357) -- the employee create/update paths do not.",
        "Wrap the employee create + update (row + roles + schedule) in sqlite.transaction(), mirroring the instruments PUT.",
        "OPEN"),
    (5, "Competency schedule save nulls first_annual_due_at for CLIA-only labs on every save, erasing the create-time milestone.",
        "Data quality", "routes.ts:23978-24000 / 24053-24075 (vs create-time helper 21948)",
        "MEDIUM",
        "The competency PUT unconditionally recomputes six_month/first_annual/annual due dates from the request body; firstAnnualDue is assigned ONLY inside the includesTJCorCAP branch (23945). For a CLIA-only lab, first_annual_due_at is written null on every save, even though ensureCompetencyScheduleMilestones back-filled it at employee-create. Saving the dialog (even just Notes) silently drops the 1st-Annual milestone from the roster status. The two write paths disagree on the schema's meaning.",
        "Preserve first_annual_due_at when the accreditor branch does not compute it (COALESCE with the existing value, or compute it for CLIA-only too). Confirm intended CLIA cadence first.",
        "OPEN - needs a look"),
    (6, "Dashboard tiles + roster swallow fetch errors: a broken stats endpoint or a failed roster load looks identical to 'no data' / 'no employees yet'.",
        "Reliability/UX", "CompetencyStatusTile.tsx:42,50; CredentialExpirationTile.tsx:54,65; VeritaStaffAppPage.tsx:203-206,416-426",
        "MEDIUM",
        "Both tiles: if (!r.ok) return zeros; then if (total === 0) return null -> the tile VANISHES on a 500, indistinguishable from 'nobody overdue'. The roster useQuery has no error branch; on a failed employees fetch it defaults to [] and renders 'No employees yet. Add your first employee', so a director with 20 staff may believe the roster was wiped and re-add duplicates.",
        "Add isError branches: a tile 'stats unavailable, retry' state; a roster error state distinct from the empty state. Same class as the VeritaComp/VeritaScan empty-state fixes.",
        "OPEN - clean fix ready"),
    (7, "Staff bulk-import template ships off-brand blue headers, not house teal (clashes with the VeritaComp template).",
        "Export", "server/staffBulkImport.ts:172,102 (vs competencyBulkImport.ts:125)",
        "MEDIUM",
        "Header + title fill = FF1F4E78 (Microsoft blue); CLAUDE.md Sec 6/6.7 mandates teal 01696F, 'brand colors only, no off-shades'. The VeritaComp import template correctly uses FF01696F, so the two templates visibly disagree. (Blank fill-in template, so lower blast radius than a populated export.)",
        "Change the staff template header + title fill to FF01696F. Mechanical.",
        "OPEN - clean fix ready"),
    (8, "?filter=overdue/expiring roster filter reads window.location.search imperatively, so the shown list can disagree with the URL after Clear.",
        "UX", "VeritaStaffAppPage.tsx:377,397,431",
        "LOW-MED",
        "The filter is read via new URLSearchParams(window.location.search).get('filter') inside render, not from wouter's reactive location. The drilldown paint is correct, but clicking Clear (navigate to the unfiltered route) can leave the 'Showing N overdue' banner + filtered list because the imperative read is not re-triggered.",
        "Read the filter from wouter's useSearch()/reactive location so it re-renders on navigation.",
        "OPEN"),
    (9, "Bulk import UX rough edges: success toast even on 0 imported rows; raw server error string on a bad/corrupt file; two same-name people hard-block the whole import.",
        "UX", "VeritaStaffAppPage.tsx:2022,1993-1998; staffBulkImport.ts:391,500-506",
        "LOW-MED",
        "Commit toasts 'Import complete: 0 added, 0 updated' with no indication nothing happened (the sibling CompetencyBulkImportDialog handles this better). Preview failures surface a backend parse string with no guidance. Dedup is name-only (first+last); two legitimately-distinct 'John Smith' TPs in one file hard-error and block the entire commit rather than importing with a warning.",
        "Toast a real 'nothing imported' state; friendlier file-error guidance; allow same-name rows with a warning (dedup on a stronger key). Your call on the dedup policy.",
        "OPEN - your ruling on dedup"),
    (10, "CMS-209 PDF interpolates lab + employee fields without HTML-escaping; a name/address/qualifications value with &, <, or > breaks the federal-form layout a surveyor sees.",
        "Export/PDF", "server/pdfReport.ts:5005-5007, 4943-4955",
        "LOW",
        "buildCMS209HTML injects ${lab.lab_name}, ${address}, ${r.lastName}, ${r.quals} raw; every other generator routes user data through esc() (4207). Robustness bug, not a formatting-standard break.",
        "Wrap the interpolated lab/employee fields in esc() in buildCMS209HTML.",
        "OPEN - clean fix ready"),
    (11, "Termination dialog allows a blank Reason with no nudge, undercutting the surveyor record the dialog exists to capture.",
        "UX", "VeritaStaffAppPage.tsx:1455,1463,1474",
        "LOW",
        "The dialog copy promises 'records preserved for CMS records retention' but the code comment (1455) confirms 'Empty reason is allowed' and nothing prompts on a blank reason. It IS a real confirm dialog (good), so this is a soft record-quality gap, not a missing-confirm defect.",
        "Nudge (not block) on a blank termination reason. Your call.",
        "OPEN - your ruling"),
    (12, "'Import departments from VeritaMap' button has no loading/disabled state and an empty catch, so on a slow/failing call it reads as a dead button.",
        "UX", "VeritaStaffAppPage.tsx:638-651,757-759",
        "LOW",
        "loadSuggestions wraps fetch in try{}catch{} with an EMPTY catch; the toast only fires inside if(res.ok); the button has no disabled/spinner. On a lab with no map or a slow endpoint, clicking does nothing visible and a thrown error is eaten.",
        "Add a loading/disabled state + a toast on non-ok/failure. Optional-path, low priority.",
        "OPEN"),
    (13, "GS-complexity + role-qualification CFR citations: for Michael's regulatory ruling, not a code fix.",
        "Compliance ruling", "VeritaStaffAppPage.tsx:866,933-957; VeritaStaffPage.tsx:161",
        "LOW-MED",
        "Lens 3 flagged 'General Supervisor tied to HIGH complexity' as contradicting CLAUDE.md's 'Waived = General Supervisor'. Main-agent read: the CODE is regulatorily sound (CLIA GS 493.1461 IS high-complexity; PPM has its own GS) -> CLAUDE.md Sec 5's one-liner is the imprecise side. Separately, the role-qualification reminder block (933-957) has several suspect/unlisted CFR cites (LD moderate .1405, CC .1417, GS-'not used' .1459, TP moderate .1421-1423) that should be reconciled against the Master Citation Index, not resolved in code.",
        "Clarify CLAUDE.md Sec 5 GS wording (do NOT change the code); send the 933-957 citation block to the Master Citation Index for a ruling. Also verify the '17 CMS specialty categories' claim vs the CMS-209 enumeration.",
        "OPEN - your ruling"),
]

PASS_NOTES = [
    "No silent-save class: every VeritaStaff mutation (employee, lab-setup, competency, delete, bulk-commit, position-descriptions, documents, instruments) shows a toast AND invalidates its query. onError toasts present.",
    "No empty-footer PDF bug in VeritaStaff: the competency PDF uses a full COMPETENCY_FOOTER (brand + Page X of Y on every page); CMS-209 legitimately keeps its own FORM CMS-209 (09/2025) federal-form footer + a per-page license band. Signature is on page 1; author metadata is 'Perplexity Computer'; CLIA header + fallback correct; 'laboratory director or designee' correct.",
    "General CRUD is correctly lab-scoped: the /api/labs/:labId/staff/* employee/competency/documents/instruments/position-description/CMS-209 endpoints scope by req.scope.labId (tier2_lab_id). The multi-lab hole is confined to the bulk-import trio (#2) + the two dropped toggle columns (#3).",
    "Position-descriptions + instruments writes ARE atomic (sqlite.transaction) and correctly scoped. Em-dash clean in all generated artifacts; TM not (R); no Mayo/placeholder/TODO leakage; controlled-vocab titles (staffTitles.ts) honored in the UI.",
    "Staff Portal module views (policies, inventory, audit, competency, quizzes) all have proper loading AND error AND empty states -- a good contrast to the director-side tiles (#6).",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaStaff Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaStaff - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-10)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 34
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
print("HIGH:", sum(1 for x in ROWS if x[4]=="HIGH"),
      "| MEDIUM:", sum(1 for x in ROWS if x[4]=="MEDIUM"),
      "| LOW-MED:", sum(1 for x in ROWS if x[4]=="LOW-MED"),
      "| LOW:", sum(1 for x in ROWS if x[4]=="LOW"))
