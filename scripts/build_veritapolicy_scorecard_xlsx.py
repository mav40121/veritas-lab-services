#!/usr/bin/env python3
"""Build the VeritaPolicy module audit scorecard (internal deliverable for Michael).

Findings came from a 4-lens read-only audit (reliability/multi-lab, export/PDF/DOCX,
compliance-copy/data-truth, UX) and were then EACH verified by the main agent against
current code. Main-agent recalibrations vs the sub-agents:
  - Sub-agent 3 counted 57 master-list rows; main-agent recount = 58 (the 58th quoted
    "policy_id" is a real row, not the interface field, which is unquoted). Real count = 58.
  - Sub-agent 4's reported client-side "notes/field mismatch" was verified NOT present
    (edit-binding is keyed per policy_id) -> recorded as a PASS. A DIFFERENT notes defect
    IS real: the master-list SEED-DATA notes column is shifted off its rows (finding 7).
  - The two multi-lab collisions (master_status + settings) are one bug class = one PR.

Usage:  python scripts/build_veritapolicy_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaPolicy_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaPolicy_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 22)]

ROWS = [
    (1, "Multi-lab data loss: a multi-lab owner's per-lab policy STATUS collides into one shared row. Marking a policy in Lab B silently erases Lab A's status for that policy.",
        "Multi-lab", "routes.ts:28626-28642 (write) / 28593 (read); db.ts:2293, 2656",
        "HIGH",
        "Write upserts ON CONFLICT(user_id, policy_id) with userIdForRow = owner_user_id (28622), which is IDENTICAL across every lab that owner holds; read is WHERE lab_id = ? (28593). UNIQUE is (user_id, policy_id) (db.ts:2293); lab_id was added as a plain column (2656) but the UNIQUE was never recreated (comment 2645 admits the follow-up is pending). verilabguy (labs Michaels + Riverside under user 17) is the exact trigger. Every one of the 58 rows behaves this way; both labs' readiness scores corrupt.",
        "Rebuild the table with UNIQUE(lab_id, policy_id), backfill lab_id from existing rows, change ON CONFLICT to (lab_id, policy_id). Same PR as #2.",
        "OPEN - HIGH, fix ready"),
    (2, "Multi-lab data loss: veritapolicy_settings holds ONE row per owner across ALL their labs. is_independent / accreditation_body flip when the owner switches labs.",
        "Multi-lab", "routes.ts:28567-28576; db.ts:2226",
        "HIGH",
        "PUT settings upserts ON CONFLICT(user_id) (28570) with owner_user_id; UNIQUE(user_id) (db.ts:2226) means a multi-lab owner can never hold distinct VeritaPolicy settings per lab. Same class as #1, confined to the two legacy user_id-keyed tables (the newer approval-workflow tables are correctly lab_id-scoped and guarded).",
        "UNIQUE(lab_id) + backfill + ON CONFLICT(lab_id). Fold into the #1 PR (one migration touches both tables).",
        "OPEN - HIGH, fix ready"),
    (3, "VeritaPolicy readiness PDF has NO per-page footer and NO 'Page X of Y' on any page (CLAUDE.md Sec 5 non-negotiable footer).",
        "Export/PDF", "pdfReport.ts:5669; licenseStamp.ts:113-114",
        "HIGH",
        "generateVeritaPolicyPDF calls applyLicenseToPuppeteer(html, \"\", ctx) with an EMPTY baseFooter, so licenseAugmentedFooterTemplate hits `if (!baseTemplate) return band-only` and the Puppeteer running footer carries only the yellow license band: no page numbers, no VeritaAssure/VeritaPolicy/Confidential line. Every other generator passes FOOTER_TEMPLATE (3484). Overflow pages of the requirements table get no footer at all. NOTE: the per-policy DOCX footer is CORRECT; only this PDF is broken.",
        "Pass the brand+pageNumber FOOTER_TEMPLATE (module 'VeritaPolicy') as baseFooter so the license band augments it, matching the other generators.",
        "OPEN - HIGH, fix ready"),
    (4, "Search box is dead: typing in 'Search policies by title, description, or manual' never re-filters the list.",
        "UX", "VeritaPolicyMyPoliciesPage.tsx:897-910",
        "MEDIUM",
        "The rendered table is built from `grouped` (memo), which iterates filteredDocs (899) but its dependency array is [documents, manuals] (910) -- it omits filteredDocs/searchQ. filteredDocs recomputes on keystroke; grouped returns its CACHED value, so the table never changes. Reads as a broken product on first use.",
        "Change the grouped deps to [filteredDocs, manuals]. One line.",
        "OPEN - clean fix ready"),
    (5, "False count: app + demo copy claim '96 policies' but the real master list has 58 rows.",
        "Data truth", "VeritaPolicyAppPage.tsx:40,465,467; DemoLabPage.tsx:1294,1303",
        "MEDIUM",
        "VERITAPOLICY_MASTER_LIST has 58 rows (IDs 1-111 non-contiguous, a curated subset); the lab-scoped readiness route computes over these 58. The how-to card ('96-policy master list', 'Browse the 96-row master list') and demo copy ('96 CFR-anchored policies', '96 generic templates') hardcode 96, shown on the same page as the live 58-row table. DemoLabPage:1303 also lists 'Personnel' as a spanned section, which no master row uses.",
        "Drive the number from VERITAPOLICY_MASTER_LIST.length (58) or hardcode 58; reconcile the section list. Public-facing accuracy.",
        "OPEN - clean fix ready"),
    (6, "Mayo placeholder ships in a user-facing input: the 'Upload new version' change-summary field example names Mayo, violating the standing Mayo-removal rule.",
        "Compliance copy", "VeritaPolicyMyPoliciesPage.tsx:1913",
        "MEDIUM",
        "placeholder=\"e.g., Updated critical value list per Mayo Q2 2026.\" renders to every user. The Mayo critical-value layer was removed 2026-07-10 and Mayo must not surface as a default anywhere. Only Mayo/placeholder/vendor-stub leak found.",
        "Neutral example, e.g. 'e.g., Updated critical value list per MEC review Q2 2026.' Mechanical.",
        "OPEN - clean fix ready"),
    (7, "Master-list NOTES column is shifted off its rows: several policies carry guidance describing a DIFFERENT policy. Ships in the Excel/UI as wrong per-row instructions.",
        "Data quality", "veritapolicyMasterList.ts (policy 10, 11, 35, 38, 39, 40, 66)",
        "MEDIUM",
        "Verified 2 of 7: policy 10 'System Downtime' notes = 'Reference lab list, current CLIA cert ... See policy 96.' (that is referral-lab content); policy 11 'Critical Value Reporting' notes = 'documented method evaluation, calibration verification, reportable range ... CFR 493.1253 is the anchor' (that is method-verification content). The whole notes column looks misaligned from a prior merge. (Distinct from the client edit-binding, which is CLEAN.)",
        "Re-align the notes column to their policies (data pass). Confirm scope of the shift before editing.",
        "OPEN - needs a data pass"),
    (8, "Em-dash in the customer-facing Excel About sheets of both VeritaPolicy export routes (CLAUDE.md Sec 6 rule 6 + Sec 3).",
        "Export/copy", "routes.ts:29088, 29108 (+ legacy 28393, 28417)",
        "MEDIUM",
        "About-sheet 'Sections' and 'Coverage gaps' paragraphs use a literal em-dash in all four VeritaPolicy Excel downloads. Already surfaced by script/audit.py as a warning. The two routes are near-duplicate copies (copy-paste class bug).",
        "Replace the em-dashes with a comma/period/semicolon in both the lab-scoped and legacy About copy.",
        "OPEN - clean fix ready"),
    (9, "Silent no-op: 'Edit policy details' cannot move a policy to Unassigned (clear its manual). The save toasts success but the change is discarded.",
        "Reliability", "routes.ts:30152; VeritaPolicyMyPoliciesPage.tsx:727",
        "MEDIUM",
        "The documents PATCH uses manual_id = COALESCE(?, manual_id); the client sends manualId: null to clear the manual (727). COALESCE(null, manual_id) keeps the OLD value, then the toast says 'Saved' and the refetch shows the policy still in its old manual.",
        "Allow an explicit null to clear manual_id (drop COALESCE for that column, or use a sentinel), so 'move to Unassigned' works.",
        "OPEN - clean fix ready"),
    (10, "Comprehension-quiz authoring UI is fully built but unreachable: PolicyQuizAuthorDialog is never imported or rendered anywhere.",
        "Dead feature", "components/PolicyQuizAuthorDialog.tsx (defined, 0 importers)",
        "MEDIUM",
        "Repo-wide grep for PolicyQuizAuthorDialog returns only its own definition; grep for 'quiz' in MyPoliciesPage returns nothing. The dialog (gate-vs-record, threshold, question-bank CRUD, backed by live /quiz-config + /quiz endpoints) has no entry point. A whole PR's UI is orphaned; the director cannot author an attestation quiz.",
        "Wire a 'Comprehension quiz' action into the policy row/detail, or confirm it is intentionally shelved and note it.",
        "OPEN - your call (wire or shelve)"),
    (11, "Non-atomic approve sign-off: signoff INSERT + final-step status UPDATE run without a transaction; a mid-sequence throw wedges the document.",
        "Reliability", "routes.ts:30381 (insert), 30416 (update)",
        "MEDIUM",
        "On final-step approval the signoff row commits, then UPDATE policy_documents SET status='approved' can throw (constraint/disk). The doc stays in_review with every step signed, so getCurrentPendingStep returns null and both approve and reject hit 409 'No pending step' -- only the owner can recall it. better-sqlite3 is single-threaded so this needs a mid-sequence throw, not a race (hence MED not HIGH). Same non-atomic shape in the new-version upload (30940/30946/30966).",
        "Wrap the approve sequence (and the new-version sequence) in sqlite.transaction().",
        "OPEN"),
    (12, "Curated set implied as exhaustive + unverified accreditor counts. 'All/Every policy required by [AO]' over a 58-row curated subset; 88/65/81/286 counts have no in-repo data source.",
        "Compliance copy", "VeritaPolicyPage.tsx:26,34,42,56,84,110,144",
        "LOW-MED",
        "The page renders 'All {count} {AO}-required policies pre-loaded' and 'Track all {count} policies required by {AO}' -- 'All/Every' frames the curated subset as complete coverage. The 88/65/81/286 constants are commented as 'from server requirement files auto-generated from the master citation index' but no such file exists in the repo (only VERITAPOLICY_MASTER_LIST.length=58). Provenance unconfirmed.",
        "Your ruling: soften 'All/Every' to 'key/curated', and confirm 88/65/81/286 against the Master Citation Index or replace with a sourced number.",
        "OPEN - your ruling"),
    (13, "Suspect CFR / citation format in the master list (flag for the Master Citation Index, do not resolve in code).",
        "Compliance copy", "veritapolicyMasterList.ts (policy 66, 11; CAP ranges)",
        "LOW-MED",
        "policy 66 'Manual Hematology QC' cites 42 CFR 493.927 -- but 493.927 is General Immunology; the hematology subpart is 493.941. policy 11 'Critical Value Reporting' cites four HIPAA 45 CFR 164 sections (stretch; likely bleed-over, reinforced by its mismatched notes). Two CAP cites use a hyphen RANGE (GEN.40491-GEN.40507) while every other CAP cite is a semicolon list -- inconsistent, won't split the same way.",
        "Master Citation Index ruling on 66 and 11; normalize CAP ranges to semicolon lists.",
        "OPEN - your ruling"),
    (14, "'validate' used for what the LAB does in two master rows (labs verify, manufacturers validate). Ships in the Excel/PDF export.",
        "Compliance copy", "veritapolicyMasterList.ts (policy 28, 36)",
        "LOW",
        "policy 36 description 'how the laboratory validates new methods' is the clear one; policy 28 'LIS Validation and Verification' is borderline (LIS validation is a defensible term of art).",
        "Change policy 36 'validates new methods' to 'verifies performance specifications for new methods'; policy 28 is your call.",
        "OPEN - mostly clean fix"),
    (15, "Submit-for-review is not blocked when zero eligible reviewers exist: the red warning renders but Submit stays enabled, sending the doc to in_review with nobody able to approve it.",
        "UX", "VeritaPolicyMyPoliciesPage.tsx:1729-1747, 1781-1789",
        "LOW",
        "The warning 'No eligible reviewer exists ... will leave the document stuck in review' shows, but the Submit disabled condition ignores eligibility.minCount===0. A first-time director with no reviewers invited can wedge a doc; the only escape is the separate Recall button. Arguably intentional (invite a reviewer later) so not HIGH.",
        "Disable Submit (or add a confirm) when eligibility.minCount===0, or auto-point the user to invite a reviewer first.",
        "OPEN"),
    (16, "New-version upload is offered in every status incl in_review, with no confirm; it silently resets the doc to draft and invalidates in-progress sign-offs.",
        "UX", "VeritaPolicyMyPoliciesPage.tsx:1265-1274",
        "LOW",
        "The upload-new-version icon renders whenever the viewer owns the doc, regardless of status. The dialog copy warns 'the approval workflow runs again' but there is no confirm step (unlike Archive and Recall, which both confirm). An owner can blow away a review mid-flight.",
        "Add a confirm step when status is in_review/approved, mirroring Archive/Recall.",
        "OPEN"),
    (17, "Approve/Reject row buttons + bulk 'Mark all N/A' lack in-flight guards (double-submit / optimistic desync).",
        "UX", "VeritaPolicyMyPoliciesPage.tsx:1278-1293; VeritaPolicyAppPage.tsx:358-386",
        "LOW",
        "Row-level Approve/Reject have no disabled=isPending (the dialog's final button IS guarded, so exposure is a confusing 2nd toast if the 30s refetch races the sign). Bulk N/A sets all rows optimistically then loops PATCHes; failed rows are not rolled back per-row until a full reload (the single-row path DOES roll back).",
        "Disable row buttons while signMutation.isPending; reconcile per-row failures in the bulk path.",
        "OPEN"),
    (18, "DOCX signature block Title label is space-padded in a proportional font, so the fill-lines do not align vertically (cosmetic).",
        "Export/DOCX", "veritapolicyDocx.ts:212",
        "LOW",
        "'Title:           ' uses literal spaces in Calibri; Print Name / Title / Signature / Date underscores misalign. No compliance impact (the Title field, the designee hint, the annual-review line, and the 493.1251(b)(13) citation are all present and correct).",
        "Use a tab stop or a 2-column table for the signature grid. Polish only.",
        "OPEN - cosmetic"),
]

PASS_NOTES = [
    "No silent-save class: every approval-workflow mutation (create-manual, submit, sign, rename, recertify, assign, attest, archive, recall) has an onSuccess toast + query invalidation + onError toast. Unlike VeritaMap, there is no false-success write here.",
    "Per-policy DOCX footer is CORRECT: a real Word section Footer with PageNumber.CURRENT/TOTAL_PAGES repeats on every page (veritapolicyDocx.ts:567-578). The Sec-5 breach is PDF-only (finding 3).",
    "Client edit-binding is CLEAN: 'Our Policy Name' and notes are keyed per policy_id (editingPolicyName[p.policy_id], updatePolicy by r.policy_id===p.policy_id). The reported client-side notes/field mismatch is NOT present in current code. (Finding 7 is a separate seed-DATA defect.)",
    "Excel Sec-6 structure COMPLIANT on both routes: About sheet 1 + activeTab 0, 3-layer lab identity from the live lab record (About row + oddHeader.right + oddFooter.left), teal headers, freeze B2, auto-filter, sheet protection. Only the em-dash (finding 8) breaks it.",
    "Loading / empty / error states are covered across the app (docs list, master table, view modal, audit trail, loadAll catch). No blank-screen dead-ends. Tab nav + destructive confirms (Archive, Recall) all correct.",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaPolicy Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaPolicy - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-10)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

# Pass-note banner rows (green)
for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 28
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
print("HIGH:", sum(1 for x in ROWS if x[4]=="HIGH"),
      "| MEDIUM:", sum(1 for x in ROWS if x[4]=="MEDIUM"),
      "| LOW-MED:", sum(1 for x in ROWS if x[4]=="LOW-MED"),
      "| LOW:", sum(1 for x in ROWS if x[4]=="LOW"))
