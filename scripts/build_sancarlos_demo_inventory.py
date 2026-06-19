#!/usr/bin/env python3
"""
build_sancarlos_demo_inventory.py

Demo inventory for the San Carlos Apache Healthcare enterprise demo (Tuesday).
Seven locations: 1 warehouse + 6 stockrooms (ED, Main Lab, Clarence Wesley Lab,
Pharmacy, Inpatient Unit, Clinic). NOT the real SCAHC production labs (2/6); this
is a fresh demo tenant.

Design goals so the demo lands:
  - Warehouse holds the bulk of every shared item (the source for transfers).
  - Shared catalog numbers across locations so items reconcile in the roll-up and
    a warehouse -> stockroom multi-item transfer matches and merges cleanly.
  - A handful of stockroom items deliberately staged BELOW reorder point so the
    low-stock alerts light up, and the warehouse always has stock of those.
  - Item mix matched to each location (lab reagents/tubes for the two labs, meds
    and IV stock for pharmacy, nursing/med-surg supplies for inpatient, etc).

No PHI: supplies, reagents, controls, and over-the-counter meds only.

Usage:
  python scripts/build_sancarlos_demo_inventory.py            # write review xlsx
  python scripts/build_sancarlos_demo_inventory.py --commit   # seed via API
      requires:  ADMIN_SECRET env  AND  --labmap '{"San Carlos Warehouse": 12, ...}'
      (lab ids come from creating the 7 labs under the demo account first)
"""
import os, sys, json, datetime, urllib.request
from collections import Counter

TAG = f"SCAHC demo seed {datetime.date.today().isoformat()}"

# catalog#: (name, order_unit, pack_size, category, department, vendor)
CATALOG = {
    # ----- shared general supplies (warehouse + most stockrooms) -----
    "GLV-NIT-S":  ("Nitrile Exam Gloves, Small",            "box",  100, "Supply",   "General",     "Medline"),
    "GLV-NIT-M":  ("Nitrile Exam Gloves, Medium",           "box",  100, "Supply",   "General",     "Medline"),
    "GLV-NIT-L":  ("Nitrile Exam Gloves, Large",            "box",  100, "Supply",   "General",     "Medline"),
    "ALC-PREP":   ("Alcohol Prep Pads, Sterile Medium",     "box",  200, "Supply",   "General",     "Medline"),
    "GZ-4X4":     ("Gauze Sponges 4 in x 4 in, Non-Sterile","pack", 200, "Supply",   "General",     "Cardinal Health"),
    "TAPE-SILK":  ("Silk Tape 1 in x 10 yd",                "box",  12,  "Supply",   "General",     "3M"),
    "SANI-WIPE":  ("Disinfectant Wipes, Purple Top Canister","each",1,   "Supply",   "General",     "PDI"),
    "HAND-SAN":   ("Hand Sanitizer 8 oz",                   "each", 1,   "Supply",   "General",     "Purell"),
    "SHARPS-1Q":  ("Sharps Container 1 Quart",              "each", 1,   "Supply",   "General",     "Becton Dickinson"),
    "SPEC-BAG":   ("Specimen Transport Bag, Biohazard",     "pack", 100, "Supply",   "General",     "Medline"),
    "SYR-10ML":   ("Syringe 10 mL Luer Lock",               "box",  200, "Supply",   "General",     "Becton Dickinson"),
    "NDL-21G":    ("Hypodermic Needle 21G x 1 in",          "box",  100, "Supply",   "General",     "Becton Dickinson"),
    # ----- phlebotomy / collection (labs + warehouse) -----
    "TUBE-RED":   ("Red Top Serum Tube 3.5 mL",             "pack", 50,  "Supply",   "Phlebotomy",  "Greiner"),
    "TUBE-LAV":   ("Lavender Top K2EDTA Tube 4 mL",         "pack", 50,  "Supply",   "Phlebotomy",  "Greiner"),
    "TUBE-BLUE":  ("Blue Top Sodium Citrate Tube 3.5 mL",   "pack", 50,  "Supply",   "Phlebotomy",  "Greiner"),
    "TUBE-GRN":   ("Green Top Lithium Heparin Tube 3.5 mL", "pack", 50,  "Supply",   "Phlebotomy",  "Greiner"),
    "TOURNIQUET": ("Tourniquet, Latex-Free",                "pack", 100, "Supply",   "Phlebotomy",  "Medline"),
    "BCS-21G":    ("Blood Collection Set 21G Safety-Lok",   "box",  50,  "Supply",   "Phlebotomy",  "Becton Dickinson"),
    "URINE-CUP":  ("Urine Specimen Cup, Sterile 120 mL",    "each", 1,   "Supply",   "Specimen",    "Medline"),
    # ----- lab reagents / controls (main lab + warehouse) -----
    "RGT-TOSOH1": ("#1 Reagent for Tosoh",                  "kit",  1,   "Reagent",  "Chemistry",   "Tosoh"),
    "RGT-TOSOH2": ("#2 Reagent for Tosoh",                  "kit",  1,   "Reagent",  "Chemistry",   "Tosoh"),
    "CELLPACK":   ("Cellpack for Sysmex XN, 10 L",          "each", 1,   "Reagent",  "Hematology",  "Sysmex"),
    "SULFOLYSER": ("SLS Sulfolyser, 5 L",                   "each", 1,   "Reagent",  "Hematology",  "Sysmex"),
    "QC-MAS":     ("MAS QC for Vitros, Level 1",            "box",  1,   "Control",  "Chemistry",   "Thermo MAS"),
    "STRIP-UA":   ("Urinalysis Test Strips, 100 ct",        "each", 1,   "Reagent",  "Urinalysis",  "Siemens"),
    # ----- point of care rapid kits (labs + clinic) -----
    "KIT-STREP":  ("Strep A Rapid Test Kit",                "kit",  1,   "Reagent",  "Point of Care","Abbott"),
    "KIT-FLU":    ("Influenza A/B Rapid Test Kit",          "kit",  1,   "Reagent",  "Point of Care","Abbott"),
    "KIT-COVID":  ("COVID-19 Antigen Test Kit",             "kit",  1,   "Reagent",  "Point of Care","Abbott"),
    "KIT-RSV":    ("RSV Rapid Test Kit",                    "kit",  1,   "Reagent",  "Point of Care","Abbott"),
    "KIT-HCG":    ("hCG Urine Pregnancy Test Kit",          "kit",  1,   "Reagent",  "Point of Care","McKesson"),
    "LANCET":     ("Safety Lancet 21G",                     "box",  100, "Supply",   "Point of Care","Becton Dickinson"),
    # ----- pharmacy -----
    "NS-1000":    ("Sodium Chloride 0.9% IV Bag 1000 mL",   "case", 12,  "Supply",   "Pharmacy",    "Baxter"),
    "D5W-1000":   ("Dextrose 5% in Water IV Bag 1000 mL",   "case", 12,  "Supply",   "Pharmacy",    "Baxter"),
    "SYR-3ML":    ("Syringe 3 mL Luer Lock",                "box",  100, "Supply",   "Pharmacy",    "Becton Dickinson"),
    "FLUSH-NACL": ("Sodium Chloride 0.9% Prefilled Flush 10 mL","box",30, "Supply",  "Pharmacy",    "BD PosiFlush"),
    "MED-CUP":    ("Unit Dose Medication Cups 1 oz",        "case", 5000,"Supply",   "Pharmacy",    "Medline"),
    "VIAL-LBL":   ("Pharmacy Vial Labels, Roll",            "roll", 500, "Supply",   "Pharmacy",    "McKesson"),
    "APAP-500":   ("Acetaminophen 500 mg Tablets, 100 ct",  "bottle",100,"Supply",   "Pharmacy",    "McKesson"),
    "IBU-200":    ("Ibuprofen 200 mg Tablets, 100 ct",      "bottle",100,"Supply",   "Pharmacy",    "McKesson"),
    # ----- emergency department -----
    "IV-START":   ("IV Start Kit",                          "case", 100, "Supply",   "Emergency",   "Medline"),
    "CATH-IV20":  ("IV Catheter 20G",                       "box",  50,  "Supply",   "Emergency",   "Becton Dickinson"),
    "CATH-FOLEY": ("Foley Catheter 16 Fr",                  "each", 1,   "Supply",   "Emergency",   "Bard"),
    "ABG-KIT":    ("Arterial Blood Gas Kit",                "box",  25,  "Supply",   "Emergency",   "Becton Dickinson"),
    "SUTURE-40":  ("Nylon Suture 4-0",                      "box",  12,  "Supply",   "Emergency",   "Ethicon"),
    "ECG-ELEC":   ("ECG Electrodes",                        "pack", 100, "Supply",   "Emergency",   "3M"),
    # ----- inpatient / med-surg -----
    "DRESS-FILM": ("Transparent Film Dressing 4 in x 4.75 in","box",100, "Supply",   "Med-Surg",    "3M Tegaderm"),
    "UNDERPAD":   ("Underpad (Chux) 23 in x 36 in",         "pack", 10,  "Supply",   "Med-Surg",    "Medline"),
    "BRIEF-L":    ("Adult Brief, Large",                    "bag",  20,  "Supply",   "Med-Surg",    "Medline"),
    "BATH-WIPE":  ("Bath Wipes, Pre-Moistened, 8 ct",       "pack", 8,   "Supply",   "Med-Surg",    "Medline"),
    "FEED-SET":   ("Enteral Feeding Bag Set",               "each", 1,   "Supply",   "Med-Surg",    "Cardinal Health"),
    "FOAM-DRESS": ("Foam Dressing 4 in x 4 in",             "box",  10,  "Supply",   "Wound Care",  "Medline"),
    # ----- clinic / outpatient -----
    "DEPRESSOR":  ("Tongue Depressors, Non-Sterile",        "box",  500, "Supply",   "Clinic",      "Puritan"),
    "SWAB-COT":   ("Cotton-Tipped Applicators 6 in",        "box",  100, "Supply",   "Clinic",      "Puritan"),
    "EXAM-PAPER": ("Exam Table Paper Roll",                 "case", 12,  "Supply",   "Clinic",      "Medline"),
    "OTO-TIP":    ("Otoscope Specula, Disposable",          "bag",  850, "Supply",   "Clinic",      "Welch Allyn"),
    "BANDAID":    ("Adhesive Bandages 3/4 in x 3 in",       "box",  100, "Supply",   "Clinic",      "Band-Aid"),
    # ----- second batch (added on request) -----
    "SLIDE-MICRO":   ("Microscope Slides, Frosted End",       "box",  72,  "Supply",  "Laboratory",   "Globe Scientific"),
    "PIP-1000":      ("Pipette Tips 1000 uL, Racked",         "case", 960, "Supply",  "Laboratory",   "Rainin"),
    "CONICAL-15":    ("Conical Centrifuge Tube 15 mL",        "case", 500, "Supply",  "Laboratory",   "Corning"),
    "BC-AERO":       ("Blood Culture Bottle, Aerobic",        "box",  50,  "Supply",  "Microbiology", "BD BACTEC"),
    "BC-ANAERO":     ("Blood Culture Bottle, Anaerobic",      "box",  50,  "Supply",  "Microbiology", "BD BACTEC"),
    "MICROTAINER":   ("Capillary Microtainer Tube, K2EDTA",   "box",  50,  "Supply",  "Phlebotomy",   "Becton Dickinson"),
    "GLUCOSE-STRIP": ("Blood Glucose Test Strips, 50 ct",     "box",  1,   "Reagent", "Point of Care","Roche Accu-Chek"),
    "CULTURETTE":    ("Transport Swab, Liquid Amies",         "box",  50,  "Supply",  "Microbiology", "Copan"),
    "QC-HEME":       ("Hematology Control, Tri-Level",        "box",  1,   "Control", "Hematology",   "Sysmex"),
    "N95-RESP":      ("N95 Particulate Respirator",           "box",  20,  "Supply",  "General",      "3M"),
}

# location -> list of (catalog#, on_hand, reorder_point)  [both in COUNT units, e.g. boxes]
# low-stock (on_hand <= reorder) is deliberate where noted; the warehouse always
# carries those items in quantity so a transfer demos cleanly.
PLACEMENT = {
    "San Carlos Warehouse": [
        ("GLV-NIT-S", 48, 12), ("GLV-NIT-M", 60, 15), ("GLV-NIT-L", 44, 12),
        ("ALC-PREP", 40, 10), ("GZ-4X4", 55, 12), ("TAPE-SILK", 36, 10),
        ("SANI-WIPE", 80, 20), ("HAND-SAN", 64, 16), ("SHARPS-1Q", 50, 12),
        ("SPEC-BAG", 30, 8), ("SYR-10ML", 28, 8), ("NDL-21G", 34, 10),
        ("TUBE-RED", 40, 10), ("TUBE-LAV", 42, 10), ("TUBE-BLUE", 30, 8),
        ("TUBE-GRN", 28, 8), ("TOURNIQUET", 26, 8), ("BCS-21G", 24, 6),
        ("URINE-CUP", 600, 150), ("RGT-TOSOH1", 12, 3), ("RGT-TOSOH2", 12, 3),
        ("CELLPACK", 16, 4), ("SULFOLYSER", 14, 4), ("QC-MAS", 10, 3),
        ("STRIP-UA", 22, 6), ("KIT-STREP", 30, 8), ("KIT-FLU", 26, 6),
        ("KIT-COVID", 40, 10), ("KIT-RSV", 22, 6), ("KIT-HCG", 28, 8),
        ("NS-1000", 30, 8), ("D5W-1000", 20, 6), ("FLUSH-NACL", 36, 10),
        ("IV-START", 18, 5), ("CATH-IV20", 26, 8), ("DRESS-FILM", 24, 6),
        ("APAP-500", 30, 8), ("IBU-200", 28, 8), ("EXAM-PAPER", 22, 6),
        ("BANDAID", 40, 10),
        ("SLIDE-MICRO", 30, 8), ("PIP-1000", 24, 6), ("CONICAL-15", 26, 6),
        ("BC-AERO", 28, 8), ("BC-ANAERO", 28, 8), ("MICROTAINER", 34, 10),
        ("GLUCOSE-STRIP", 40, 10), ("CULTURETTE", 30, 8), ("QC-HEME", 12, 3),
        ("N95-RESP", 50, 12),
    ],
    "ED Stockroom": [
        ("GLV-NIT-M", 4, 6),            # LOW
        ("GLV-NIT-L", 7, 4), ("ALC-PREP", 6, 3), ("GZ-4X4", 8, 4),
        ("IV-START", 5, 3), ("CATH-IV20", 2, 4),  # LOW
        ("CATH-FOLEY", 9, 4), ("ABG-KIT", 3, 4),  # LOW
        ("SUTURE-40", 6, 3), ("ECG-ELEC", 7, 4), ("SHARPS-1Q", 6, 3),
        ("FLUSH-NACL", 5, 3), ("NDL-21G", 4, 3),
        ("BC-AERO", 5, 3), ("GLUCOSE-STRIP", 6, 3), ("CULTURETTE", 4, 3), ("N95-RESP", 8, 4),
    ],
    "San Carlos Main Lab": [
        ("TUBE-RED", 9, 6), ("TUBE-LAV", 3, 6),    # LOW
        ("TUBE-BLUE", 7, 4), ("TUBE-GRN", 6, 4), ("TOURNIQUET", 5, 3),
        ("BCS-21G", 4, 3), ("RGT-TOSOH1", 1, 3),   # LOW
        ("RGT-TOSOH2", 4, 3), ("CELLPACK", 3, 3), ("SULFOLYSER", 4, 2),
        ("QC-MAS", 2, 3),                          # LOW
        ("STRIP-UA", 6, 3), ("URINE-CUP", 80, 60), ("GLV-NIT-M", 6, 4),
        ("ALC-PREP", 5, 3), ("KIT-COVID", 8, 4),
        ("SLIDE-MICRO", 6, 3), ("PIP-1000", 5, 2), ("CONICAL-15", 6, 3),
        ("BC-AERO", 7, 4), ("BC-ANAERO", 3, 4),   # LOW
        ("MICROTAINER", 6, 3), ("CULTURETTE", 5, 3), ("QC-HEME", 2, 2),
    ],
    "Clarence Wesley Lab": [
        ("TUBE-RED", 5, 3), ("TUBE-LAV", 2, 3),    # LOW
        ("STRIP-UA", 3, 2), ("KIT-STREP", 1, 3),   # LOW
        ("KIT-FLU", 4, 2), ("KIT-COVID", 5, 3), ("KIT-HCG", 4, 2),
        ("URINE-CUP", 40, 30), ("GLV-NIT-M", 4, 3), ("ALC-PREP", 3, 2),
        ("LANCET", 5, 3), ("TOURNIQUET", 3, 2),
        ("SLIDE-MICRO", 4, 2), ("MICROTAINER", 4, 2),
    ],
    "Pharmacy": [
        ("NS-1000", 4, 6),             # LOW
        ("D5W-1000", 8, 4), ("FLUSH-NACL", 10, 5), ("SYR-3ML", 9, 4),
        ("SYR-10ML", 6, 3), ("MED-CUP", 5, 2), ("VIAL-LBL", 7, 3),
        ("APAP-500", 6, 4), ("IBU-200", 5, 3), ("GLV-NIT-S", 5, 3),
        ("ALC-PREP", 4, 3), ("SHARPS-1Q", 4, 2),
        ("N95-RESP", 5, 3),
    ],
    "Inpatient Unit": [
        ("DRESS-FILM", 2, 4),          # LOW
        ("UNDERPAD", 9, 4), ("BRIEF-L", 7, 3), ("BATH-WIPE", 6, 3),
        ("FEED-SET", 8, 4), ("FOAM-DRESS", 5, 3), ("GLV-NIT-L", 6, 4),
        ("GLV-NIT-M", 5, 4), ("GZ-4X4", 7, 4), ("TAPE-SILK", 5, 3),
        ("FLUSH-NACL", 6, 3), ("SHARPS-1Q", 5, 3),
        ("GLUCOSE-STRIP", 6, 3), ("N95-RESP", 6, 3),
    ],
    "Clinic": [
        ("DEPRESSOR", 6, 3), ("SWAB-COT", 7, 3), ("EXAM-PAPER", 5, 3),
        ("OTO-TIP", 4, 2), ("BANDAID", 8, 4), ("URINE-CUP", 60, 40),
        ("KIT-STREP", 5, 3), ("KIT-FLU", 4, 2),
        ("KIT-COVID", 3, 4),           # LOW
        ("KIT-HCG", 5, 3), ("LANCET", 6, 3), ("GLV-NIT-M", 5, 3),
        ("ALC-PREP", 4, 3),
        ("MICROTAINER", 5, 3), ("GLUCOSE-STRIP", 3, 4),   # GLUCOSE-STRIP LOW
        ("CULTURETTE", 4, 2),
    ],
}

WAREHOUSE_LOC = "San Carlos Warehouse"


def expand(loc):
    """Yield full item dicts for a location."""
    for cat, on_hand_ct, reorder_ct in PLACEMENT[loc]:
        name, order_unit, pack, category, dept, vendor = CATALOG[cat]
        usage_unit = "each" if pack > 1 else order_unit
        count_unit = order_unit
        on_hand_each = on_hand_ct * pack
        reorder_each = reorder_ct * pack
        yield {
            "catalog_number": cat, "item_name": name, "category": category,
            "department": dept, "vendor": vendor,
            "order_unit": order_unit, "usage_unit": usage_unit, "count_unit": count_unit,
            "units_per_order_unit": pack, "units_per_count_unit": pack,
            "on_hand_ct": on_hand_ct, "reorder_ct": reorder_ct,
            "quantity_on_hand": on_hand_each, "reorder_point": reorder_each,
            "low": on_hand_ct <= reorder_ct,
        }


def to_payload(loc):
    out = []
    for r in expand(loc):
        out.append({
            "item_name": r["item_name"], "catalog_number": r["catalog_number"],
            "quantity_on_hand": r["quantity_on_hand"], "reorder_point": r["reorder_point"],
            "category": r["category"], "department": r["department"], "vendor": r["vendor"],
            "unit": r["usage_unit"], "order_unit": r["order_unit"], "usage_unit": r["usage_unit"],
            "count_unit": r["count_unit"], "units_per_order_unit": r["units_per_order_unit"],
            "units_per_count_unit": r["units_per_count_unit"], "status": "active", "notes": TAG,
        })
    return out


def write_xlsx(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    teal = PatternFill("solid", fgColor="01696F")
    amber = PatternFill("solid", fgColor="FFF3E0")
    wb = openpyxl.Workbook()
    # Summary sheet
    ws = wb.active; ws.title = "Summary"
    ws.append(["Location", "Items", "Low-stock", "Warehouse?"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = teal
    for loc in PLACEMENT:
        rows = list(expand(loc))
        ws.append([loc, len(rows), sum(1 for r in rows if r["low"]), "YES" if loc == WAREHOUSE_LOC else ""])
    for col, w in zip("ABCD", [26, 10, 12, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    # Per-location sheets
    for loc in PLACEMENT:
        s = wb.create_sheet(title=loc[:31])
        s.append(["Catalog #", "Item", "Category", "Department", "On Hand", "Unit", "Reorder Pt", "Low?", "Vendor"])
        for c in s[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = teal
        for r in expand(loc):
            s.append([r["catalog_number"], r["item_name"], r["category"], r["department"],
                      r["on_hand_ct"], r["count_unit"], r["reorder_ct"],
                      "LOW" if r["low"] else "", r["vendor"]])
            if r["low"]:
                for c in s[s.max_row]:
                    c.fill = amber
        for col, w in zip("ABCDEFGHI", [12, 42, 11, 14, 9, 8, 10, 7, 18]):
            s.column_dimensions[col].width = w
        s.freeze_panes = "A2"
    wb.save(path)


def main():
    total = sum(len(PLACEMENT[l]) for l in PLACEMENT)
    low = sum(1 for l in PLACEMENT for r in expand(l) if r["low"])
    print(f"Locations: {len(PLACEMENT)}   Total items: {total}   Low-stock: {low}")
    for loc in PLACEMENT:
        rows = list(expand(loc))
        print(f"  {loc:<24} {len(rows):>2} items, {sum(1 for r in rows if r['low'])} low")
    print("Category mix:", dict(Counter(r["category"] for l in PLACEMENT for r in expand(l))))

    xlsx = r"C:\Users\veril\Desktop\Verita Products\SanCarlos_Demo_Inventory_Review.xlsx"
    write_xlsx(xlsx)
    print("Review workbook:", xlsx)

    commit = "--commit" in sys.argv
    provision = "--provision" in sys.argv
    if not (commit or provision):
        return

    base = "https://www.veritaslabservices.com"

    def _arg(flag, default=None):
        for i, a in enumerate(sys.argv):
            if a == flag and i + 1 < len(sys.argv):
                return sys.argv[i + 1]
        return default

    def _post(path, payload):
        req = urllib.request.Request(base + path, data=json.dumps(payload).encode(),
                                     headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode())

    sec = os.environ.get("ADMIN_SECRET", "")
    if not sec:
        print("ADMIN_SECRET not set; cannot commit/provision."); return

    labmap = {}
    created = {}
    if provision:
        # Stand up the 7 demo labs under one owner and wire the warehouse group,
        # all via ADMIN_SECRET (no per-lab JWT). Idempotent: re-running finds the
        # existing labs and only re-seeds ones it freshly created.
        owner = _arg("--owner-email", "verilabguy@gmail.com")
        plan = _arg("--plan", "enterprise")
        print(f"Provisioning 7 demo labs under {owner} (plan {plan})...")
        wh = _post("/api/admin/provision-demo-lab",
                   {"secret": sec, "ownerEmail": owner, "labName": WAREHOUSE_LOC, "plan": plan, "isWarehouse": True})
        wh_id = wh["labId"]; labmap[WAREHOUSE_LOC] = wh_id; created[WAREHOUSE_LOC] = wh.get("created")
        print(f"  {WAREHOUSE_LOC:<24} lab {wh_id}  ({'created' if wh.get('created') else 'exists'}, warehouse)")
        for loc in PLACEMENT:
            if loc == WAREHOUSE_LOC:
                continue
            r = _post("/api/admin/provision-demo-lab",
                      {"secret": sec, "ownerEmail": owner, "labName": loc, "plan": plan, "warehouseLabId": wh_id})
            labmap[loc] = r["labId"]; created[loc] = r.get("created")
            print(f"  {loc:<24} lab {r['labId']}  ({'created' if r.get('created') else 'exists'}, -> warehouse {wh_id})")
        print("labmap:", json.dumps(labmap))
    else:
        # Legacy path: caller supplies --labmap and seeds all 7.
        labmap = json.loads(_arg("--labmap") or "{}")
        if set(labmap) != set(PLACEMENT):
            print("labmap must map all 7 locations to lab ids:", list(PLACEMENT)); return
        created = {loc: True for loc in labmap}

    # Seed inventory. import-inventory is idempotent per lab (it skips catalog
    # numbers already present), so we always seed: a fresh lab gets everything,
    # an already-provisioned lab gets only the new items. No duplicates either
    # way, so adding items here and re-running --provision just tops up.
    for loc in PLACEMENT:
        lab_id = labmap[loc]
        res = _post("/api/admin/import-inventory", {"secret": sec, "labId": lab_id, "items": to_payload(loc)})
        print(f"{loc} (lab {lab_id}): inserted {res.get('inserted')}  skipped {res.get('skipped', 0)}  errors {len(res.get('errors', []))}")


if __name__ == "__main__":
    main()
