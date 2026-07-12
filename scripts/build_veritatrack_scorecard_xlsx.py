#!/usr/bin/env python3
"""Build the VeritaTrack module audit scorecard (internal deliverable for Michael).

4-lens read-only audit (reliability/multi-lab, export/PDF, compliance-copy/data-truth,
UX), each finding verified by the main agent against current code. Main-agent
recalibrations vs the sub-agents:
  - Lens 2 called the Excel About em-dashes HIGH; recalibrated to MED (punctuation
    in an About paragraph, low blast-radius, matches how VeritaLab/VeritaScan scored
    the identical item).
  - Lens 3 called daily/weekly-cadence overstatement HIGH; kept as top-MED
    (public data-truth copy fix, not a functional break).
  - Confirmed HIGH: the /worklist IDOR (missing labScopeMiddleware), the tasks
    error-as-empty, and the sign-off VeritaMap writeback wrong-lab write.
  - CLEAN confirmed: every unscoped route EXCEPT /worklist is lab-membership-safe
    via resolveLegacyLabId / resolveRowForMutation; Excel export is fully compliant
    and lab-scoped; cadences + CFR citations are accurate.

Usage:  python scripts/build_veritatrack_scorecard_xlsx.py
Output: C:/Users/veril/Downloads/VeritaTrack_Scorecard.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\veril\Downloads\VeritaTrack_Scorecard.xlsx"
TEAL, WHITE, ALT, TEXT = "01696F", "FFFFFF", "EBF3F8", "28251D"
GREEN, RED, AMBER, GRAY = "437A22", "A12C7B", "964219", "7A7974"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [("#", 5), ("Finding", 42), ("Area", 15), ("File : Line", 34),
        ("Severity", 10), ("Evidence (verified against code)", 52),
        ("Recommendation", 42), ("Status", 20)]

ROWS = [
    (1, "SECURITY / cross-lab IDOR: the /worklist endpoint has NO lab-membership guard, so any authenticated track-plan user can read ANY lab's compliance calendar plus cross-module data including staff names.",
        "Security/Multi-lab", "server/veritatrack.ts:93-99 (route) + 100-277 (queries)",
        "HIGH",
        "GET /api/labs/:labId/veritatrack/worklist is chained with authMiddleware ONLY and reads Number(req.params.labId) directly. Its siblings /tasks (942) and /dashboard (973) both chain labScopeMiddleware (the only thing that validates an active lab_members row and 403s otherwise); worklist is registered OUTSIDE the `if (labScopeMiddleware)` block and never sets req.scope, so hasTrackAccess(req.user, req.scope?.lab) falls back to the REQUESTER's plan. A user on lab 10 can GET /api/labs/14/veritatrack/worklist and read lab 14's tasks + cert numbers, PT AT-RISK analytes, QC corrective actions, findings, and competency EMPLOYEE NAMES (compMilestonesDue 231-256).",
        "Add labScopeMiddleware to the handler chain and read req.scope.labId (mirror /tasks at 942). Fix first, ahead of everything else.",
        "OPEN - HIGH (security)"),
    (2, "A failed task-list load returns [] and renders 'No tasks yet', so a transient error looks identical to an empty calendar; the page then auto-opens Quick Setup, and re-adding creates true duplicates.",
        "Reliability/UX", "VeritaTrackAppPage.tsx:680-682 (query) + 747-753 (auto-open) + server 382-401 (no dedupe on create)",
        "HIGH",
        "The tasks queryFn does `if (!r.ok) return []`, so a 403/500/network blip resolves to [], isLoading clears, and the empty state 'No tasks yet' (1071) renders while the auto-open effect (749, tasks.length===0) pops Quick Setup. The create endpoint does not dedupe by name, so re-seeding produces duplicate tasks. Same class fixed in VeritaLab/VeritaStaff/VeritaScan this sweep. (Dashboard query 690-692 and History dialog 370-378 share the error-as-empty shape at lower stakes.)",
        "throw on !r.ok, render a distinct error state (retry) separate from empty, and gate the auto-open effect on !isError. Fix the dashboard + history queries in the same pass.",
        "OPEN - HIGH, clean fix ready"),
    (3, "Sign-off's VeritaMap writeback targets the OWNER's home lab, not the task's lab, so a multi-lab owner signing off a map-linked task on Lab B writes the completion date onto Lab A's map and never updates Lab B. (#107 wrong-lab class.)",
        "Multi-lab/Reliability", "server/veritatrack.ts:463 (correct lab) vs 484-501 (uses users.lab_id)",
        "HIGH",
        "signoffLabId = task.lab_id ?? resolveLegacyLabId(...) is computed correctly at 463, but the map lookup 20 lines later does SELECT lab_id FROM users WHERE id=? (owner's legacy/home lab, 485) then veritamap_maps WHERE lab_id = ownerLabId (491) and UPDATE veritamap_tests. users.lab_id can drift from the active lab (labAccessGuard warns of this). Reachable in the normal /labs/B/veritatrack-app flow. Mitigated (wrapped in try/catch; veritamap_tests is the denormalized copy, repairable via resync-complexity), but it is the exact task-#107 wrong-lab-write shape.",
        "Use signoffLabId (or task.lab_id) for the veritamap_maps lookup, not users.lab_id.",
        "OPEN - HIGH"),
    (4, "Public + export copy advertises 'daily, weekly, and custom-cadence' tasks, but the product's minimum cadence is Monthly and the date math is month-only, so sub-monthly cadences are structurally impossible. A prospect can disprove it in one click.",
        "Data-truth/Copy", "VeritaTrackPage.tsx:19,44,47; export server/veritatrack.ts:819,822 (vs FREQUENCIES AppPage 70-76 / frequencyToMonths 30-39 / nextDue setMonth 41-45)",
        "MEDIUM",
        "FREQUENCIES = Monthly/Quarterly/Biannual/Annual/Biennial only; nextDue uses base.setMonth(+months), so there is no Daily/Weekly/custom path. Yet the public page says 'document daily, weekly, and monthly QC' (44) and 'Recurring daily, weekly, monthly, and custom-cadence tasks' (47), and the export About sheet says 'daily, weekly, monthly...checks' (819). False capability claim (the class Michael flags as fabricated).",
        "Either add true Daily/Weekly cadences (day-granular next-due math), or strike 'daily, weekly' + 'custom-cadence' from the public page + About sheet and state the real floor (Monthly and longer). Your call.",
        "OPEN - top MED (data-truth)"),
    (5, "Lab-scoped dashboard returns a mismatched shape, so multi-lab accounts see BLANK 'Due This Month' and 'Due Soon' compliance cards.",
        "Multi-lab/Data-truth", "server/veritatrack.ts:991 (returns due_soon snake_case, no dueThisMonth) vs VeritaTrackAppPage.tsx:1017-1022 (renders dashboard.dueThisMonth / dashboard.dueSoon)",
        "MEDIUM",
        "The lab-scoped dashboard endpoint returns { overdue, due_soon: dueSoon, current, not_started, total, tasks } with no dueThisMonth and snake_case due_soon. The client Dashboard cards read dashboard.dueThisMonth and dashboard.dueSoon (camelCase), so on the multi-lab path (activeLabId set) both cards render undefined (blank) while Overdue/Current show numbers. The legacy unscoped endpoint (643-667) returns the correct camelCase shape, so single-lab users are unaffected.",
        "Make the lab-scoped endpoint return the same keys as the legacy one (dueThisMonth, dueSoon, notStarted, *Items).",
        "OPEN - clean fix ready"),
    (6, "'Due today' tasks are flagged as '1 day overdue': the status math compares a UTC-midnight due date against a live timestamp, so in any US timezone a task reads overdue on its actual due date. The 'Due Today' tile is effectively always 0 and inflates Overdue.",
        "Data-truth/Reliability", "server/veritatrack.ts:47-54 (taskStatus), 118-124 (worklist), 660-662 (dashboard); client 157-163",
        "MEDIUM",
        "taskStatus: due = new Date('YYYY-MM-DD') (UTC midnight), now = new Date() (live). daysUntil = floor((due-now)/86400000). For a negative-UTC-offset (US) timezone, on the due date now is already past UTC midnight, so daysUntil floors to -1 and the task returns 'overdue'. In the worklist, the daysUntil<0 overdue push (121) fires before the due-today check (122), so the Due Today tile is ~always 0. Same math in dashboard, export, and the client daysLabel.",
        "Compare date-only (floor now to local midnight, or compare YYYY-MM-DD strings) so due-today is daysUntil === 0.",
        "OPEN - clean fix ready"),
    (7, "Two contradictory 'Overdue' numbers on the same screen: the Worklist tile folds never-signed-off tasks into Overdue while the Dashboard 'Overdue' card counts them separately, so right after Quick Setup the two read e.g. '20 Overdue' vs '0'.",
        "Data-truth", "server/veritatrack.ts:114-116 (worklist push) + 332 (count) vs 988/658 (dashboard) ; client 895-896 vs 1015-1028",
        "MEDIUM",
        "The worklist buckets never-signed-off tasks into buckets.overdue with status 'not_started' (114-116) and counts them as counts.overdue (332); the Worklist tile renders that. The Dashboard cards come from the dashboard endpoint, which counts not_started separately, so its Overdue card excludes them. After seeding ~20 default tasks (none signed off), the Worklist tile shows '20 Overdue' and the Dashboard Overdue card shows 0 on the same page. Also 'Overdue' is the wrong word for a task that has never come due.",
        "Give not_started its own worklist bucket/label ('Not started'), and reconcile the two views to one Overdue definition.",
        "OPEN"),
    (8, "'Auto-imports schedules from VeritaMap so adding a new instrument creates its cadence automatically' overstates a manual, on-demand Import button. The same false line is copy-pasted onto 3 other surfaces.",
        "Data-truth/Copy", "VeritaTrackAppPage.tsx:843,846 (+ class: DemoLabPage.tsx:1357,1364; VeritaAssurePage.tsx:107; scripts/inject_howto_cards.py:87)",
        "MEDIUM",
        "Import fires only from the 'Import from VeritaMap' button -> handleImport (779-798) -> POST /import-from-map. Grep confirms no VeritaMap-side hook creates a task on instrument add. Nothing is automatic; adding an instrument does nothing until the user clicks Import. Reword to reflect the one-click manual action, and fix all 4 copies in the same pass (fix-the-class).",
        "Reword to 'One-click import of your VeritaMap test menu; re-run after adding instruments.' Fix DemoLab + VeritaAssure + the generator script too.",
        "OPEN - clean fix (class of 4)"),
    (9, "Legacy dashboard endpoint still keys by user_id, not lab_id (missed in the sweep that migrated /tasks), so a multi-lab owner on the bare route gets summary cards that aggregate across all their labs.",
        "Multi-lab", "server/veritatrack.ts:643-648",
        "MEDIUM",
        "The /tasks list was migrated off user_id to lab_id (its own comment 342-346 explains the bleed it fixed), but this legacy dashboard endpoint still does SELECT ... WHERE user_id = ?. Reached when activeLabId is null (bare /veritatrack-app). On that route /tasks uses WHERE lab_id while the dashboard uses WHERE user_id, so the cards count across every lab owned by that user_id and disagree with the visible list.",
        "Scope the legacy dashboard by resolveLegacyLabId like the /tasks read, or retire it in favor of the lab-scoped endpoint (973).",
        "OPEN"),
    (10, "The Calendar view silently drops every 'Not Started' task (its next_due is null), so filtering to Not Started + Calendar shows 'nothing due' while the header count still says N.",
        "UX/Data-truth", "VeritaTrackAppPage.tsx:596 (if !t.next_due continue) + 1063 (filter option) + 1067 (count) + 1088 (calendar gets filtered)",
        "MEDIUM",
        "The calendar builder does `if (!t.next_due) continue`; never-signed-off tasks carry next_due null (server 110/115/358), so they are omitted from every month while the task-count header still reflects them. Not-started tasks are exactly the ones a director most needs to see.",
        "Surface not-started tasks in the calendar (a 'no baseline yet' bucket) or disable/annotate the Not Started filter in calendar view; reconcile the count.",
        "OPEN"),
    (11, "Sign-off Save is enabled with blank initials AND blank performer, producing an anonymous, unattributable completion in a surveyor-facing audit trail, contradicting the module's own 'audit trail for who attests' copy.",
        "Data quality/UX", "VeritaTrackAppPage.tsx:226-237 (Save disabled only on !date) ; server 455-472 (only completed_date required)",
        "MEDIUM",
        "Save is disabled={!date || mut.isPending}; the Initials field has no asterisk while 'Completion Date *' does. Server checks only `if (!completed_date)`. A completion can be recorded with no initials and no performed_by, and the audit detail degrades to 'Completed <date>' with no name, contradicting the export copy that calls Performed By 'the audit trail for who attests the task was done' (server 822).",
        "Require at least one of initials/performer (disable Save until one is filled) and validate server-side.",
        "OPEN"),
    (12, "Em-dashes in three customer-facing Excel About paragraphs.",
        "Export/Copy", "server/veritatrack.ts:819, 825, 831 (aboutBody strings; 831 has two)",
        "MEDIUM",
        "Three aboutBody(...) strings contain \\u2014 that render as literal em-dashes in the exported .xlsx (e.g. 825 '...in VeritaTrack \\u2014 it does not validate...'). CLAUDE.md Sec 3 + Sec 6 item 6 (NON-NEGOTIABLE, audit-script-enforced) ban em-dashes in every cell/header/footer/About paragraph. The other \\u2014 at 365/431/445/509 are code comments (allowed). Same class as the VeritaLab/VeritaScan About em-dash fixes.",
        "Replace each \\u2014 with a period, semicolon, or comma.",
        "OPEN - clean fix ready"),
    (13, "Banned 'Cal Ver' / 'cal ver' abbreviation and 'method comparison' alone in customer-facing UI copy, inconsistent with the module's own correctly-named seeded tasks.",
        "Copy", "VeritaTrackAppPage.tsx:308 (placeholder 'e.g. Cal Ver - Sodium'), 843 ('cal-ver cadence'), 864 (tooltip 'cal ver, method comparison')",
        "MEDIUM",
        "CLAUDE.md Sec 5: 'Calibration Verification / Linearity, never Cal Ver'; 'Correlation / Method Comparison, never Method Comparison alone.' The banned strings appear in a form placeholder, a how-to line, and the import-button tooltip, even though the module's own seeded task names use the correct full forms ('Calibration Verification' 609, 'Correlation / Method Comparison' 610), so the copy is internally inconsistent.",
        "Use the full labels: 'e.g. Calibration Verification - Sodium'; 'calibration-verification cadence'; tooltip 'calibration verification, correlation / method comparison, precision, and SOP schedules.'",
        "OPEN - clean fix ready"),
    (14, "deleteTask never checks r.ok, so a failed delete (403 seat/plan gate, 404) still runs onSuccess and reports success; the task reappears on refetch with no error shown.",
        "Reliability/UX", "VeritaTrackAppPage.tsx:422-431",
        "MEDIUM",
        "mutationFn does `await fetch(... DELETE ...)` with no r.ok check; onSuccess invalidates the queries regardless. A 403/404 resolves without throwing, so the UI reports success (false success). Server DELETE is a soft-delete + audit event, so no data loss, but the user is misinformed.",
        "throw on !r.ok and add onError surfacing. (Server-side the DELETE is correctly soft + audited.)",
        "OPEN - clean fix ready"),
    (15, "Export Excel / Import / Seed buttons swallow failures: no loading/disabled on Export, no error surface on any, and the module has no toast system, so failures are invisible and heavy builds can be double-fired.",
        "UX", "VeritaTrackAppPage.tsx:800-811 (export), 779-798 (import), 755-777 (seed)",
        "LOW-MED",
        "handleExcelExport has no exporting state and `if(!r.ok) return` (silent); the server build is a synchronous 2-sheet workbook (seconds), so the button looks inert and can be re-clicked into multiple builds. handleImport/handleSeedDefaults act only on r.ok with no else and no catch (a non-JSON 502 throws unhandled), discarding the server's actionable 'No VeritaMap found, build your test menu first' guidance (server 576). No toast system is imported anywhere in the module.",
        "Add an exporting/disabled state; surface server error messages (toast or inline banner) on non-2xx; read data.error before assuming success.",
        "OPEN"),
    (16, "'Due today' shows as overdue is compounded by nextDue's setMonth month-end overflow: a monthly task last done Jan 31 lands on Mar 2/3.",
        "Reliability", "server/veritatrack.ts:41-45",
        "LOW",
        "nextDue does base.setMonth(base.getMonth()+frequencyMonths); Jan 31 + 1mo overflows Feb into Mar 2/3, so month-end tasks drift later each cycle. Also mixes UTC-parsed dates with local getMonth/setMonth (latent off-by-one for positive-offset zones; US operators are negative-offset, low impact).",
        "Clamp to end-of-month or add months on a normalized UTC date.",
        "OPEN"),
    (17, "Import creates the task and its seeded sign-off in two un-transactioned inserts; a throw between them leaves a task with no imported date that the dedupe then skips forever ('Not Started' despite a real map date).",
        "Reliability", "server/veritatrack.ts:622-637",
        "LOW",
        "The per-analyte loop does INSERT veritatrack_tasks then a separate INSERT veritatrack_signoffs with no transaction. If the sign-off insert throws after the task insert, the task exists without its date; on re-run the WHERE lab_id=? AND name=? dedupe skips it and the date is never applied. Low probability (better-sqlite3 synchronous, additive-only, no destructive delete-then-insert).",
        "Wrap each task+signoff pair (or the whole import) in a db.transaction().",
        "OPEN"),
    (18, "Legacy POST /tasks sets lab_id in a second, swallowed UPDATE; if it throws the task persists with lab_id NULL and is invisible to every lab-scoped read.",
        "Reliability/Multi-lab", "server/veritatrack.ts:389-398",
        "LOW",
        "The legacy create does INSERT (...user_id...) then try { UPDATE veritatrack_tasks SET lab_id=? WHERE id=? } catch {}. A swallowed UPDATE leaves lab_id NULL, and every lab-scoped read filters WHERE lab_id=?, so the task vanishes. Only reachable on the bare /veritatrack-app fallback (the scoped POST at 958 sets lab_id in the INSERT).",
        "Set lab_id in the initial INSERT (as the scoped POST already does).",
        "OPEN"),
    (19, "Audit trail (append-only sign-off history) is hidden behind a hover-only 10px ghost icon, so a director doing a first walk-through likely never finds the surveyor-critical history.",
        "UX", "VeritaTrackAppPage.tsx:474 (opacity-0 group-hover) + 476 (History trigger size=10)",
        "LOW",
        "The row actions are opacity-0 group-hover:opacity-100 and the History trigger is a size=10 icon with only a title tooltip. The audit view is where deleted sign-offs are the only surviving record, yet it is discoverable only by hovering a dense row and spotting a tiny icon.",
        "Give the audit trail a labeled, always-visible affordance (a 'History' text link or persistent icon).",
        "OPEN"),
    (20, "Public-page overstatements / stale labels: 'Never miss a due date' (no notification system exists), 'Excel matches your existing regulatory calendar format' (fixed template), and stale pricing ($499/yr, 'Enterprise' tier).",
        "Copy/Data-truth", "VeritaTrackPage.tsx:837 (never-miss), VeritaTrackAppPage.tsx:848 (matches-your-format), VeritaTrackPage.tsx:58/63 (pricing)",
        "LOW",
        "There is no reminder/email mechanism in the module (due dates are visible only when the page is open), so 'Never miss a due date' overpromises. The export is a fixed teal VeritaAssure template, not adaptive to the lab's file. Public pricing shows 'From $499/yr' + 'Enterprise' vs the current Clinic $999 / Community $2,125 / Hospital $4,995 / System tiers (Sec 10); 'Enterprise' is a grandfathered-only name. Pricing display is your call.",
        "Soften 'Never miss...' (or add reminders), reword the Excel-format claim, and confirm the intended price anchor + rename Enterprise -> System.",
        "OPEN - your ruling on pricing"),
]

PASS_NOTES = [
    "Multi-lab is SAFE on every unscoped route EXCEPT /worklist: /tasks, /tasks/:id, PUT/DELETE, /signoff, /signoffs/:id, /audit, import-from-map, seed-defaults, and export/excel all resolve the lab through resolveLegacyLabId / resolveRowForMutation, which membership-validate before honoring X-Active-Lab-Id and ignore a forged foreign labId header. The sign-off RECORD lands on the correct lab (task.lab_id); only its VeritaMap side-writeback is mis-scoped (#3).",
    "Excel export is one of the most compliant in the suite: ExcelJS only; teal 01696F headers; freeze D2; auto-filter; alt rows; thin D0D0D0 borders; exact status colors; null cells blank; About sheet is sheet 1 and opens active; 3-layer lab identity from the LIVE labs row; sheet protection with the env password; author 'Perplexity Computer'; correctly lab-scoped (X-Active-Lab-Id chain verified end-to-end). No cross-lab bleed. No PDF export exists (so the PDF footer/signature rules are N/A).",
    "Default cadences + CFR citations are ACCURATE: Calibration Verification 6mo (42 CFR 493.1255(b)(3)), Correlation/Method Comparison 6mo (493.1281(a)), SOP biennial, Bloodborne training annual (OSHA 1910.1030), QC monthly / PT quarterly (defensible lab-defined); the only surfaced CFR (PT AT-RISK 493.803) is correct; internal cites (493.1282, 493.1235) correct. No fabricated or wrong CFR; no specialty->CFR map so the 493.927 trap is N/A.",
    "Copy hygiene is otherwise clean: TM not (R) throughout (zero registered marks); no EP Evaluator / CAMLAB / LabVine / Mayo / TODO / 'coming soon'; no dated accreditor manual references; 'former TJC surveyor' (not CMS); 'medical director or designee' correct. The em-dash + Cal-Ver items (#12/#13) are the only copy-rule breaches.",
    "Delete Task is the only destructive UI action and it is correctly gated by ConfirmDialog (and the server DELETE is a soft-delete + audit event). No write reports FALSE success except deleteTask (#14, missing r.ok). trackAudit is correctly append-only and non-throwing so an audit-write failure can't break a sign-off. Filters are declarative React state (no window.location desync).",
]

SEV_COLOR = {"HIGH": RED, "MED-HIGH": RED, "MEDIUM": AMBER, "LOW-MED": AMBER, "LOW": GRAY}

wb = Workbook(); ws = wb.active; ws.title = "VeritaTrack Scorecard"
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "VeritaTrack - Module Audit Scorecard   (4-lens audit, main-agent verified, 2026-07-11)"
t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=TEAL)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[1].height = 26

hdr = 2
for i, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=hdr, column=i, value=name)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border; ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[hdr].height = 20

r = hdr + 1
for idx, row in enumerate(ROWS):
    fill = WHITE if idx % 2 == 0 else ALT
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=fill); c.border = border
    sc = ws.cell(row=r, column=5)
    sc.font = Font(name="Calibri", size=10, bold=True, color=SEV_COLOR.get(row[4], TEXT))
    r += 1

for note in PASS_NOTES:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    pc = ws.cell(row=r, column=1, value="PASS  |  " + note)
    pc.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    pc.fill = PatternFill("solid", fgColor="E6F2F2")
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 58
    r += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{hdr+len(ROWS)}"
ws.sheet_view.showGridLines = False
wb.save(OUT)
print(f"wrote {OUT} with {len(ROWS)} finding rows + {len(PASS_NOTES)} pass-notes")
for sev in ("HIGH", "MEDIUM", "LOW-MED", "LOW"):
    print(f"  {sev}: {sum(1 for x in ROWS if x[4]==sev)}")
